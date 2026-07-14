"""Orchestration / build order (METHODOLOGY.md section 8).

Wires the per-company flow end to end: load inputs -> NAV -> dividend strip ->
cycle weighting -> blend -> breakeven -> sensitivity -> report, then the
watchlist roll-up across all names.

Build order (section 8): DHT first (single-class validator), then FRO, ECO,
INSW.
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

from dataclasses import replace

from .blend import blend_fair_value
from .breakeven import implied_breakeven_tce
from .carveout import (
    crude_carve_out,
    lng_carve_out,
    product_carve_out,
    product_read,
    sector_carve_out,
)
from .cycle import compute_cycle
from .dividend_strip import compute_dividend_strip
from .loaders import INPUTS_DIR, load_company_inputs, load_watchlist
from .marks import (TXN_PURE_PLAY_K_BAND, scale_vessel_marks,
                    solve_broker_premium)
from .nav import compute_nav
from .report import OUTPUTS_DIR, CompanyReport, write_company_report, write_watchlist_summary
from .scenarios import (
    PRODUCT_SCENARIO_CLASS_MAP,
    SCENARIO_CLASS_MAP_BY_SECTOR,
    ScenarioFV,
    ScenarioReport,
    load_scenarios,
    position_recommendation,
    run_scenarios,
    write_scenario_report,
    write_scenario_summary,
)
from .sensitivity import compute_sensitivity, payout_sensitivity
from .transactions import (
    apply_transaction_anchored_curves,
    load_all_transactions,
)
from .validate import validate_inputs

V1_TICKERS = ("DHT", "FRO", "ECO", "INSW")

# Hybrid crude+product operators valued via crude+product carve-outs and
# aggregated to whole-company FV (METHODOLOGY 6 v2).
HYBRID_TICKERS = {"INSW"}

# Three-sleeve hybrids (crude + product + lng). TEN onboarded 2026-06-06 —
# METHODOLOGY §11.6 (off-curve-at-contracted-book convention for the DP2
# shuttle sleeve, which sits in `shuttle_contracted_book` rather than the
# carve-out).
THREE_SLEEVE_TICKERS: set[str] = {"TEN"}

# Multi-sleeve hybrids over an ARBITRARY ordered set of sectors (not the
# crude/product/lng combos baked into HYBRID_TICKERS / THREE_SLEEVE_TICKERS).
# CMBT (CMB.TECH, ex-Euronav) onboarded 2026-06-26 — first crude + dry_bulk +
# containerships hybrid (METHODOLOGY §11.9). Chemical / offshore (Windcat) /
# FSO / held-for-sale / the multi-segment newbuild book sit OFF-CURVE at the
# corporate level (working_capital_net / shuttle_contracted_book / newbuild
# lines), pro-rated across the on-curve sleeves like the rest of the stack.
MULTI_SLEEVE_TICKERS: dict[str, list[str]] = {
    "CMBT": ["crude", "dry_bulk", "containerships"],
}


def _load_all_sectors(inputs_dir: Path = INPUTS_DIR) -> dict[str, dict]:
    """Pre-load all sector scenario sub-docs once per pipeline run.

    Currently {crude, lng, product, dry_bulk, containerships, lpg}. Pipeline
    picks the right doc per ticker (and per sleeve, for hybrid INSW carve-outs:
    crude sleeve runs through sectors.crude, product sleeve through
    sectors.product). lpg added 2026-07-08 (WO3 Phase 1); no names route
    through it until the Phase-4 validators (Dorian LPG / BW LPG) onboard.
    """
    from .scenarios import SCENARIOS_PATH
    path = inputs_dir / "scenario_inputs.yaml" if inputs_dir != INPUTS_DIR else SCENARIOS_PATH
    return {
        "crude": load_scenarios(path, "crude"),
        "lng": load_scenarios(path, "lng"),
        "product": load_scenarios(path, "product"),
        "dry_bulk": load_scenarios(path, "dry_bulk"),
        "containerships": load_scenarios(path, "containerships"),
        "lpg": load_scenarios(path, "lpg"),
    }


def _resolve_sector(ticker: str, watchlist: dict) -> str:
    """The scenario sector for a ticker; defaults to 'crude'."""
    entry = watchlist.get(ticker) or {}
    return entry.get("sector") or "crude"


def _whole_company_basis(crude_share: float) -> str:
    """Banner text for a hybrid name's whole-company aggregation (METHODOLOGY 6 v2)."""
    return (
        f"WHOLE-COMPANY = crude sleeve ({crude_share:.1%} of vessel value) + product "
        f"sleeve ({1 - crude_share:.1%}) AGGREGATED. Compared to the WHOLE-COMPANY tape "
        f"price (not the crude-allocated proxy). Each sleeve is probability-weighted by "
        f"its OWN sector's scenario weights (cross-sector independence; METHODOLOGY 6 v2, "
        f"rank-1 pairing removed 2026-07-02)."
    )


def _aggregate_hybrid_report(
    crude_report: ScenarioReport,
    product_report: ScenarioReport,
    *, ticker: str, whole_price: float, whole_target: float,
    crude_share: float, product_share: float,
) -> ScenarioReport:
    """Sum the two sleeve ScenarioReports into a single whole-company report.

    Delegates to ``_aggregate_multi_sleeve_report`` (2026-07-02, review C-3):
    each sleeve's ladder is weighted by ITS OWN sector's probabilities. The
    prior index-pairing applied crude weights to the product sleeve — correct
    only while the two families carried identical weights, which the Jun-9
    crude reweight silently broke.
    """
    return _aggregate_multi_sleeve_report(
        [(crude_report, crude_share), (product_report, product_share)],
        ticker=ticker, whole_price=whole_price, whole_target=whole_target,
        sectors=["crude", "product"],
        basis=_whole_company_basis(crude_share),
    )


def _aggregate_three_sleeve_report(
    crude_report: ScenarioReport,
    product_report: ScenarioReport,
    lng_report: ScenarioReport,
    *, ticker: str, whole_price: float, whole_target: float,
    crude_share: float, product_share: float, lng_share: float,
) -> ScenarioReport:
    """3-sleeve analog of ``_aggregate_hybrid_report`` (METHODOLOGY §11.6).

    Delegates to ``_aggregate_multi_sleeve_report`` (2026-07-02, review C-3):
    per-sleeve sector weights, rank-1 index pairing removed.
    """
    basis = (
        f"WHOLE-COMPANY 3-SLEEVE = crude ({crude_share:.1%}) + product "
        f"({product_share:.1%}) + lng ({lng_share:.1%}) AGGREGATED (METHODOLOGY "
        f"§11.6). Off-curve shuttle-contracted-book sleeve sits at the corporate "
        f"level (`shuttle_contracted_book`) and flows through NAV uniformly "
        f"across scenarios. Compared to the WHOLE-COMPANY tape price."
    )
    return _aggregate_multi_sleeve_report(
        [(crude_report, crude_share), (product_report, product_share),
         (lng_report, lng_share)],
        ticker=ticker, whole_price=whole_price, whole_target=whole_target,
        sectors=["crude", "product", "lng"],
        basis=basis,
    )


def _class_map_for_sector(sector: str):
    """Scenario class map for a sleeve sector (None ⇒ the crude/lng default map)."""
    if sector == "product":
        return PRODUCT_SCENARIO_CLASS_MAP
    if sector in ("dry_bulk", "containerships", "lpg"):
        return SCENARIO_CLASS_MAP_BY_SECTOR[sector]
    return None


def _multi_sleeve_basis(sectors: list[str], shares: list[float]) -> str:
    """Banner for an N-sleeve aggregate. MUST start with 'WHOLE-COMPANY' so the
    scenario renderer / roll-up tag it [WHOLE-CO] (scenarios.py)."""
    parts = " + ".join(f"{sec} ({sh:.1%})" for sec, sh in zip(sectors, shares))
    return (
        f"WHOLE-COMPANY MULTI-SLEEVE = {parts} AGGREGATED (METHODOLOGY §11.9). "
        f"Off-curve segments (chemical / offshore / FSO / held-for-sale / newbuild "
        f"book) sit at the corporate level and flow through NAV uniformly across "
        f"sleeves. Compared to the WHOLE-COMPANY tape price."
    )


def _aggregate_multi_sleeve_report(
    sleeves: list[tuple[ScenarioReport, float]],
    *, ticker: str, whole_price: float, whole_target: float, sectors: list[str],
    basis: str | None = None,
) -> ScenarioReport:
    """N-sleeve aggregation core (METHODOLOGY §11.9) — ALL hybrid aggregators
    delegate here (2-sleeve INSW, 3-sleeve TEN, N-sleeve CMBT).

    **Cross-sector independence assumed; rank-1 pairing removed (2026-07-02,
    review C-3).** The headline probability-weighted FV is the SUM of each
    sleeve's OWN probability-weighted FV — each sleeve's ladder weighted by its
    own sector's probabilities. The prior form paired ladders by index and
    applied the FIRST sleeve's weights to every sleeve, which was only correct
    while the sector families carried identical weights at each index; the
    Jun-9 crude reweight broke that silently (CMBT's dry-bulk sleeve — 72.7 %
    of vessel value — was being probability-weighted by the Hormuz state).
    Per-share sleeve FVs add directly (carve-outs pro-rate balance-sheet items,
    never shares).

    The per-scenario DISPLAY rows remain index-paired (lead sleeve's
    name/weight/cycle shown) as an illustrative correlated slice — the footer
    PW FV is deliberately NOT the weight-dot-product of those rows.
    upside_best / downside_worst are independence envelopes (sum of per-sleeve
    bests / worsts). Per-sleeve strip horizons are not the aggregator's
    concern: each sleeve's ``run_scenarios`` built its own strip.
    """
    reports = [r for r, _ in sleeves]
    shares = [s for _, s in sleeves]
    n = min(len(r.scenarios) for r in reports)
    agg: list[ScenarioFV] = []
    for i in range(n):
        cells = [r.scenarios[i] for r in reports]
        lead = cells[0]
        agg.append(ScenarioFV(
            name=lead.name,
            weight=lead.weight,
            fair_value=sum(c.fair_value for c in cells),
            fair_value_low=sum(c.fair_value_low for c in cells),
            fair_value_high=sum(c.fair_value_high for c in cells),
            nav_per_share=sum(c.nav_per_share for c in cells),
            vessel_scale=lead.vessel_scale,
            divstrip_npv=sum(c.divstrip_npv for c in cells),
            cycle_position=lead.cycle_position,
            w_nav=lead.w_nav,
            assumed_tce=sum(c.assumed_tce * sh for c, sh in zip(cells, shares)),
        ))
    pw_fv = sum(r.probability_weighted_fv for r in reports)
    ev = pw_fv - whole_price
    return ScenarioReport(
        ticker=ticker,
        current_price=whole_price,
        analyst_target=whole_target,
        base_nav_per_share=sum(r.base_nav_per_share for r in reports),
        breakeven_tce=reports[0].breakeven_tce,
        scenarios=agg,
        probability_weighted_fv=pw_fv,
        upside_best=sum(max(s.fair_value for s in r.scenarios) for r in reports) - whole_price,
        downside_worst=sum(min(s.fair_value for s in r.scenarios) for r in reports) - whole_price,
        expected_value_vs_current=ev,
        position_recommendation=position_recommendation(ev / whole_price * 100.0),
        basis=(basis or _multi_sleeve_basis(sectors, shares)),
        sector=reports[0].sector,
        # WO1 V-1: per-sleeve PW-FV contributions (per-share; sum == pw_fv by
        # the C-3 per-sleeve identity) — the governance repo's sleeve watch.
        sleeve_fvs={sec: r.probability_weighted_fv for sec, r in zip(sectors, reports)},
    )


def _run_scenarios_for_ticker(
    ticker: str, whole_inputs, whole_price: float, whole_target: float,
    sector_docs: dict[str, dict], watchlist: dict,
    asof_quarter: str | None = None,
) -> tuple[ScenarioReport, ScenarioReport | None, ScenarioReport | None]:
    """Headline ScenarioReport per ticker.

    Returns ``(headline, crude_sleeve_or_None, product_sleeve_or_None)``. For
    pure-plays the headline is the only thing — sleeve reports are None. For
    2-sleeve hybrids (INSW) the headline aggregates crude+product. For
    3-sleeve hybrids (TEN; ``THREE_SLEEVE_TICKERS``) the headline aggregates
    crude+product+lng; the LNG sleeve report is consumed internally and the
    public return surface stays 3-tuple for back-compat. TEN-specific
    per-sleeve breakdown reports can route through a dedicated accessor at
    onboarding time if needed.
    """
    sector = _resolve_sector(ticker, watchlist)
    doc = sector_docs[sector]

    sleeve_sectors = MULTI_SLEEVE_TICKERS.get(ticker)
    if sleeve_sectors is not None:
        # Generic N-sleeve hybrid (CMBT): carve each sleeve sector out of the
        # whole-company fleet, run its scenarios through its own sector doc +
        # class map, and aggregate to a whole-company headline vs the tape
        # (METHODOLOGY §11.9). Returns (headline, crude_sleeve_or_None, None) to
        # keep the 3-tuple contract; the LNG slot is unused here. Per-sleeve
        # breakdown table is skipped (matches the TEN 3-sleeve precedent).
        sleeves = []
        for sec in sleeve_sectors:
            carve = sector_carve_out(whole_inputs, sec)
            cmap = _class_map_for_sector(sec)
            kwargs = {"scenario_class_map": cmap} if cmap is not None else {}
            r = run_scenarios(
                carve.sleeve_inputs, carve.carved_price(whole_price),
                carve.carved_price(whole_target), sector_docs[sec],
                asof_quarter=asof_quarter, **kwargs,
            )
            sleeves.append((r, carve.sleeve_share, sec))
        headline = _aggregate_multi_sleeve_report(
            [(r, s) for r, s, _ in sleeves],
            ticker=ticker, whole_price=whole_price, whole_target=whole_target,
            sectors=[sec for _, _, sec in sleeves],
        )
        crude_r = next((r for r, _, sec in sleeves if sec == "crude"), None)
        return headline, crude_r, None

    if ticker not in HYBRID_TICKERS and ticker not in THREE_SLEEVE_TICKERS:
        # Pure-play routing: pure-product names (sector=product) need the
        # product class map so MR/LR1/LR2 route to clean-trade rate keys
        # (mr / lr1_clean / lr2_clean) rather than the default crude+lng map
        # (which lacks MR and would map LR1->aframax_dirty). For crude/lng
        # pure-plays the default map is correct and we pass it implicitly.
        # Sectors whose vessel classes are NOT in the module-level default
        # SCENARIO_CLASS_MAP need the explicit per-sector map passed through.
        # product: MR/LR1_clean/LR2_clean route via PRODUCT_SCENARIO_CLASS_MAP.
        # dry_bulk: Cape/Pana/Supra-Ultra route via the dry_bulk sector map.
        if sector == "product":
            kwargs = {"scenario_class_map": PRODUCT_SCENARIO_CLASS_MAP}
        elif sector in ("dry_bulk", "containerships", "lpg"):
            from .scenarios import SCENARIO_CLASS_MAP_BY_SECTOR
            kwargs = {"scenario_class_map": SCENARIO_CLASS_MAP_BY_SECTOR[sector]}
        else:
            kwargs = {}
        return run_scenarios(whole_inputs, whole_price, whole_target, doc,
                             asof_quarter=asof_quarter, **kwargs), None, None

    crude_doc = sector_docs["crude"]
    product_doc = sector_docs["product"]
    co = crude_carve_out(whole_inputs)
    po = product_carve_out(whole_inputs)
    crude_r = run_scenarios(
        co.crude_inputs, co.carved_price(whole_price), co.carved_price(whole_target), crude_doc,
        asof_quarter=asof_quarter,
    )
    product_r = run_scenarios(
        po.product_inputs, po.carved_price(whole_price), po.carved_price(whole_target), product_doc,
        scenario_class_map=PRODUCT_SCENARIO_CLASS_MAP, asof_quarter=asof_quarter,
    )

    if ticker in THREE_SLEEVE_TICKERS:
        # 3-sleeve hybrid (TEN): add the LNG sleeve and aggregate all three.
        lng_doc = sector_docs["lng"]
        lo = lng_carve_out(whole_inputs)
        lng_r = run_scenarios(
            lo.lng_inputs, lo.carved_price(whole_price), lo.carved_price(whole_target), lng_doc,
            asof_quarter=asof_quarter,
        )
        headline = _aggregate_three_sleeve_report(
            crude_r, product_r, lng_r, ticker=ticker, whole_price=whole_price, whole_target=whole_target,
            crude_share=co.crude_share, product_share=po.product_share, lng_share=lo.lng_share,
        )
        return headline, crude_r, product_r

    # 2-sleeve hybrid INSW: crude sleeve runs through sectors.crude, product
    # sleeve through sectors.product (METHODOLOGY §6 v2, §11.5). The product
    # sleeve was previously a v2 shortcut using MR forwards embedded in
    # sectors.crude; 2026-06-01 formalised that into its own sector.
    headline = _aggregate_hybrid_report(
        crude_r, product_r, ticker=ticker, whole_price=whole_price, whole_target=whole_target,
        crude_share=co.crude_share, product_share=po.product_share,
    )
    return headline, crude_r, product_r


def _crude_sleeve_basis(crude_share: float, whole_price: float, alloc_price: float) -> str:
    """Standard banner text for a hybrid name's crude sleeve (used everywhere)."""
    return (
        f"CRUDE SLEEVE only ({crude_share:.1%} of vessel value). FV and price are the "
        f"CRUDE sleeve / CRUDE-ALLOCATED price ${alloc_price:,.2f} (= whole-company "
        f"${whole_price:,.2f} × crude_share). Product sleeve (~{1 - crude_share:.0%}) "
        f"is EXCLUDED from the model FV — covered qualitatively only (v2 product strip "
        f"pending). Do not compare directly to whole-company P/NAV without re-aggregating."
    )


def _maybe_apply_transactions(ci, inputs_dir: Path, use_transaction_anchored: bool):
    """Apply transaction-anchored mid-age anchors if toggled on (METHODOLOGY 9.9).

    Returns (possibly-recalibrated CompanyInputs, per-class CurveFit dict). Called
    BEFORE the carve-out so that crude_share is computed on the recalibrated marks.
    """
    if not use_transaction_anchored:
        return ci, {}
    txs = load_all_transactions(inputs_dir)
    if not txs:
        return ci, {}
    new_md, fits = apply_transaction_anchored_curves(ci.market_data, txs)
    return replace(ci, market_data=new_md), fits


def value_company(
    ticker: str,
    quarter: str,
    current_price: float,
    analyst_target: float,
    inputs_dir: Path = INPUTS_DIR,
    use_transaction_anchored: bool = True,
    strip_horizon: int = 8,
) -> CompanyReport:
    """Run the full valuation flow for one ticker and return its report bundle.

    For hybrid names (HYBRID_TICKERS) the crude sleeve is carved out first and the
    price/target are crude-allocated, so the report compares like-for-like.

    ``use_transaction_anchored`` recalibrates the mid-age leg of any class
    with a populated ``inputs/market_data/transactions/<class>.yaml`` BEFORE
    carving / valuing (METHODOLOGY 9.9). DEFAULT ON since 2026-06-09 (owner
    decision): transaction-validated marks ARE the tool's marks; the un-anchored
    broker-resale curve is the diagnostic alternative (pass False to compare).
    """
    ci = load_company_inputs(ticker, quarter, inputs_dir)
    ci, txn_fits = _maybe_apply_transactions(ci, inputs_dir, use_transaction_anchored)
    carve = None
    if ticker in HYBRID_TICKERS:
        carve = crude_carve_out(ci)
        ci = carve.crude_inputs
        current_price = carve.carved_price(current_price)
        analyst_target = carve.carved_price(analyst_target)
    warnings = validate_inputs(ci)
    constructed = {"Aframax", "LR2"} & {v.cls for v in ci.fleet.vessels}
    if constructed:
        warnings.append(
            f"{'/'.join(sorted(constructed))} FFA forward curve is CONSTRUCTED (no market "
            f"anchor) — built from the 12M TC + spot, not a Baltic / $MT / Worldscale series. "
            f"Treat its dividend-strip contribution as indicative."
        )
    nav = compute_nav(ci)
    cycle = compute_cycle(ci)
    strip = compute_dividend_strip(
        ci, nav.nav_per_share, strip_horizon=strip_horizon,
        terminal_multiple=cycle.terminal_multiple,
    )
    blended = blend_fair_value(nav, strip, cycle)
    breakeven = implied_breakeven_tce(ci, current_price)
    sensitivity = compute_sensitivity(ci, current_price)
    payouts = payout_sensitivity(ci)

    notes: list[str] = []
    basis = ""
    if txn_fits:
        applied = ", ".join(
            f"{cls} 5yr ${f.new_5yr/1e6:.1f}M ({f.delta_5yr_pct:+.0f}%) / "
            f"10yr ${f.new_10yr/1e6:.1f}M ({f.delta_10yr_pct:+.0f}%) [n={f.n_used}]"
            for cls, f in txn_fits.items() if not f.fallback
        )
        if applied:
            notes.append(
                f"Mid-age value anchors **transaction-recalibrated** (METHODOLOGY 9.9): "
                f"{applied}. Newbuild + old-age anchors unchanged."
            )
    if carve is not None:
        whole_price = current_price / carve.crude_share if carve.crude_share else current_price
        basis = _crude_sleeve_basis(carve.crude_share, whole_price, current_price)
        crude_ev_pct = (blended.fair_value_per_share / current_price - 1.0) * 100.0
        notes.append(
            f"HYBRID crude carve-out (METHODOLOGY 6): crude sleeve = {carve.crude_share:.1%} "
            f"of vessel value (${carve.crude_value/1e6:,.0f}M crude vs "
            f"${carve.product_value/1e6:,.0f}M product). Price/target shown are crude-ALLOCATED "
            f"(whole-company x crude_share); balance sheet, G&A and corporate debt pro-rated, "
            f"LR1-secured ECA debt held with the product sleeve."
        )
        notes.append(product_read(carve, crude_ev_pct))
    if ci.fleet.fleet_schedule:
        notes.append(
            "Earning fleet varies over the strip per the manifest fleet_schedule "
            "(e.g. newbuild deliveries / sales); NAV is anchored at the report date."
        )
    if inputs_has_yard_discounts(ci):
        notes.append(
            "Vessel values carry a yard-quality discount (Chinese / ex-Hanjin-Subic yards); "
            "NAV is shown with and without it."
        )
    if any(v.cls == "LR2" for v in ci.fleet.vessels):
        notes.append(
            "LR2/Aframax vessels modeled as Aframax-equivalent (crude/dirty proxy) for v1; "
            "true clean-LR2 product rates would differ (v2: max of Aframax-crude and LR2-product)."
        )

    return CompanyReport(
        ticker=ticker,
        report_date=str(ci.fleet.report_date),
        current_price=current_price,
        analyst_target=analyst_target,
        nav=nav,
        strip=strip,
        cycle=cycle,
        blended=blended,
        breakeven=breakeven,
        sensitivity=sensitivity,
        warnings=warnings,
        payout_sensitivity=payouts,
        notes=notes,
        basis=basis,
    )


def inputs_has_yard_discounts(ci) -> bool:
    discounts = ci.market_data.yard_discounts
    return bool(discounts) and any(v.yard in discounts for v in ci.fleet.vessels)


def _append_hybrid_breakdown(
    path: Path, crude_r: ScenarioReport, product_r: ScenarioReport,
    whole_r: ScenarioReport,
) -> None:
    """Append the per-sleeve breakdown table to a hybrid name's scenario .md."""
    crude_ev = crude_r.expected_value_vs_current / crude_r.current_price * 100.0
    product_ev = product_r.expected_value_vs_current / product_r.current_price * 100.0
    whole_ev = whole_r.expected_value_vs_current / whole_r.current_price * 100.0
    crude_share = crude_r.current_price / whole_r.current_price
    product_share = product_r.current_price / whole_r.current_price
    text = (
        "\n## Hybrid sleeve breakdown (v2 whole-company aggregation)\n\n"
        "| Sleeve | Share | Allocated price | Weighted FV | EV% | Position |\n"
        "|---|--:|--:|--:|--:|---|\n"
        f"| Crude | {crude_share:.1%} | ${crude_r.current_price:,.2f} | "
        f"${crude_r.probability_weighted_fv:,.2f} | {crude_ev:+.1f}% | "
        f"{crude_r.position_recommendation.split()[0]} |\n"
        f"| Product | {product_share:.1%} | ${product_r.current_price:,.2f} | "
        f"${product_r.probability_weighted_fv:,.2f} | {product_ev:+.1f}% | "
        f"{product_r.position_recommendation.split()[0]} |\n"
        f"| **WHOLE-COMPANY** | 100% | **${whole_r.current_price:,.2f}** | "
        f"**${whole_r.probability_weighted_fv:,.2f}** | **{whole_ev:+.1f}%** | "
        f"**{whole_r.position_recommendation.split()[0]}** |\n\n"
        f"_Whole-company FV = crude FV + product FV (both per shares-outstanding); "
        f"compared against the whole-company tape price, not the carved proxy. "
        f"The product sleeve uses CLEAN trading rates (LR1/LR2 via clean curves, MR "
        f"via its own scenario forwards). The product sleeve carries MORE downside than "
        f"crude because product is leading the MoU rate normalisation (MR -52% w/w, "
        f"LR2 -28% w/w as of 2026-05-29) — flagged in METHODOLOGY 6 v2._\n"
    )
    with open(path, "a") as fh:
        fh.write(text)


def _append_product_read(path: Path, carve, report) -> None:
    """Legacy v1 qualitative product read — kept for backward-compat callers."""
    crude_ev_pct = report.expected_value_vs_current / report.current_price * 100.0
    text = (
        "\n## Hybrid carve-out & product sleeve (v1 qualitative)\n\n"
        f"- **Crude sleeve = {carve.crude_share:.1%} of vessel value** "
        f"(${carve.crude_value/1e6:,.0f}M crude vs ${carve.product_value/1e6:,.0f}M product). "
        f"All figures above are the CRUDE sleeve vs the crude-allocated price "
        f"(whole-company x crude_share).\n"
        f"- {product_read(carve, crude_ev_pct)}\n"
    )
    with open(path, "a") as fh:
        fh.write(text)


def run_watchlist(
    quarter: str,
    tickers: list[str] | None = None,
    inputs_dir: Path = INPUTS_DIR,
    outputs_dir: Path = OUTPUTS_DIR,
    live_prices: bool = False,
) -> list[CompanyReport]:
    """Value every ticker with populated inputs, write reports + the roll-up."""
    watchlist = load_watchlist(inputs_dir, live_prices=live_prices)
    tickers = tickers or list(watchlist.keys())
    # Per-sector strip horizon (METHODOLOGY §11.8.6.4) for the single-point FV,
    # consistent with the scenario layer's doc-driven horizon.
    horizons = {
        sector: int(doc.get("strip_horizon", 8))
        for sector, doc in _load_all_sectors(inputs_dir).items()
    }

    reports: list[CompanyReport] = []
    for ticker in tickers:
        entry = watchlist.get(ticker)
        if entry is None:
            print(f"skip {ticker}: no watchlist price")
            continue
        try:
            report = value_company(
                ticker, quarter, entry["current_price"], entry["analyst_target"], inputs_dir,
                strip_horizon=horizons.get(entry.get("sector", "crude"), 8),
            )
        except FileNotFoundError as exc:
            print(f"skip {ticker}: inputs not populated ({exc})")
            continue
        path = write_company_report(report, outputs_dir)
        tag = " [CRUDE SLEEVE]" if report.basis else ""
        print(f"{ticker}{tag}: FV ${report.blended.fair_value_per_share:,.2f} -> {path}")
        for warning in report.warnings:
            print(f"  ! {ticker} {warning}")
        reports.append(report)

    if reports:
        summary = write_watchlist_summary(reports, outputs_dir)
        print(f"watchlist roll-up -> {summary}")
    return reports


def run_scenarios_watchlist(
    quarter: str,
    tickers: list[str] | None = None,
    inputs_dir: Path = INPUTS_DIR,
    outputs_dir: Path = OUTPUTS_DIR,
    use_transaction_anchored: bool = True,
    live_prices: bool = False,
    asof_quarter: str | None = None,
):
    """Run the sector-layered scenario framework across the watchlist.

    For hybrids (HYBRID_TICKERS) the scenarios are run twice — once per sleeve —
    and aggregated into a whole-company ScenarioReport (METHODOLOGY 6 v2).

    ``asof_quarter`` routes the scenario strip timeline to a historical vintage
    (PLAN Phase 3b — the engine EV% as-of plumbing). ``None`` (default) keeps the
    live q3_2026 anchor, so the standard run is byte-unchanged; the single-point
    NAV/strip path is already as-of-correct via ``quarter`` (the strip is
    positional), so only the scenario quarter-key labels needed parametrizing.
    For a historical run the caller passes ``asof_quarter=quarter`` *and* must
    supply that vintage's scenario curves (else run_scenarios fails fast).
    """
    sector_docs = _load_all_sectors(inputs_dir)
    watchlist = load_watchlist(inputs_dir, live_prices=live_prices)
    tickers = tickers or list(watchlist.keys())

    reports = []
    for ticker in tickers:
        entry = watchlist.get(ticker)
        if entry is None:
            continue
        try:
            ci = load_company_inputs(ticker, quarter, inputs_dir)
        except FileNotFoundError:
            continue
        ci, _ = _maybe_apply_transactions(ci, inputs_dir, use_transaction_anchored)
        whole_price = entry["current_price"]
        whole_target = entry["analyst_target"]
        report, crude_r, product_r = _run_scenarios_for_ticker(
            ticker, ci, whole_price, whole_target, sector_docs, watchlist,
            asof_quarter=asof_quarter,
        )
        path = write_scenario_report(report, outputs_dir)
        if ticker in HYBRID_TICKERS and crude_r is not None:
            _append_hybrid_breakdown(path, crude_r, product_r, report)
        ev_pct = report.expected_value_vs_current / report.current_price * 100
        if ticker in HYBRID_TICKERS:
            tag = " [WHOLE-CO: crude+product]"
        elif ticker in MULTI_SLEEVE_TICKERS:
            tag = " [WHOLE-CO: " + "+".join(MULTI_SLEEVE_TICKERS[ticker]) + "]"
        else:
            tag = ""
        print(f"{ticker}{tag}: weighted FV ${report.probability_weighted_fv:,.2f} "
              f"(EV {ev_pct:+.1f}% -> {report.position_recommendation}) -> {path}")
        reports.append(report)

    if reports:
        summary = write_scenario_summary(reports, outputs_dir)
        print(f"scenario roll-up -> {summary}")
    return reports


@dataclass
class BrokerSweepRow:
    """One name's headline at tool / midpoint / broker vessel marks (section 9)."""

    ticker: str
    hybrid: bool
    consensus_pnav: float
    k_broker: float            # uniform mark premium that hits the broker NAV
    ev_tool: float             # EV% vs price at tool marks (k=1.00)
    ev_mid: float              # EV% at the midpoint
    ev_broker: float           # EV% at broker-equivalent marks
    pos_tool: str
    pos_broker: str
    be_tool: float             # blended breakeven TCE at tool marks
    be_broker: float           # blended breakeven TCE at broker marks
    fv_tool: float             # probability-weighted FV at tool marks
    fv_broker: float

    @property
    def spread(self) -> float:
        """EV swing from tool to broker marks — how mark-driven the call is."""
        return self.ev_broker - self.ev_tool


def _scenario_at_premium(
    ticker: str, ci_whole, k: float, price: float, target: float,
    sector_docs: dict, watchlist: dict,
) -> ScenarioReport:
    """Headline scenario report at vessel-mark premium ``k``.

    For hybrids returns the whole-company aggregation (vs the tape price), so
    the broker sweep's EV% / spread are whole-co quantities.
    """
    scaled = scale_vessel_marks(ci_whole, k)
    report, _, _ = _run_scenarios_for_ticker(
        ticker, scaled, price, target, sector_docs, watchlist,
    )
    return report


def run_broker_sweep(
    quarter: str,
    tickers: list[str] | None = None,
    inputs_dir: Path = INPUTS_DIR,
    outputs_dir: Path = OUTPUTS_DIR,
    use_transaction_anchored: bool = True,
    live_prices: bool = False,
) -> list[BrokerSweepRow]:
    """Value every name at tool / midpoint / broker vessel marks; write the roll-up.

    The broker endpoint is each name's OWN consensus-implied NAV (broker NAV =
    price / consensus_pnav), so validated pure-plays show ~zero spread and only
    genuinely mark-uncertain names (hybrids) show a wide tool-vs-broker swing.
    "Tool marks" = transaction-anchored marks (the pipeline default), so k_broker
    measures broker premium over TRANSACTION-VALIDATED levels.
    """
    sector_docs = _load_all_sectors(inputs_dir)
    watchlist = load_watchlist(inputs_dir, live_prices=live_prices)
    tickers = tickers or list(watchlist.keys())

    rows: list[BrokerSweepRow] = []
    for ticker in tickers:
        entry = watchlist.get(ticker)
        if entry is None or entry.get("consensus_pnav") is None:
            continue
        try:
            ci = load_company_inputs(ticker, quarter, inputs_dir)
        except FileNotFoundError:
            continue
        ci, _ = _maybe_apply_transactions(ci, inputs_dir, use_transaction_anchored)
        price, target = entry["current_price"], entry["analyst_target"]
        broker_nav = price / entry["consensus_pnav"]
        k_broker = solve_broker_premium(ci, broker_nav)
        k_mid = (1.0 + k_broker) / 2.0

        r_tool = _scenario_at_premium(ticker, ci, 1.0, price, target, sector_docs, watchlist)
        r_mid = _scenario_at_premium(ticker, ci, k_mid, price, target, sector_docs, watchlist)
        r_brk = _scenario_at_premium(ticker, ci, k_broker, price, target, sector_docs, watchlist)

        def _ev(r: ScenarioReport) -> float:
            return r.expected_value_vs_current / r.current_price * 100.0

        rows.append(BrokerSweepRow(
            ticker=ticker, hybrid=ticker in HYBRID_TICKERS or ticker in MULTI_SLEEVE_TICKERS,
            consensus_pnav=entry["consensus_pnav"], k_broker=k_broker,
            ev_tool=_ev(r_tool), ev_mid=_ev(r_mid), ev_broker=_ev(r_brk),
            pos_tool=r_tool.position_recommendation, pos_broker=r_brk.position_recommendation,
            be_tool=r_tool.breakeven_tce, be_broker=r_brk.breakeven_tce,
            fv_tool=r_tool.probability_weighted_fv, fv_broker=r_brk.probability_weighted_fv,
        ))

    if rows:
        path = _write_broker_sweep(rows, outputs_dir)
        for r in rows:
            print(f"{r.ticker}: EV tool {r.ev_tool:+.1f}% -> broker {r.ev_broker:+.1f}% "
                  f"(k_broker {r.k_broker:.2f}, spread {r.spread:+.0f}pp, "
                  f"{r.pos_tool.split()[0]}->{r.pos_broker.split()[0]})")
        print(f"broker-NAV sweep -> {path}")
    return rows


def _write_broker_sweep(rows: list[BrokerSweepRow], outputs_dir: Path) -> Path:
    """Render outputs/broker_nav_sweep.md (+ .xlsx) — METHODOLOGY 9 mark-uncertainty band."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    outputs_dir.mkdir(parents=True, exist_ok=True)
    out: list[str] = []
    w = out.append
    w("# Broker-NAV sensitivity sweep\n")
    lo, hi = TXN_PURE_PLAY_K_BAND
    w("Each name valued at three vessel-mark levels: **tool marks** (k=1.00, "
      "transaction-anchored since 2026-06-09), **midpoint**, and **broker-equivalent** "
      "(k lifts the tool NAV to the consensus broker NAV = price / consensus P/NAV). "
      "EV% = probability-weighted scenario FV vs price (crude-allocated for hybrids). "
      "k_broker is two-regime: on transaction-anchored sectors (crude/product/dry-bulk) "
      "it is the broker premium over transaction levels, and validated pure-plays are "
      f"EXPECTED inside the uniform band k {lo:.2f}-{hi:.2f} (~1.12-1.14 at the Jun-2026 "
      "fit, ~+13-17pp spread); on un-anchored sectors (LNG, containerships) it keeps the "
      "original broker-vs-independent-curve reading (validated ≈ 1.0). The **Read** column "
      "is mechanical spread width only — per-name mark-driven / mark-validated "
      "classification lives in METHODOLOGY 6.\n")
    w("| Name | Cons. P/NAV | k_broker | EV @tool | EV @mid | EV @broker | Pos tool→broker | "
      "Breakeven tool→broker | Spread (pp) | Read |")
    w("|---|--:|--:|--:|--:|--:|---|--:|--:|---|")
    def _be(x: float) -> str:
        return "NAV>px" if x < 5000 else f"${x:,.0f}"

    for r in sorted(rows, key=lambda x: x.ev_broker, reverse=True):
        read = "wide-spread" if abs(r.spread) >= 10 else "narrow-spread"
        tag = " **(WHOLE-CO)**" if r.hybrid else ""
        w(f"| {r.ticker}{tag} | {r.consensus_pnav:.2f}× | {r.k_broker:.2f} | "
          f"{r.ev_tool:+.1f}% | {r.ev_mid:+.1f}% | {r.ev_broker:+.1f}% | "
          f"{r.pos_tool.split()[0]}→{r.pos_broker.split()[0]} | "
          f"{_be(r.be_tool)}→{_be(r.be_broker)} | {r.spread:+.0f} | {read} |")
    if any(r.hybrid for r in rows):
        w("\n_**(WHOLE-CO)** = hybrid name valued via crude + product sleeve carve-outs "
          "aggregated against the whole-company tape price (METHODOLOGY 6 v2). The breakeven "
          "shown is the crude-sleeve breakeven (proxy)._\n")
    w(f"\n_On txn-anchored sectors, k_broker inside the {lo:.2f}-{hi:.2f} pure-play band "
      "⇒ the broker premium is the expected uniform one (mark-validated); k_broker outside "
      "the band (either side) ⇒ the name's apparent cheapness/richness is partly a NAV-mark "
      "choice (mark-driven — see the METHODOLOGY 6 entry for the per-name thesis). "
      "Pre-2026-06-09, validated read as k ≈ 1.0; that semantics survives only on "
      "un-anchored sectors. Uniform k also lifts the disposal-validated old-age leg (a "
      "known overshoot); leg-specific recalibration is the follow-up (METHODOLOGY 9)._\n")
    md_path = outputs_dir / "broker_nav_sweep.md"
    md_path.write_text("\n".join(out))

    wb = Workbook()
    ws = wb.active
    ws.title = "Broker-NAV sweep"
    headers = ["ticker", "basis", "consensus_pnav", "k_broker", "ev_tool_pct", "ev_mid_pct",
               "ev_broker_pct", "pos_tool", "pos_broker", "breakeven_tool", "breakeven_broker",
               "fv_tool", "fv_broker", "spread_pp"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in sorted(rows, key=lambda x: x.ev_broker, reverse=True):
        basis = "WHOLE-COMPANY (hybrid aggregation)" if r.hybrid else "whole-company"
        ws.append([r.ticker, basis, r.consensus_pnav, round(r.k_broker, 3),
                   round(r.ev_tool, 1), round(r.ev_mid, 1), round(r.ev_broker, 1),
                   r.pos_tool, r.pos_broker, round(r.be_tool, 0), round(r.be_broker, 0),
                   round(r.fv_tool, 2), round(r.fv_broker, 2), round(r.spread, 1)])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 26)
    wb.save(outputs_dir / "broker_nav_sweep.xlsx")
    return md_path


@dataclass
class TxnComparisonRow:
    """One name's headline at baseline vs transaction-anchored vessel marks."""

    ticker: str
    hybrid: bool
    nav_base: float
    nav_txn: float
    ev_base: float        # EV% vs price (crude-allocated for hybrids)
    ev_txn: float
    pos_base: str
    pos_txn: str
    fv_base: float        # probability-weighted FV
    fv_txn: float
    price_base: float
    price_txn: float      # crude-allocated price (may shift for hybrids if crude_share moves)

    @property
    def ev_delta(self) -> float:
        return self.ev_txn - self.ev_base


def _scenario_baseline_or_txn(
    ticker: str, quarter: str, price: float, target: float,
    sector_docs: dict, watchlist: dict, inputs_dir: Path, use_txn: bool,
) -> tuple[ScenarioReport, float]:
    """Run one name's whole-company scenario report. Returns (report, NAV/sh).

    NAV/sh is the whole-company NAV for hybrids (sum of sleeves) since the
    aggregated ScenarioReport is the headline.
    """
    ci = load_company_inputs(ticker, quarter, inputs_dir)
    ci, _ = _maybe_apply_transactions(ci, inputs_dir, use_txn)
    nav_per_share = compute_nav(ci).nav_per_share
    report, _, _ = _run_scenarios_for_ticker(
        ticker, ci, price, target, sector_docs, watchlist,
    )
    return report, nav_per_share


def run_transaction_anchored_comparison(
    quarter: str,
    tickers: list[str] | None = None,
    inputs_dir: Path = INPUTS_DIR,
    outputs_dir: Path = OUTPUTS_DIR,
    live_prices: bool = False,
) -> list[TxnComparisonRow]:
    """Value every name at baseline vs transaction-anchored marks side-by-side.

    Writes outputs/transaction_anchor_comparison.md / .xlsx and prints a summary.
    Diagnostic for the recalibration: shows which calls flip and by how much.
    """
    sector_docs = _load_all_sectors(inputs_dir)
    watchlist = load_watchlist(inputs_dir, live_prices=live_prices)
    tickers = tickers or list(watchlist.keys())

    rows: list[TxnComparisonRow] = []
    for ticker in tickers:
        entry = watchlist.get(ticker)
        if entry is None:
            continue
        try:
            r_base, nav_base = _scenario_baseline_or_txn(
                ticker, quarter, entry["current_price"], entry["analyst_target"],
                sector_docs, watchlist, inputs_dir, use_txn=False,
            )
            r_txn, nav_txn = _scenario_baseline_or_txn(
                ticker, quarter, entry["current_price"], entry["analyst_target"],
                sector_docs, watchlist, inputs_dir, use_txn=True,
            )
        except FileNotFoundError:
            continue

        def _ev(r: ScenarioReport) -> float:
            return r.expected_value_vs_current / r.current_price * 100.0

        rows.append(TxnComparisonRow(
            ticker=ticker, hybrid=ticker in HYBRID_TICKERS or ticker in MULTI_SLEEVE_TICKERS,
            nav_base=nav_base, nav_txn=nav_txn,
            ev_base=_ev(r_base), ev_txn=_ev(r_txn),
            pos_base=r_base.position_recommendation, pos_txn=r_txn.position_recommendation,
            fv_base=r_base.probability_weighted_fv, fv_txn=r_txn.probability_weighted_fv,
            price_base=r_base.current_price, price_txn=r_txn.current_price,
        ))

    if rows:
        path = _write_txn_comparison(rows, inputs_dir, outputs_dir)
        for r in rows:
            flip = " (FLIP)" if r.pos_base.split()[0] != r.pos_txn.split()[0] else ""
            print(f"{r.ticker}: NAV ${r.nav_base:,.2f}→${r.nav_txn:,.2f}  "
                  f"EV {r.ev_base:+.1f}%→{r.ev_txn:+.1f}%  "
                  f"{r.pos_base.split()[0]}→{r.pos_txn.split()[0]}{flip}")
        print(f"transaction-anchor comparison -> {path}")
    return rows


def _write_txn_comparison(
    rows: list[TxnComparisonRow], inputs_dir: Path, outputs_dir: Path
) -> Path:
    """Render outputs/transaction_anchor_comparison.md / .xlsx (METHODOLOGY 9.9)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    outputs_dir.mkdir(parents=True, exist_ok=True)
    txs = load_all_transactions(inputs_dir)
    out: list[str] = []
    w = out.append
    w("# Transaction-anchor recalibration — comparison\n")
    w("Each name valued twice: with the existing curve (baseline) and with the "
      "mid-age leg of every class with a populated transactions file recalibrated "
      "to disclosed S&P prints (recency-weighted, ~15mo half-life; quality-flag "
      "uplifts: financing +5%, distressed +10%). Newbuild + old-age anchors are "
      "left unchanged. Toggle: `use_transaction_anchored=True`.\n")
    if txs:
        w("## Transaction inputs in scope\n")
        for cls, ts in txs.items():
            w(f"- **{cls}** ({len(ts.prints)} prints, as_of {ts.as_of}):")
            for p in ts.prints:
                tag = "" if p.quality_flag == "standard" else f", {p.quality_flag} (+{int(100*0 if p.quality_flag=='newbuild_resale' else (5 if p.quality_flag=='financing' else 10 if p.quality_flag=='distressed' else 0))}%)"
                w(f"  - {p.date} • age {p.age:g} • ${p.price_usd_m}M{tag} • {p.source}")
        w("")
    w("## Per-name impact\n")
    w("| Name | NAV base→txn | Δ% | EV base→txn | Δpp | Position |")
    w("|---|--:|--:|--:|--:|---|")
    for r in sorted(rows, key=lambda x: x.ev_txn, reverse=True):
        d_nav = (r.nav_txn / r.nav_base - 1.0) * 100.0 if r.nav_base else 0.0
        tag = " **(WHOLE-CO)**" if r.hybrid else ""
        flip = " ⚠️" if r.pos_base.split()[0] != r.pos_txn.split()[0] else ""
        w(f"| {r.ticker}{tag} | ${r.nav_base:.2f}→${r.nav_txn:.2f} | {d_nav:+.1f}% | "
          f"{r.ev_base:+.1f}%→{r.ev_txn:+.1f}% | {r.ev_delta:+.1f} | "
          f"{r.pos_base.split()[0]}→{r.pos_txn.split()[0]}{flip} |")
    w("\n_Δ% = (txn − base) / base. Names with no exposure to a recalibrated class "
      "show Δ ≈ 0 — useful as a control. ⚠️ flags a position-call flip._")
    if any(r.hybrid for r in rows):
        w("\n_**(WHOLE-CO)** = hybrid name valued via crude + product carve-outs "
          "aggregated against the whole-company tape price (METHODOLOGY 6 v2)._\n")
    else:
        w("")

    md_path = outputs_dir / "transaction_anchor_comparison.md"
    md_path.write_text("\n".join(out))

    wb = Workbook()
    ws = wb.active
    ws.title = "Txn anchor comparison"
    ws.append(["ticker", "basis", "nav_base", "nav_txn", "nav_delta_pct",
               "ev_base_pct", "ev_txn_pct", "ev_delta_pp",
               "pos_base", "pos_txn", "fv_base", "fv_txn"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in sorted(rows, key=lambda x: x.ev_txn, reverse=True):
        d_nav = (r.nav_txn / r.nav_base - 1.0) * 100.0 if r.nav_base else 0.0
        basis = "WHOLE-COMPANY (hybrid aggregation)" if r.hybrid else "whole-company"
        ws.append([r.ticker, basis, round(r.nav_base, 2), round(r.nav_txn, 2),
                   round(d_nav, 2), round(r.ev_base, 1), round(r.ev_txn, 1),
                   round(r.ev_delta, 1), r.pos_base, r.pos_txn,
                   round(r.fv_base, 2), round(r.fv_txn, 2)])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 26)
    wb.save(outputs_dir / "transaction_anchor_comparison.xlsx")
    return md_path


def main() -> None:
    """Console entry point (see [project.scripts] in pyproject.toml)."""
    quarter = sys.argv[1] if len(sys.argv) > 1 else "2026-Q1"
    # Fail fast on anything that isn't a quarter ('--help', a typo'd '2026-Q5'):
    # a bogus quarter would skip every name on missing balance-sheet files yet
    # still run scenarios/xrefs and OVERWRITE state/last_run.json + touch
    # decisions/*.md (audit 2026-07-02, F-6).
    if not re.fullmatch(r"\d{4}-Q[1-4]", quarter):
        print(f"usage: python -m crude_tanker_fv.pipeline <QUARTER>  (e.g. 2026-Q1); "
              f"got {quarter!r}", file=sys.stderr)
        raise SystemExit(2)
    fv_reports = run_watchlist(quarter, live_prices=True)
    if not fv_reports:
        print(f"no ticker produced a report for {quarter} (inputs not populated?) — "
              f"aborting before scenario/xref/state writes", file=sys.stderr)
        raise SystemExit(1)
    print("--- scenarios ---")
    scenario_reports = run_scenarios_watchlist(quarter, live_prices=True)
    print("--- broker-NAV sweep ---")
    broker_rows = run_broker_sweep(quarter, live_prices=True)
    print("--- consensus forward-EPS cross-check ---")
    from crude_tanker_fv.consensus_eps import run_consensus_eps_xref
    run_consensus_eps_xref(quarter)
    print("--- justified P/NAV diagnostic ---")
    from crude_tanker_fv.justified_pnav import run_justified_pnav_xref
    run_justified_pnav_xref(quarter)
    print("--- book-wide scorecard (Thread 4) — consolidated verdict + validation ---")
    from crude_tanker_fv.scorecard import run_scorecard_xref
    run_scorecard_xref(
        quarter,
        fv_reports=fv_reports,
        scenario_reports=scenario_reports,
        broker_rows=broker_rows,
    )
    print("--- §12 dividend-window test ---")
    from crude_tanker_fv.dividend_window import run_dividend_window_xref
    run_dividend_window_xref(quarter)
    print("--- transaction-anchor comparison ---")
    run_transaction_anchored_comparison(quarter, live_prices=True)
    print("--- delta + decision log ---")
    _run_delta_and_decision_log(quarter, fv_reports, scenario_reports, broker_rows)


def _run_delta_and_decision_log(quarter, fv_reports, scenario_reports, broker_rows) -> None:
    """Snapshot the current run, diff vs the last one, write delta report and
    prepend per-ticker decision-log entries (METHODOLOGY §7.7, §7.8).

    Imported locally so the delta module doesn't pollute the top-level pipeline
    import graph — keeps the original pipeline path identical when called
    programmatically (e.g. from tests) via the individual run_* functions.
    """
    from crude_tanker_fv import delta

    current = delta.snapshot_current_run(quarter, fv_reports, scenario_reports, broker_rows)
    previous = delta.load_previous_snapshot()
    report = delta.compute_deltas(current, previous)
    delta_path = delta.write_delta_report(report)
    touched = delta.prepend_decision_log_entries(report)
    delta.save_snapshot(current)

    if report.is_first_run:
        print(f"delta report (FIRST RUN — no prior snapshot) -> {delta_path}")
    else:
        material_count = sum(1 for d in report.deltas if d.material)
        new_count = sum(1 for d in report.deltas if d.is_new)
        print(f"delta report ({material_count} material, {new_count} new, "
              f"{len(report.changed_input_files)} input files changed) -> {delta_path}")
    print(f"decision logs updated: {len(touched)} files (decisions/*.md)")


if __name__ == "__main__":
    main()
