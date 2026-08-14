"""Justified P/NAV diagnostic (METHODOLOGY.md §17).

A **coverage-independent** companion to the broker-NAV sweep (§9.9) and the
consensus-EPS cross-check (§9.11). Where those two lean on Pareto coverage (a
broker P/NAV, a consensus forward P/E), this leg asks a question answerable from
fundamentals alone — **"does the fleet earn its cost of capital on its own marked
NAV?"** — so it gives the APPROX / no-Pareto names (SB, CMDB, GSL, MPCC, CCEC,
NAT, ASC, TEN, CMBT) a NAV benchmark they otherwise lack, and lets the subsector
multiple structure fall out of fundamentals rather than hand-set scenario weights.

Formula (Gordon / residual-income justified price-to-book, applied to NAV):

    P/NAV*           = (RONAV_norm − g) / (r − g)
                     = 1 + (RONAV_norm − r) / (r − g)        # residual-income form
    Justified FV/sh  = P/NAV* × NAV_per_share
    RONAV_implied    = g + (price / NAV_per_share) × (r − g)  # the market's implied return
    P/NAV(mkt)       = price / NAV_per_share

where:

- ``NAV_per_share`` is the tool's CLEAN, un-haircut marked NAV (``compute_nav`` —
  the governance discount is applied downstream of it, never inside it, so this is
  the denominator the P/NAV signal needs).
- ``RONAV_norm`` is return on (marked) NAV at THROUGH-CYCLE earning power:
  ``normalized_annual_EPS / NAV_per_share``, where ``normalized_annual_EPS`` runs
  the existing dividend-strip earnings machinery with every vessel class's TCE set
  to its cycle anchor (``historical_tce_means`` — the same anchor ``cycle.py`` uses
  for cycle position), NOT the FFA forward curve (the hot NTM number the strip's
  front end carries near a peak, which would inflate the multiple). It is return on
  marked NAV, not on depreciated book (book always "earns well" mid-cycle and says
  nothing about whether the market value is justified), and it is mid-cycle, not
  next-twelve-months.
- ``r`` = ``COST_OF_EQUITY`` (11%), constant across all names in v1.
- ``g`` = per-subsector sustainable nominal growth (``sectors.<sector>.g`` in
  ``scenario_inputs.yaml``).

This is a DIAGNOSTIC ONLY — it is not wired into the headline FV, and whether
justified- (or subsector-demeaned-) P/NAV predicts forward returns is a separate,
pre-registered study (§17). ``r − g`` is a small denominator, so the multiple is
hypersensitive to ``g`` and ``RONAV_norm`` (±1pp swings it 10-20%): read it as an
ORDERING tool, not a precision estimate. See §17 for the dry-bulk anchor-bias and
dwt-scaling caveats.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, replace
from pathlib import Path
from statistics import median
from typing import Optional

from .dividend_strip import STRIP_HORIZON_QUARTERS, compute_dividend_strip
from .loaders import ALLOWED_CLASSES, INPUTS_DIR, load_company_inputs, load_watchlist
from .nav import COST_OF_EQUITY, compute_nav
from .normal_rates import normal_rate_table
from .report import OUTPUTS_DIR
from .scenarios import SCENARIOS_PATH, load_scenarios
from .schemas import CompanyInputs
from .vessel_values import vessel_market_value

# Hybrid name registries + the transaction-anchoring helper (single source of
# truth — pipeline owns them). The import is acyclic: pipeline imports THIS module
# only lazily inside main().
from .pipeline import (
    HYBRID_TICKERS,
    MULTI_SLEEVE_TICKERS,
    THREE_SLEEVE_TICKERS,
    _maybe_apply_transactions,
)

# Fallback per-subsector growth if a sector lacks an explicit ``g:`` in the YAML.
DEFAULT_G = 0.01

# Newbuild value share above which RONAV_norm is unreliable (a not-yet-delivered
# hull earns a full anchor-year in the strip while its NAV is PV-haircut and its
# capex commitment never hits strip EPS — the numerator/denominator bases are
# irreconcilable). The observed distribution is bimodal: newbuild-DOMINATED names
# (BRUT 100%, CAPT 73%, MPCC 40%) sit well above a flagship with a modest program
# (FRO ~17%), so 25% cleanly separates them — FRO computes (with a mild residual
# bias, footnoted) and stays in its sector median.
NEWBUILD_HEAVY_SHARE = 0.25

# |Justified P/NAV − P/NAV(mkt)| within this fraction of P/NAV(mkt) reads "fair".
FAIR_BAND = 0.10

# Deadband on the GOVERNED read flag (tier semantics amendment 2026-08-13, §3; owner-set at
# ratification). The instantaneous `robust` is truthful every run and is DISPLAY; `read_flag` is
# what governance consumes, and it adopts a new state only once the price sits this far (in %)
# beyond the boundary that changed it. Motivation: SBLK closed 0.62% off its re-promotion
# boundary — inside a strobe zone where a 62 bp move restates the sizing input either way.
READ_FLAG_HYST_PCT = 2.0

_NTM_QUARTERS = 4


# ---------------------------------------------------------------------------
# Pure formula functions (unit-tested to agree / round-trip)
# ---------------------------------------------------------------------------
def justified_pnav(ronav_norm: float, r: float, g: float) -> float:
    """Gordon form: (RONAV_norm − g) / (r − g)."""
    return (ronav_norm - g) / (r - g)


def justified_pnav_resid(ronav_norm: float, r: float, g: float) -> float:
    """Residual-income form: 1 + (RONAV_norm − r) / (r − g). Equals the Gordon form."""
    return 1.0 + (ronav_norm - r) / (r - g)


def cheap_fair_boundary(nav_per_share: float, justified: float) -> float:
    """Price at which this basis stops reading `cheap`: NAV × J / (1 + FAIR_BAND).

    Inverts the band form in ``_read_from`` — which measures the gap as a fraction of P/NAV(mkt),
    not of J — so `cheap` is pnav_mkt < J/(1+FAIR_BAND), i.e. price below this. The boundary price
    itself reads `fair` (the comparison is <=)."""
    return nav_per_share * justified / (1.0 + FAIR_BAND)


def fair_rich_boundary(nav_per_share: float, justified: float) -> float:
    """Price at which this basis starts reading `rich`: NAV × J / (1 − FAIR_BAND)."""
    return nav_per_share * justified / (1.0 - FAIR_BAND)


def _robust_from(read: str, read_hist: str) -> str:
    """Does the cheap/rich call survive the basis choice? Single definition, shared by the row
    property and the boundary scan so a margin can never disagree with the flag it measures."""
    if read in ("cheap", "fair", "rich") and read_hist in ("cheap", "fair", "rich"):
        return "robust" if read == read_hist else f"flips ({read}/{read_hist})"
    return "n/a"


def _robust_at_price(nav_per_share: float, j_par: Optional[float],
                     j_hist: Optional[float], price: float) -> str:
    """Recompute the robust state at an arbitrary price (both J's and NAV held fixed)."""
    pnav = price / nav_per_share if nav_per_share > 0 else None
    return _robust_from(_read_from(j_par, pnav, None), _read_from(j_hist, pnav, None))


def boundary_prices(nav_per_share: float, j_par: Optional[float],
                    j_hist: Optional[float]) -> dict:
    """The four §17 band edges as PRICES, keyed ``<basis>_<edge>``. None where the basis is blocked."""
    out: dict = {}
    for basis, j in (("par", j_par), ("hist", j_hist)):
        out[f"{basis}_cheap_fair"] = cheap_fair_boundary(nav_per_share, j) if j is not None else None
        out[f"{basis}_fair_rich"] = fair_rich_boundary(nav_per_share, j) if j is not None else None
    return out


# Relative nudge used to test whether crossing a boundary changes the robust state. Far below any
# economically meaningful move, far above double-precision noise at these price magnitudes.
_BOUNDARY_PROBE = 1e-9


def flip_margin_pct(nav_per_share: float, j_par: Optional[float],
                    j_hist: Optional[float], price: float) -> Optional[float]:
    """Signed % distance from the nearest boundary whose crossing would change the robust state.

    Not every band edge qualifies: crossing one changes a single basis's read, which only matters
    where that moves AGREEMENT. Each candidate is therefore probed either side and kept only if the
    robust state actually differs across it — so the margin always measures the distance to a real
    state change, never to a cosmetic one. Positive = price sits above the boundary.

    None when no boundary can change the state (a §17-blocked name has no read to flip)."""
    if nav_per_share <= 0 or price <= 0:
        return None
    changing = [
        b for b in boundary_prices(nav_per_share, j_par, j_hist).values()
        if b is not None and b > 0
        and _robust_at_price(nav_per_share, j_par, j_hist, b * (1 - _BOUNDARY_PROBE))
        != _robust_at_price(nav_per_share, j_par, j_hist, b * (1 + _BOUNDARY_PROBE))
    ]
    if not changing:
        return None
    nearest = min(changing, key=lambda b: abs(price - b) / b)
    return (price - nearest) / nearest * 100.0


READ_FLAG_STATE_FILE = Path(__file__).resolve().parents[2] / "state" / "read_flag_state.json"


def govern_read_flag(instantaneous: str, margin_pct: Optional[float],
                     prior: Optional[str]) -> str:
    """Apply the READ_FLAG_HYST_PCT deadband to the instantaneous read state.

    A transition is adopted only once the price sits at least the deadband beyond the boundary that
    caused it; inside the band the prior state HOLDS, so a name parked on a boundary reports one
    stable sizing input instead of strobing between two. Bootstraps to the instantaneous state when
    there is no prior (state/ is machine-local and regenerated, so a fresh clone starts here).

    A ``None`` margin with a changed state means no boundary can explain the change — the J's
    themselves moved (new quarter data), which is an estimate change, not a strobe: adopt it."""
    if instantaneous == "n/a":
        return "n/a"
    if prior is None or prior == "n/a" or prior == instantaneous:
        return instantaneous
    if margin_pct is None or abs(margin_pct) >= READ_FLAG_HYST_PCT:
        return instantaneous
    return prior


def load_read_flag_state(path: Path = READ_FLAG_STATE_FILE) -> dict:
    """Prior governed read flags, ``{ticker: read_flag}``. Missing file = no priors."""
    if not path.exists():
        return {}
    return json.loads(path.read_text()).get("flags", {})


def save_read_flag_state(flags: dict, path: Path = READ_FLAG_STATE_FILE,
                         asof: str = "") -> None:
    """Persist the governed flags. Blocked names are NOT written (Addendum A §2): a name with no
    producible read has no governed state to carry forward."""
    keep = {t: f for t, f in flags.items() if f != "n/a"}
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({"asof": asof, "hyst_pct": READ_FLAG_HYST_PCT,
                                "flags": keep}, indent=1, sort_keys=True) + "\n")


def ronav_implied(price: float, nav_per_share: float, r: float, g: float) -> float:
    """The return on NAV the market is pricing in: g + (price/NAV) × (r − g)."""
    return g + (price / nav_per_share) * (r - g)


# ---------------------------------------------------------------------------
# Normalized through-cycle earning power
# ---------------------------------------------------------------------------
def normalized_annual_eps(
    ci: CompanyInputs, nav_per_share: float, anchors: dict[str, float]
) -> tuple[Optional[float], list[str]]:
    """Annual EPS with every vessel class's TCE pinned to a through-cycle anchor.

    ``anchors`` is the per-class flat rate to pin TCE to — either the ``parity``
    or the ``historical_mean`` basis (P1; PRE_REGISTRATION_NORMAL_RATES.md). The
    same machinery serves both, so the two RONAVs differ only by the anchor.

    Returns ``(eps, missing_anchor_classes)``. ``eps`` is None iff a fleet class
    has no anchor (then ``missing`` is non-empty). The strip is reused unchanged
    via a normalized ``CompanyInputs``:

    - ``ffa_forward_curve`` -> flat per-class anchor (the through-cycle rate),
    - coverage neutralized (cov=0 ⇒ blended TCE = anchor for EVERY vessel, so a
      name's idiosyncratic locked-charter book does not leak into a through-cycle
      number — load-bearing for names with real disclosed charters like CCEC),
    - ``fleet_schedule`` neutralized ⇒ all current manifest vessels earn a full
      anchor-year (steady-state), matched to the full-fleet NAV denominator.

    Annual EPS = sum of the first four quarters (flat rates + static fleet ⇒ the
    four quarters are identical, so this is 4× the quarterly figure). The strip's
    terminal value is computed but discarded — only ``eps_by_quarter`` is read.
    """
    md = ci.market_data
    fleet_classes = {v.cls for v in ci.fleet.vessels}
    missing = sorted(c for c in fleet_classes if c not in anchors)
    if missing:
        return None, missing
    norm_ffa = {cls: [anchors[cls]] * STRIP_HORIZON_QUARTERS for cls in fleet_classes}
    norm_md = replace(md, ffa_forward_curve=norm_ffa)
    norm_fleet = replace(
        ci.fleet, coverage_schedule={}, spot_coverage_pct={}, fleet_schedule={}
    )
    norm_ci = replace(ci, market_data=norm_md, fleet=norm_fleet)
    strip = compute_dividend_strip(norm_ci, nav_per_share)
    return sum(strip.eps_by_quarter[:_NTM_QUARTERS]), []


def _newbuild_value_share(ci: CompanyInputs) -> float:
    """Fraction of fleet market value carried by not-yet-delivered newbuilds."""
    curves = ci.market_data.vessel_value_curves
    yard_discounts = ci.market_data.yard_discounts
    total = 0.0
    newbuild = 0.0
    for v in ci.fleet.vessels:
        value = vessel_market_value(v, curves[v.cls], yard_discounts) * v.count
        total += value
        if v.years_to_delivery > 0:
            newbuild += value
    return newbuild / total if total else 0.0


# ---------------------------------------------------------------------------
# Row model
# ---------------------------------------------------------------------------
def _read_from(justified: Optional[float], pnav_mkt: Optional[float],
               flag: Optional[str]) -> str:
    """cheap / fair / rich vs the market multiple, or the blocking flag."""
    if flag is not None:
        return flag
    if justified is None or pnav_mkt is None:
        return "n/a"
    diff = justified - pnav_mkt
    if abs(diff) <= FAIR_BAND * pnav_mkt:
        return "fair"
    return "cheap" if diff > 0 else "rich"


@dataclass
class JustifiedPnavRow:
    """One name's justified-P/NAV read under BOTH normal-rate bases (P1).

    The PRIMARY fields are the ``parity`` basis (the headline — it closes the §17
    loop: parity is the rate that makes justified-P/NAV = 1 for a newbuild). The
    ``*_hist`` fields are the ``historical_mean`` cross-check. ``pnav_mkt`` and
    ``ronav_implied`` are basis-independent (market-priced). Multiple/FV/label are
    None when a guard trips. The deliverable is the DIVERGENCE — whether the
    cheap/rich call survives the choice between the two legitimate bases.
    """

    ticker: str
    hybrid: bool
    sector: str
    nav_per_share: float
    price: float
    pnav_mkt: Optional[float]              # price / NAV (basis-independent)
    ronav_implied: Optional[float]         # g + P/NAV(mkt)·(r − g) (basis-independent)
    r: float
    g: float
    # parity basis (headline)
    ronav_norm: Optional[float]            # parity normalized annual EPS / NAV
    justified_pnav: Optional[float]        # (RONAV_parity − g)/(r − g)
    justified_fv: Optional[float]
    gap: Optional[float]                   # RONAV_parity − RONAV_implied
    flag: Optional[str]
    # historical_mean basis (cross-check)
    ronav_norm_hist: Optional[float]
    justified_pnav_hist: Optional[float]
    gap_hist: Optional[float]
    flag_hist: Optional[str]
    # §15 governance realisation haircut — applied at the blend layer + strip terminal, NOT in
    # this leg (which uses CLEAN marked NAV). >0 only for TEN/CMDB. Surfaced for the dual-read.
    governance_discount_pct: float = 0.0

    @property
    def read(self) -> str:
        """Headline (parity-basis) read."""
        return _read_from(self.justified_pnav, self.pnav_mkt, self.flag)

    @property
    def read_hist(self) -> str:
        return _read_from(self.justified_pnav_hist, self.pnav_mkt, self.flag_hist)

    @property
    def robust(self) -> str:
        """Does the cheap/rich call survive the basis choice? The P1 deliverable.

        INSTANTANEOUS and display-only since the 2026-08-13 tier semantics amendment: it is an EDGE
        fact (a function of where the price sits), so it no longer feeds the confidence tier, and
        governance consumes the hysteresis-governed ``read_flag`` rather than this."""
        return _robust_from(self.read, self.read_hist)

    @property
    def read_blocked(self) -> Optional[str]:
        """The §17 blocking guard's label when no multiple could be produced, else None.

        The CONSTRUCTION half of the §17 signal and the tier's only §17 input (Addendum A
        2026-08-13): the tier asks whether a read was producible, never what it said. Every blocker
        in ``evaluate`` is price-independent, which is what keeps the tier price-invariant —
        ``test_tier_is_price_invariant`` is the standing guard on that."""
        return self.flag or self.flag_hist

    @property
    def boundary_prices(self) -> dict:
        """The four §17 band edges as prices (``par_cheap_fair`` … ``hist_fair_rich``)."""
        return boundary_prices(self.nav_per_share, self.justified_pnav, self.justified_pnav_hist)

    @property
    def flip_margin_pct(self) -> Optional[float]:
        """Signed % distance from the nearest boundary that would change ``robust``.

        Measured at the row's own (watchlist-vintage) price, the same price every other §17 number
        on the row uses — NOT the live tape, which moves between vintage promotes."""
        return flip_margin_pct(self.nav_per_share, self.justified_pnav,
                               self.justified_pnav_hist, self.price)


def _g_by_sector(inputs_dir: Path) -> dict[str, float]:
    """Per-sector sustainable growth from ``sectors.<sector>.g``."""
    path = inputs_dir / "scenario_inputs.yaml" if inputs_dir != INPUTS_DIR else SCENARIOS_PATH
    out: dict[str, float] = {}
    for sector in ("crude", "product", "lng", "dry_bulk", "containerships"):
        try:
            doc = load_scenarios(path, sector)
        except KeyError:
            continue
        gv = doc.get("g")
        if gv is not None:
            out[sector] = float(gv)
    return out


@dataclass
class _Eval:
    """Outcome of the guard+formula decision (pure; testable with scalars)."""

    flag: Optional[str]
    ronav_norm: Optional[float]
    pnav_mkt: Optional[float]
    ronav_implied: Optional[float]
    justified_pnav: Optional[float]
    justified_fv: Optional[float]
    gap: Optional[float]


def evaluate(
    nav: float,
    price: float,
    eps: Optional[float],
    missing_anchor: bool,
    has_cost: bool,
    newbuild_share: float,
    r: float,
    g: float,
) -> _Eval:
    """Apply the ordered guards and (if all pass) the justified-P/NAV formula.

    ``eps`` is the normalized annual EPS (None when a guard upstream made it
    uncomputable). The first blocking guard wins; a flagged row carries no
    multiple/FV but may still show RONAV_norm (newbuild-heavy) and a gap.
    """
    pnav_mkt = price / nav if nav > 0 else None
    implied = ronav_implied(price, nav, r, g) if nav > 0 and r > g else None
    flag: Optional[str] = None
    ronav: Optional[float] = None
    if nav <= 0:
        flag = "non-positive NAV"
    elif not has_cost:
        flag = "no cost data"
    elif r - g <= 0:
        flag = "r≤g (invalid)"
    elif missing_anchor:
        flag = "no anchor"
    else:
        # newbuild-heavy still shows RONAV_norm (for transparency) but no multiple
        if newbuild_share > NEWBUILD_HEAVY_SHARE:
            flag = "newbuild-heavy (unreliable)"
        ronav = eps / nav
        if eps < 0 and flag is None:
            flag = "negative mid-cycle EPS"
        elif ronav < g and flag is None:
            flag = "sub-growth returns"
    gap = ronav - implied if (ronav is not None and implied is not None) else None
    justified = justified_pnav(ronav, r, g) if (flag is None and ronav is not None) else None
    fv = justified * nav if justified is not None else None
    return _Eval(flag, ronav, pnav_mkt, implied, justified, fv, gap)


def compute_justified_pnav_rows(
    quarter: str, inputs_dir: Path = INPUTS_DIR, use_transaction_anchored: bool = True
) -> list[JustifiedPnavRow]:
    """Build the justified-P/NAV row for every valued watchlist name.

    Coverage-independent: iterates ALL names (no consensus_fwd_pe / consensus_pnav
    gate), so the APPROX names get a benchmark too.

    ``use_transaction_anchored`` (default True, the production basis) recalibrates
    the marks via ``_maybe_apply_transactions`` BEFORE ``compute_nav`` — the same
    NAV every other surface uses (broker sweep, headline FV; METHODOLOGY §9.9,
    "transaction-validated marks ARE the tool's marks"). Pass False for the
    un-anchored diagnostic NAV.
    """
    watchlist = load_watchlist(inputs_dir)
    g_by_sector = _g_by_sector(inputs_dir)
    # Parity anchors — computed ONCE for all classes (PRE_REGISTRATION §; P1).
    parity_anchors = {
        cls: nr.parity
        for cls, nr in normal_rate_table(quarter, sorted(ALLOWED_CLASSES), inputs_dir=inputs_dir).items()
        if nr.parity is not None
    }
    rows: list[JustifiedPnavRow] = []
    for ticker, entry in watchlist.items():
        try:
            ci = load_company_inputs(ticker, quarter, inputs_dir)
        except FileNotFoundError:
            continue
        ci, _ = _maybe_apply_transactions(ci, inputs_dir, use_transaction_anchored)
        sector = entry.get("sector", "crude")
        g = g_by_sector.get(sector, DEFAULT_G)
        r = COST_OF_EQUITY
        # Vintage price, same convention as consensus_eps (deterministic; tool-NAV based).
        price = float(entry.get("as_of_price") or entry["current_price"])
        nav_res = compute_nav(ci)
        nav = nav_res.nav_per_share
        has_cost = bool(ci.cost_structure.opex_per_day)
        nb_share = _newbuild_value_share(ci)

        def _basis(anchors: dict[str, float]) -> _Eval:
            eps: Optional[float] = None
            missing: list[str] = []
            if nav > 0 and has_cost and r > g:
                eps, missing = normalized_annual_eps(ci, nav, anchors)
            return evaluate(nav, price, eps, bool(missing), has_cost, nb_share, r, g)

        par = _basis(parity_anchors)                              # headline
        hist = _basis(ci.market_data.historical_tce_means)        # cross-check

        rows.append(JustifiedPnavRow(
            ticker=ticker,
            hybrid=(
                ticker in HYBRID_TICKERS
                or ticker in THREE_SLEEVE_TICKERS
                or ticker in MULTI_SLEEVE_TICKERS
            ),
            sector=sector,
            nav_per_share=nav,
            price=price,
            pnav_mkt=par.pnav_mkt,
            ronav_implied=par.ronav_implied,
            r=r,
            g=g,
            ronav_norm=par.ronav_norm,
            justified_pnav=par.justified_pnav,
            justified_fv=par.justified_fv,
            gap=par.gap,
            flag=par.flag,
            ronav_norm_hist=hist.ronav_norm,
            justified_pnav_hist=hist.justified_pnav,
            gap_hist=hist.gap,
            flag_hist=hist.flag,
            governance_discount_pct=nav_res.governance_discount_pct,
        ))
    return rows


# ---------------------------------------------------------------------------
# Summaries
# ---------------------------------------------------------------------------
# Sectors whose medians compose into the headline subsector vector (METHODOLOGY §10).
# crude/product/dry_bulk anchors compose; LNG (spike-inclusive, §18.2) and containerships
# (fy_calendar_avg, boom-tilted) do NOT — they are SUPPRESSED from the headline vector
# (their per-name rows are retained in the table above). Per-name reads still stand.
HEADLINE_VECTOR_SECTORS = frozenset({"crude", "product", "dry_bulk"})


def subsector_median_pnav(rows: list[JustifiedPnavRow]) -> dict[str, float]:
    """Median Justified P/NAV by sector (parity basis, valid + composable rows) — headline."""
    by_sector: dict[str, list[float]] = {}
    for r in rows:
        if r.justified_pnav is not None and r.sector in HEADLINE_VECTOR_SECTORS:
            by_sector.setdefault(r.sector, []).append(r.justified_pnav)
    return {s: median(v) for s, v in by_sector.items() if v}


def subsector_median_pnav_hist(rows: list[JustifiedPnavRow]) -> dict[str, float]:
    """Median Justified P/NAV by sector on the historical_mean cross-check basis. LNG +
    containerships are suppressed (non-composable, §10/§18.2) — same headline-vector filter."""
    by_sector: dict[str, list[float]] = {}
    for r in rows:
        if r.justified_pnav_hist is not None and r.sector in HEADLINE_VECTOR_SECTORS:
            by_sector.setdefault(r.sector, []).append(r.justified_pnav_hist)
    return {s: median(v) for s, v in by_sector.items() if v}


def _sector_base_ronav(rows: list[JustifiedPnavRow]) -> dict[str, float]:
    """Median RONAV_norm by sector (valid rows) — the sensitivity-grid base."""
    by_sector: dict[str, list[float]] = {}
    for r in rows:
        if r.justified_pnav is not None and r.ronav_norm is not None:
            by_sector.setdefault(r.sector, []).append(r.ronav_norm)
    return {s: median(v) for s, v in by_sector.items() if v}


_GRID_G = (0.0, 0.01, 0.02)
_GRID_RONAV_OFFSETS = (-0.02, 0.0, 0.02)


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
def _pct(x: Optional[float], places: int = 1, signed: bool = False) -> str:
    if x is None:
        return "n/a"
    sign = "+" if signed else ""
    return f"{x * 100.0:{sign}.{places}f}%"


def write_justified_pnav(
    rows: list[JustifiedPnavRow], outputs_dir: Path = OUTPUTS_DIR
) -> Path:
    """Render outputs/justified_pnav.md (+ .xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    outputs_dir.mkdir(parents=True, exist_ok=True)
    r0 = COST_OF_EQUITY
    out: list[str] = []
    w = out.append
    w("# Justified P/NAV diagnostic\n")
    w("A **coverage-independent** fair-multiple benchmark (METHODOLOGY §17). The broker-NAV "
      "sweep (§9.9) and consensus-EPS cross-check (§9.11) lean on Pareto coverage; this asks a "
      "question answerable from fundamentals alone — **does the fleet earn its cost of capital "
      "on its own marked NAV (net asset value)?** — so the APPROX / no-Pareto names (SB, CMDB, "
      "GSL, MPCC, CCEC, NAT, ASC, TEN, CMBT) get a NAV benchmark too.\n")
    w("**`RONAV` = return on NAV** — the asset-NAV analog of return on equity: annual earnings ÷ "
      "NAV per share. Two variants appear below. **`RONAV_norm`** (normalized) is what the fleet "
      "*would* earn on its marked NAV at mid-cycle rates; **`RONAV_implied`** is the return on NAV "
      "the *market price* is implying, backed out of the same identity.\n")
    w("`P/NAV* = (RONAV_norm − g)/(r − g) = 1 + (RONAV_norm − r)/(r − g)`; "
      "`Justified FV/sh = P/NAV* × NAV/sh`; `RONAV_implied = g + P/NAV(mkt)·(r − g)`. "
      f"`r` = cost of equity {r0:.0%} (constant in v1). `NAV/sh` is the tool's CLEAN, "
      "un-haircut marked NAV — transaction-anchored (METHODOLOGY §9.9), the SAME basis the "
      "headline FV and broker sweep use; governance discount is applied downstream, never inside it.\n")
    w("**RONAV_norm is return on *marked NAV*, not on accounting book**, and **through-cycle, not "
      "NTM (next-twelve-months)**: `normalized_annual_EPS / NAV/sh`, where the EPS runs the "
      "dividend-strip earnings machinery with every vessel class's day-rate (TCE, time-charter "
      "equivalent) pinned to a through-cycle anchor, NOT the FFA (forward freight agreement) "
      "forward curve. Book always 'earns well' mid-cycle and says nothing about whether the market "
      "value is justified; the FFA front end is the hot near-term number that would inflate the "
      "multiple — both are deliberately avoided.\n")
    w("**Two normal-rate bases (P1, §18; PRE_REGISTRATION_NORMAL_RATES.md).** Each name is shown "
      "under BOTH: **`parity`** (headline) — replacement economics, the TCE that lets a newbuild "
      "earn its cost of capital (closes the loop: justified-P/NAV = 1 for a newbuild); and "
      "**`historical_mean`** (cross-check) — the realized through-cycle anchor (current "
      "`historical_tce_means`). The per-class divergence (`historical − parity`) is the "
      "under-/over-ordered signal. The deliverable is whether the cheap/rich call **survives the "
      "basis choice** (`Robust` column): survives → a genuine read; **flips → the call depends on "
      "the normalization philosophy**, which is itself the finding. PROVISIONAL pending the §18.5b "
      "orderbook validation of the divergence.\n")
    w("**Read this as an ORDERING tool, not a precision estimate.** `r − g` is a small "
      "denominator, so the multiple is hypersensitive: ±1pp on `g` or `RONAV_norm` swings it "
      "10-20% (see the per-sector sensitivity grids below). **Anchor-bias caveats — RONAV_norm "
      "inherits the cycle anchors, which are not all true long-run means:** (1) **dry-bulk** "
      "anchors are 22-month firm-window medians, biased elevated (§11.7.5), so its multiples are "
      "an **upper bound**; (2) **containership** anchors are FY2021-2025 calendar averages "
      "(boom-tilted, `fy_calendar_avg` per §10 — NOT a through-cycle mean), so containership "
      "`RONAV_norm` is biased high even more than dry-bulk — GSL/MPCC's multiples are a **loose "
      "upper bound, not a real target**; (3) NAV dwt-scales per vessel but strip revenue is "
      "per-class count-based, so large-hull dry-bulk names (SB, CMBT) are biased toward 'rich' "
      "(partially offsetting (1)).\n")
    w("**A crude name reading `rich` near a cycle peak is cycle POSITION, not a short signal.** "
      "RONAV_norm is pinned to a through-cycle anchor while the market price embeds the hot near-peak "
      "NTM rate, so a crude pure-play at/near peak (DHT/FRO/ECO/NAT) reads `rich` BY CONSTRUCTION "
      "(the §12 NAT mechanism) — it says where in the cycle the name sits, not TRIM/SHORT. Read the "
      "crude `rich` column together with the cycle position, never as a standalone call.\n")
    w("**Not in the headline FV** (diagnostic only); whether justified-P/NAV ranking predicts "
      "forward returns is a separate pre-registered study.\n")

    w("| Ticker | Sector | NAV/sh | Price | P/NAV (mkt) | RONAV (par) | Just P/NAV (par) | "
      "RONAV (hist) | Just P/NAV (hist) | Read: par → hist | Robust? |")
    w("|---|---|--:|--:|--:|--:|--:|--:|--:|---|---|")

    def _sort_key(row: JustifiedPnavRow):
        # Valid rows by parity gap descending (cheapest first); flagged rows last.
        return (0, -(row.gap if row.gap is not None else 0.0)) if row.flag is None else (1, 0.0)

    def _jp(x: Optional[float]) -> str:
        return f"{x:.3f}×" if x is not None else "—"

    for r in sorted(rows, key=_sort_key):
        tag = " **(WHOLE-CO)**" if r.hybrid else ""
        pnav_mkt = f"{r.pnav_mkt:.3f}×" if r.pnav_mkt is not None else "n/a"
        w(f"| {r.ticker}{tag} | {r.sector} | ${r.nav_per_share:,.2f} | ${r.price:,.2f} | "
          f"{pnav_mkt} | {_pct(r.ronav_norm)} | {_jp(r.justified_pnav)} | "
          f"{_pct(r.ronav_norm_hist)} | {_jp(r.justified_pnav_hist)} | "
          f"{r.read} → {r.read_hist} | {r.robust} |")

    medians = subsector_median_pnav(rows)
    medians_h = subsector_median_pnav_hist(rows)
    w("\n## Subsector vector — median Justified P/NAV (parity headline, historical cross-check)\n")
    w("| Sector | Median Just P/NAV (parity) | Median (historical) | n |")
    w("|---|--:|--:|--:|")
    counts = {s: sum(1 for r in rows if r.sector == s and r.justified_pnav is not None)
              for s in medians}
    for s in sorted(medians, key=lambda s: medians[s], reverse=True):
        mh = f"{medians_h[s]:.3f}×" if s in medians_h else "—"
        w(f"| {s} | {medians[s]:.3f}× | {mh} | {counts[s]} |")
    w("\n_The two columns ARE the signal: where parity ≫ historical, the sector reads cheaper under "
      "replacement economics than under its (boom/firm-window-biased) historical anchor — the §18 "
      "under-ordering. The §17.6 anchor-bias caveats apply to the historical column only; parity is "
      "independent of those biases (it is built from newbuild cost, not a rate-history window)._\n")
    w("\n_The headline vector covers the COMPOSABLE sectors only (crude / product / dry_bulk). "
      "**LNG and containership medians are suppressed** as non-composable (§10: containerships on "
      "`fy_calendar_avg`, LNG spike-inclusive) — their per-name reads remain in the table above but "
      "do not roll into a cross-sector median._\n")

    bases = _sector_base_ronav(rows)
    w("\n## Sensitivity grids — Justified P/NAV across g × RONAV_norm "
      f"(r = {r0:.0%}, base = sector median RONAV_norm)\n")
    for s in sorted(bases):
        base = bases[s]
        w(f"\n**{s}** (base RONAV_norm {base * 100:.1f}%)\n")
        w("| RONAV_norm \\ g | " + " | ".join(f"g={gg:.0%}" if gg else "g=0%" for gg in _GRID_G) + " |")
        w("|---|" + "--:|" * len(_GRID_G))
        for off in _GRID_RONAV_OFFSETS:
            ronav = base + off
            cells = []
            for gg in _GRID_G:
                if r0 - gg <= 0:
                    cells.append("n/a")
                else:
                    cells.append(f"{justified_pnav(ronav, r0, gg):.2f}×")
            w(f"| {ronav * 100:.1f}% | " + " | ".join(cells) + " |")

    if any(r.hybrid for r in rows):
        w("\n_**(WHOLE-CO)** = hybrid (INSW / TEN / CMBT) valued whole-company: whole-company "
          "normalized EPS ÷ whole-company NAV, with the lead-sleeve `g` (the watchlist sector "
          "tag). Value-weighted `g` is the intended v2 refinement._\n")
    w("\n_Flags: `non-positive NAV`, `no cost data`, `no anchor`, `r≤g (invalid)`, "
      "`newbuild-heavy (unreliable)` (newbuild value share > 25% — a not-yet-delivered hull "
      "earns a full anchor-year in the strip while its NAV is PV-haircut), `negative mid-cycle "
      "EPS`, `sub-growth returns` (RONAV_norm < g ⇒ P/NAV* unstable). Flagged rows carry no "
      "multiple. Names with a sub-threshold newbuild program (e.g. FRO ~17%) compute but carry "
      "a mild residual upward RONAV bias. Per-subsector `r` is a documented v2 extension._\n")

    gov_rows = [r for r in rows if r.governance_discount_pct > 0]
    if gov_rows:
        names = ", ".join(f"{r.ticker} ({r.governance_discount_pct:.0%})"
                          for r in sorted(gov_rows, key=lambda r: r.ticker))
        w("\n_**§15 governance dual-read.** " + names + " carry a realisation haircut applied at the "
          "blend layer + dividend-strip terminal but NOT in this leg, which uses CLEAN marked NAV by "
          "design. So the P/NAV* and Justified FV above are the **clean-NAV** reads; the **haircut "
          "basis** scales NAV — and Justified FV — by (1 − haircut) and lifts P/NAV(mkt) by the same "
          "factor (e.g. a 30% haircut ⇒ FV × 0.70, P/NAV(mkt) ÷ 0.70). Read these two names on both._\n")

    md_path = outputs_dir / "justified_pnav.md"
    md_path.write_text("\n".join(out))

    wb = Workbook()
    ws = wb.active
    ws.title = "Justified PNAV"
    headers = ["ticker", "basis", "sector", "nav_per_share", "price", "pnav_mkt",
               "r", "g", "ronav_implied",
               "ronav_parity", "justified_pnav_parity", "read_parity",
               "ronav_hist", "justified_pnav_hist", "read_hist",
               "robust", "flag"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    def _rnd(x, p=4):
        return round(x, p) if x is not None else None

    for r in sorted(rows, key=_sort_key):
        basis = "WHOLE-COMPANY (hybrid)" if r.hybrid else "whole-company"
        ws.append([
            r.ticker, basis, r.sector, round(r.nav_per_share, 3), round(r.price, 2),
            _rnd(r.pnav_mkt), round(r.r, 4), round(r.g, 4), _rnd(r.ronav_implied),
            _rnd(r.ronav_norm), _rnd(r.justified_pnav), r.read,
            _rnd(r.ronav_norm_hist), _rnd(r.justified_pnav_hist), r.read_hist,
            r.robust, (r.flag or ""),
        ])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 26)

    ws2 = wb.create_sheet("sensitivity")
    ws2.append(["sector", "base_ronav_norm", "ronav_norm", *[f"g={gg}" for gg in _GRID_G]])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for s in sorted(bases):
        base = bases[s]
        for off in _GRID_RONAV_OFFSETS:
            ronav = base + off
            cells = [
                (round(justified_pnav(ronav, r0, gg), 4) if r0 - gg > 0 else None)
                for gg in _GRID_G
            ]
            ws2.append([s, round(base, 4), round(ronav, 4), *cells])
    for col in ws2.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(width + 2, 18)

    wb.save(outputs_dir / "justified_pnav.xlsx")
    return md_path


def run_justified_pnav_xref(
    quarter: str, inputs_dir: Path = INPUTS_DIR, outputs_dir: Path = OUTPUTS_DIR,
    use_transaction_anchored: bool = True,
) -> list[JustifiedPnavRow]:
    """Compute, write, and print the justified-P/NAV diagnostic."""
    rows = compute_justified_pnav_rows(quarter, inputs_dir, use_transaction_anchored)
    if rows:
        path = write_justified_pnav(rows, outputs_dir)
        for r in sorted(rows, key=lambda x: (x.gap is None, -(x.gap or 0.0))):
            jp = f"{r.justified_pnav:.2f}x" if r.justified_pnav is not None else "—"
            pm = f"{r.pnav_mkt:.2f}x" if r.pnav_mkt is not None else "n/a"
            print(f"{r.ticker}: P/NAV mkt {pm} vs justified {jp} "
                  f"[RONAV_norm {_pct(r.ronav_norm)}] {r.read}")
        for s, m in sorted(subsector_median_pnav(rows).items(), key=lambda kv: kv[1], reverse=True):
            print(f"  median justified P/NAV [{s}] = {m:.2f}x")
        print(f"justified-P/NAV diagnostic -> {path}")
    return rows
