"""Scenario engine — consumes ``scenario_inputs.yaml`` (METHODOLOGY.md §2/§3/§11).

The file is **sector-layered** (METHODOLOGY §11): the top level holds a
``sectors`` map whose entries (``crude``, ``lng``, …) each carry their own
``scenarios`` block and ``cycle_anchors``. ``load_scenarios(sector=...)``
returns one sector's sub-doc; the pipeline resolves each ticker's sector from
its ``inputs/watchlist.yaml`` entry (default ``crude``).

Within a sector, each scenario carries a full per-class 8-quarter forward curve
as ``[low, base, high]`` $/day TCE cells, a probability ``weight``, and the
sector supplies per-class 10-year ``cycle_anchors``. For each scenario we:

- override the FFA forward curve (drives the dividend strip) with the scenario
  curve at the chosen point (low/base/high),
- set the cycle-position numerator to the scenario's **forward 12M base** strip
  (front-4-quarter mean) — per the file, cycle ratio = forward-12M-base / 10yr
  mean — so each scenario's NAV/earnings weighting reflects its own rate path,
- hold NAV scenario-invariant (vessel values don't reset in T+0),

then blend, and probability-weight across scenarios for the headline fair value.

Class map: model ``VLCC/Suezmax/Aframax/LR2/LNGC`` -> scenario
``vlcc/suezmax/aframax_dirty/lr2_clean/lng``. LR2 maps to the *clean* curve
(supersedes the v1 LR2-as-Aframax proxy on the earnings side). LNGC is only
valid against ``sectors.lng``; crude classes against ``sectors.crude``.
"""

from __future__ import annotations

from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Iterable, Optional

import yaml

from .blend import blend_fair_value
from .breakeven import implied_breakeven_tce
from .cycle import compute_cycle
from .dividend_strip import compute_dividend_strip
from .nav import compute_nav
from .report import OUTPUTS_DIR
from .schemas import CompanyInputs

SCENARIOS_PATH = Path(__file__).resolve().parents[2] / "inputs" / "scenario_inputs.yaml"

# Per-sector scenario class maps (METHODOLOGY §11.4 / §11.5). Each sector's map
# routes its vessel classes to the scenario sub-keys present in that sector's
# `scenarios.<scenario>.<key>` blocks. The PRODUCT-sleeve override (used for
# hybrid INSW carve-outs and pure-product names) is the only routing that
# diverges from the default — LR1 routes to aframax_dirty in crude context but
# to lr1_clean in product context, and MR is product-only.
SCENARIO_CLASS_MAP_BY_SECTOR: dict[str, dict[str, str]] = {
    "crude": {
        "VLCC": "vlcc",
        "Suezmax": "suezmax",
        "Aframax": "aframax_dirty",
        "LR2": "lr2_clean",         # crude-context: dual-use LR2 (FRO) uses crude clean
        "LR1": "aframax_dirty",     # dual-use LR1's crude portion trades dirty (Aframax)
    },
    "lng": {
        "LNGC": "lng",
        "MGC": "mgc",
    },
    "product": {                    # NEW 2026-06-01 — METHODOLOGY §11.5
        "MR": "mr",
        "Handysize": "mr",          # clean-product Handysize: v1 earnings proxy to MR
                                    # (added 2026-06-05; NAV differentiated via its own
                                    # value curve; refine to a dedicated handysize block in v2)
        "Handymax": "mr",           # chemical Handymax: TCE proxy to MR validated empirically
                                    # by STNG Q1+Q2 2026 disclosed rates (Q1 $34k vs MR $32k;
                                    # Q2 $32k vs MR $36.5k — chem discount lives on the value
                                    # side, not rates). Added 2026-06-05; NAV differentiated
                                    # via its own value curve (§11.5 / LIMITATIONS §2).
        "LR1": "lr1_clean",         # product LR1 has its own forwards (v1 proxy = lr2_clean)
        "LR2": "lr2_clean",
    },
    "dry_bulk": {                   # NEW 2026-06-09 — METHODOLOGY §11.7
        # Methodology class names per §11.7.1 (collapse Pareto's 6 sub-classes
        # into 3): Cape / Pana / Supra-Ultra. Match ALLOWED_CLASSES in loaders.py
        # and the fleet manifest `class:` field.
        "Cape": "cape",
        "Pana": "pana",
        "Supra-Ultra": "supra_ultra",  # collapsed Supramax + Ultramax per §11.7.1
                                        # (Pareto reclassified the benchmark Sep 2025)
    },
    "containerships": {             # NEW 2026-06-12 — METHODOLOGY §11.8.1
        # 3-class collapse aligned to MB Shipbrokers' rate-table banding:
        # Feeder ≤2,000 TEU / Intermediate 2,000-5,500 / Large >5,500.
        # WB (wide-beam) variants are not classes (design premium lives in
        # the vessel's own TC rate). This sector sets strip_horizon: 10
        # (q3_2026 → q4_2028, §11.8.6.4) — its curves carry 10 quarter keys.
        "Ctr-Feeder": "ctr_feeder",
        "Ctr-Intermediate": "ctr_intermediate",
        "Ctr-Large": "ctr_large",
    },
}

# Module-level default — combines all classes whose routing is unambiguous
# across sectors. Used by run_scenarios when no explicit scenario_class_map is
# passed. The PRODUCT sleeve overrides this via PRODUCT_SCENARIO_CLASS_MAP
# (LR1 routes differently; MR is product-only).
SCENARIO_CLASS_MAP: dict[str, str] = {
    **SCENARIO_CLASS_MAP_BY_SECTOR["crude"],   # VLCC / Suezmax / Aframax / LR2 / LR1 (dirty)
    **SCENARIO_CLASS_MAP_BY_SECTOR["lng"],     # LNGC / MGC
    # MR is intentionally NOT in the default — only valid in product context
    # and routed via PRODUCT_SCENARIO_CLASS_MAP.
    # dry_bulk classes (Cape / Pana / Supra-Ultra) ALSO not in default —
    # only valid in dry_bulk context; pipeline.py routes via the explicit
    # SCENARIO_CLASS_MAP_BY_SECTOR["dry_bulk"] map for dry_bulk tickers.
}

# Product-sleeve override map for hybrid carve-outs (e.g. INSW product sleeve,
# METHODOLOGY §6 v2) and pure-product names (METHODOLOGY §11.5).
# Pipeline passes this explicitly when running the product sleeve / sector.
PRODUCT_SCENARIO_CLASS_MAP: dict[str, str] = SCENARIO_CLASS_MAP_BY_SECTOR["product"]
QUARTER_KEYS = ["q3_2026", "q4_2026", "q1_2027", "q2_2027",
                "q3_2027", "q4_2027", "q1_2028", "q2_2028"]
_POINT_IDX = {"low": 0, "base": 1, "high": 2}


def quarter_keys(n: int) -> list[str]:
    """First ``n`` strip-quarter keys, extending the q3_2026 convention.

    Sectors with ``strip_horizon`` > 8 (containerships, §11.8.6.4) carry
    longer per-scenario curves; their YAML quarter keys continue the same
    sequence (q3_2028, q4_2028, ...).
    """
    q, y = 3, 2026
    keys = []
    for _ in range(n):
        keys.append(f"q{q}_{y}")
        q += 1
        if q == 5:
            q, y = 1, y + 1
    return keys
_REC_BAND = 0.05  # +/-5% expected-value band for the position call

# Vessel-value elasticity to the forward curve (Fix B). Forward expectations feed
# second-hand/newbuild prices, so NAV is NOT scenario-invariant: a scenario whose
# value-weighted 8-quarter forward is X% above the probability-weighted reference
# scales vessel values by ELASTICITY*X%. Calibrated (pending Clarksons Asset
# Index) so the pre-MoU-vs-bear vessel swing ~ the $95M-$135M modern-VLCC range
# implied by the framework. Reference = probability-weighted forward, so the
# weighted-average NAV stays anchored to the current Compass-based NAV.
VESSEL_VALUE_ELASTICITY = 0.5
# Clamp keeps a 5-yr vessel between ~scrap-discount and ~newbuild parity: a
# modern VLCC can't be worth more than a fresh resale even at extreme rates.
_VESSEL_SCALE_CLAMP = (0.65, 1.25)


@dataclass
class ScenarioFV:
    name: str
    weight: float
    fair_value: float          # base curve
    fair_value_low: float
    fair_value_high: float
    nav_per_share: float       # scenario-flexed NAV
    vessel_scale: float        # vessel-value multiplier applied this scenario
    divstrip_npv: float
    cycle_position: float
    w_nav: float
    assumed_tce: float         # the scenario's value-weighted 12M forward TCE (NOT a breakeven)


@dataclass
class ScenarioReport:
    ticker: str
    current_price: float
    analyst_target: float
    base_nav_per_share: float   # unflexed reference NAV
    breakeven_tce: float        # ONE scenario-invariant breakeven (rate to justify the price)
    scenarios: list[ScenarioFV]
    probability_weighted_fv: float
    upside_best: float           # best-scenario FV - price (escalation tail)
    downside_worst: float        # worst-scenario FV - price (bear)
    expected_value_vs_current: float  # weighted_FV - price
    position_recommendation: str
    # Explicit valuation basis (set for hybrid carve-outs) so every FV-vs-price
    # figure in the scenario report is unambiguously labeled. Empty for pure-plays.
    basis: str = ""
    # Sector layer (METHODOLOGY §11) — `crude` or `lng` for v1. Drives the
    # framework label in the markdown title and the per-sector roll-up sheet.
    sector: str = "crude"


def load_scenarios(path: Path = SCENARIOS_PATH, sector: str = "crude") -> dict:
    """Load the scenario block for one sector. Returns a dict shaped like
    ``{"sector", "scenarios", "cycle_anchors", ...}`` — the sector sub-doc
    from ``inputs/scenario_inputs.yaml`` (METHODOLOGY §11) with the sector
    name stamped in so ``run_scenarios`` can tag its report.

    Default ``sector="crude"`` returns the three-phase MoU framework that values
    every crude-tanker ticker plus the product sleeve of hybrids (MR forwards
    live alongside crude classes for v2; v3 may split into its own product
    sector). ``sector="lng"`` returns the LNG glut-cycle framework for
    FLNG-style names.
    """
    with open(path) as fh:
        doc = yaml.safe_load(fh)
    sectors = doc.get("sectors") or {}
    if sector not in sectors:
        raise KeyError(f"scenario sector {sector!r} not found in {path}; have {list(sectors)}")
    out = sectors[sector]
    out["sector"] = sector
    return out


# ---------------------------------------------------------------------------
# Cycle-anchor basis commensurability (B5 — METHODOLOGY §10)
# ---------------------------------------------------------------------------
# A name's cycle-position ratio is (forward-12M TC) / cycle-anchor. The anchors
# come in three bases that DO NOT numerically compose, so a ratio is only
# comparable across names whose anchors share a basis:
#   tc_10yr_mean       — crude / product / lng  (TC-anchored 10-year means)
#   archive_22mo_median — dry_bulk             (22-month Pareto archive median)
#   fy_calendar_avg    — containerships        (FY2021-2025 MB calendar average)
# The valuation core is unaffected (within-sector valuation is correct); the
# basis only matters when a cross-sector view (delta report / reconcile --all)
# lines ratios up side by side. `detect_mixed_anchor_basis` powers the
# MIXED-ANCHOR-BASIS flag on those surfaces.
ANCHOR_BASIS_LABELS = {
    "tc_10yr_mean": "TC-anchored 10-year mean",
    "archive_22mo_median": "22-month archive median",
    "fy_calendar_avg": "FY2021-2025 calendar average",
}


def all_sector_anchor_bases(path: Path = SCENARIOS_PATH) -> dict:
    """Map each sector to the basis token shared by its ``cycle_anchors`` blocks.

    A sector whose class blocks disagree maps to ``"MIXED"`` (a data bug that
    the tests catch); a sector with no tagged blocks is omitted.
    """
    with open(path) as fh:
        doc = yaml.safe_load(fh)
    out: dict[str, str] = {}
    for sector, sub in (doc.get("sectors") or {}).items():
        anchors = (sub or {}).get("cycle_anchors") or {}
        bases = {a.get("anchor_basis") for a in anchors.values() if isinstance(a, dict)}
        bases.discard(None)
        if len(bases) == 1:
            out[sector] = next(iter(bases))
        elif len(bases) > 1:
            out[sector] = "MIXED"
    return out


def detect_mixed_anchor_basis(
    sectors: Iterable[str],
    path: Path = SCENARIOS_PATH,
    basis_map: Optional[dict] = None,
) -> Optional[dict]:
    """Given the sectors present in a cross-sector view, return ``{basis:
    [sectors]}`` when more than one distinct anchor basis is present (cycle
    positions not numerically comparable across them — METHODOLOGY §10), else
    ``None``. Untagged sectors are ignored.
    """
    if basis_map is None:
        basis_map = all_sector_anchor_bases(path)
    present: dict[str, list[str]] = {}
    for s in sectors:
        b = basis_map.get(s)
        if b is None:
            continue
        present.setdefault(b, [])
        if s not in present[b]:
            present[b].append(s)
    return present if len(present) > 1 else None


def format_mixed_anchor_basis(mix: dict) -> str:
    """One-line human description of a `detect_mixed_anchor_basis` result, shared
    verbatim by the delta report and reconcile surfaces."""
    parts = []
    for basis, secs in sorted(mix.items()):
        label = ANCHOR_BASIS_LABELS.get(basis, basis)
        parts.append(f"`{basis}` ({label}: {', '.join(sorted(secs))})")
    return "; ".join(parts)


def _curve(scenario_cls: dict, point: str, keys: list[str] = None) -> list[float]:
    idx = _POINT_IDX[point]
    return [float(scenario_cls[q][idx]) for q in (keys or QUARTER_KEYS)]


def _override_market_data(
    md, scenario: dict, anchors: dict, classes: set[str], point: str,
    vessel_scale: float = 1.0, scenario_class_map: dict[str, str] = None,
    keys: list[str] = None,
):
    """Replace FFA / 12M-TC / 10yr-mean (and flex vessel values) for the fleet's classes."""
    if scenario_class_map is None:
        scenario_class_map = SCENARIO_CLASS_MAP
    ffa, tc, means = {}, {}, {}
    for cls in classes:
        scen_key = scenario_class_map[cls]
        cells = scenario[scen_key]
        ffa[cls] = _curve(cells, point, keys)
        tc[cls] = sum(_curve(cells, "base", keys)[:4]) / 4.0   # cycle always uses the base front-4
        means[cls] = float(anchors[scen_key]["ten_year_mean"])

    curves = md.vessel_value_curves
    if vessel_scale != 1.0:
        curves = {
            cls: replace(
                vc,
                newbuild=vc.newbuild * vessel_scale,
                five_year_benchmark=vc.five_year_benchmark * vessel_scale,
                ten_year_benchmark=vc.ten_year_benchmark * vessel_scale,
                scrap_25yr=vc.scrap_25yr * vessel_scale,
                scrubber_premium=vc.scrubber_premium * vessel_scale,
            )
            for cls, vc in curves.items()
        }
    return replace(
        md, ffa_forward_curve=ffa, twelve_month_tc=tc,
        historical_tce_means=means, vessel_value_curves=curves,
    )


def position_recommendation(ev_pct: float) -> str:
    """Standard convention: FV above price => undervalued => buy.

    NOTE: this is the economic convention (FV > price = cheap = BUY). It is the
    *inverse* of the literal labels in scenario_inputs.yaml output_requirements
    (which read 'EV > +5% = trim/short') — flagged for the user to confirm.
    """
    if ev_pct > _REC_BAND:
        return "BUY (undervalued)"
    if ev_pct < -_REC_BAND:
        return "TRIM/SHORT (overvalued)"
    return "HOLD (fairly valued)"


def run_scenarios(
    inputs: CompanyInputs,
    current_price: float,
    analyst_target: float,
    doc: dict,
    elasticity: float = VESSEL_VALUE_ELASTICITY,
    scenario_class_map: dict[str, str] = None,
) -> ScenarioReport:
    """Run every scenario for one name and probability-weight the fair value.

    ``scenario_class_map`` defaults to the crude-sleeve map; pass
    ``PRODUCT_SCENARIO_CLASS_MAP`` for product-sleeve runs (whole-co INSW v2).
    """
    if scenario_class_map is None:
        scenario_class_map = SCENARIO_CLASS_MAP
    classes = {v.cls for v in inputs.fleet.vessels}
    # Per-sector strip horizon (METHODOLOGY §11.8.6.4): default 8; sectors
    # with longer contracted visibility (containerships) set `strip_horizon`
    # in their scenario_inputs.yaml block and carry curves of that length.
    horizon = int(doc.get("strip_horizon", 8))
    qkeys = quarter_keys(horizon)
    base_nav = compute_nav(inputs)  # unflexed reference (value weights + anchoring)
    value_weights = {
        cls: base_nav.fleet_value_by_class[cls] / base_nav.fleet_value
        for cls in base_nav.fleet_value_by_class
    }

    # Value-weighted base forward per scenario: horizon avg drives the vessel-
    # value elasticity; 12-month (front-4) avg is the scenario's "assumed TCE".
    scen_forward: dict[str, float] = {}
    scen_forward_12m: dict[str, float] = {}
    for name, scen in doc["scenarios"].items():
        f8 = f4 = 0.0
        for cls in classes:
            curve = _curve(scen[scenario_class_map[cls]], "base", qkeys)
            w = value_weights.get(cls, 0.0)
            f8 += w * (sum(curve) / len(curve))
            f4 += w * (sum(curve[:4]) / 4.0)
        scen_forward[name] = f8
        scen_forward_12m[name] = f4
    total_w = sum(float(s["weight"]) for s in doc["scenarios"].values())

    # Reference forward for the vessel-value elasticity = the CURRENT trader
    # forward (value-weighted 8q), a FIXED anchor: vessel multiplier = 1.0 at
    # today's market (where current Compass values were struck). Fixing it (vs a
    # probability-weighted reference) decouples the base-case NAV from the
    # scenario set, so adding the escalation tail doesn't move the base NAV.
    cur = inputs.market_data.ffa_forward_curve
    forward_ref = sum(
        value_weights.get(cls, 0.0) * (sum(cur[cls]) / len(cur[cls]))
        for cls in classes if cls in cur
    )

    # The breakeven TCE is scenario-INVARIANT: the value-weighted blended rate that
    # makes the blended FV equal the current price under the base (current-market)
    # cycle weight and unflexed NAV. One number per name (same value-weighted basis
    # as each scenario's "assumed TCE", so the assumed/breakeven ratio is like-for-
    # like); the scenario only changes the PROBABILITY of clearing it, not the level.
    base_be = implied_breakeven_tce(inputs, current_price)
    breakeven_tce = base_be.blended_breakeven_tce

    results: list[ScenarioFV] = []
    fv_by_name: dict[str, float] = {}
    for name, scen in doc["scenarios"].items():
        weight = float(scen["weight"])
        lo, hi = _VESSEL_SCALE_CLAMP
        vessel_scale = min(hi, max(lo, 1.0 + elasticity * (scen_forward[name] / forward_ref - 1.0)))
        # Optional per-scenario structural multiplier (e.g. LNG sectors.lng's
        # structural_reset applies an additional -10% via vessel_scale_multiplier
        # = 0.90 to reflect accelerated retirement under an energy-transition
        # tail; the field is absent on all standard cycle-rate scenarios so this
        # is a no-op for them). Reclamped to the standard band.
        vessel_scale = min(hi, max(lo, vessel_scale * float(scen.get("vessel_scale_multiplier", 1.0))))

        def build(point: str):
            return replace(
                inputs,
                market_data=_override_market_data(
                    inputs.market_data, scen, doc["cycle_anchors"], classes, point,
                    vessel_scale, scenario_class_map, qkeys,
                ),
            )

        ci_base = build("base")
        nav_s = compute_nav(ci_base)            # scenario-flexed NAV
        cyc = compute_cycle(ci_base)            # cycle weight from the scenario's forward 12M

        def fv_at(point: str) -> tuple[float, float]:
            strip = compute_dividend_strip(
                build(point), nav_s.nav_per_share, strip_horizon=horizon,
                terminal_multiple=cyc.terminal_multiple,
            )
            return blend_fair_value(nav_s, strip, cyc).fair_value_per_share, strip.implied_price

        fv_base, strip_npv = fv_at("base")
        fv_low, _ = fv_at("low")
        fv_high, _ = fv_at("high")

        results.append(ScenarioFV(
            name=name, weight=weight, fair_value=fv_base,
            fair_value_low=fv_low, fair_value_high=fv_high,
            nav_per_share=nav_s.nav_per_share, vessel_scale=vessel_scale,
            divstrip_npv=strip_npv, cycle_position=cyc.cycle_position,
            w_nav=cyc.w_nav, assumed_tce=scen_forward_12m[name],
        ))
        fv_by_name[name] = fv_base

    weighted = sum(r.weight * r.fair_value for r in results) / total_w
    upside = max(r.fair_value for r in results) - current_price     # best scenario (escalation)
    downside = min(r.fair_value for r in results) - current_price   # worst scenario (bear)
    ev = weighted - current_price

    return ScenarioReport(
        ticker=inputs.fleet.ticker,
        current_price=current_price,
        analyst_target=analyst_target,
        base_nav_per_share=base_nav.nav_per_share,
        breakeven_tce=breakeven_tce,
        scenarios=results,
        probability_weighted_fv=weighted,
        upside_best=upside,
        downside_worst=downside,
        expected_value_vs_current=ev,
        position_recommendation=position_recommendation(ev / current_price),
        sector=str(doc.get("sector") or "crude"),
    )


# --------------------------------------------------------------------------- #
# rendering
# --------------------------------------------------------------------------- #
_PRETTY = {
    # crude (three-phase MoU)
    "escalation": "Escalation",
    "pre_mou_baseline": "Pre-MoU baseline",
    "mou_base": "MoU base case",
    "mou_bear": "MoU bear",
    # lng (glut-cycle)
    "tight_resurgence": "Tight resurgence",
    "moderate_tightening": "Moderate tightening",
    "glut_base": "Glut base case",
    "glut_intensifies": "Glut intensifies",
    "structural_reset": "Structural reset",
    # dry_bulk (Bulk Set A — China-driven, METHODOLOGY §11.7.4)
    "china_acceleration": "China acceleration",
    "moderate_growth": "Moderate growth (base)",
    "china_property_drag": "China property drag",
    "coordinated_slowdown": "Coordinated slowdown",
    # containerships (Container Set A — disruption-led, METHODOLOGY §11.8.4)
    "disruption_persists": "Disruption persists",
    "gradual_normalization": "Gradual normalization (base)",
    "normalization_plus_overhang": "Normalization + orderbook overhang",
    "demand_recession": "Demand recession",
}

_SECTOR_FRAMEWORK_LABEL = {
    "crude": "three-phase MoU framework",
    "lng": "LNG glut-cycle framework",
    "product": "product margin / glut framework",
    "dry_bulk": "Bulk Set A (China-driven)",
    "containerships": "Container Set A (disruption-led)",
}


def _pct(a: float, b: float) -> float:
    return (a / b - 1.0) * 100.0 if b else float("nan")


def render_scenario_markdown(r: ScenarioReport) -> str:
    out: list[str] = []
    w = out.append
    is_whole_co = r.basis.startswith("WHOLE-COMPANY")
    is_crude_sleeve = r.basis.startswith("CRUDE SLEEVE")
    if is_whole_co:
        title_tag = " [WHOLE-CO]"
    elif is_crude_sleeve:
        title_tag = " [CRUDE SLEEVE]"
    else:
        title_tag = ""
    framework_label = _SECTOR_FRAMEWORK_LABEL.get(r.sector, "scenario framework")
    w(f"# {r.ticker}{title_tag} — Scenario Fair Value ({framework_label})\n")
    if r.basis:
        w(f"> **Valuation basis:** {r.basis}\n")
    # Only the crude-sleeve mode displays a CARVED price; whole-co and pure-play
    # both compare against the actual tape price -> no qualifier.
    price_label = "Current price (crude-allocated)" if is_crude_sleeve else "Current price"
    target_label = "Analyst target (crude-allocated)" if is_crude_sleeve else "Analyst target"
    w(f"- **{price_label}:** ${r.current_price:,.2f}")
    w(f"- **{target_label}:** ${r.analyst_target:,.2f}")
    w(f"- **NAV / share (reference, unflexed):** ${r.base_nav_per_share:,.2f} "
      f"_(flexes per scenario via vessel-value elasticity — see table)_")
    w(f"- **Probability-weighted fair value:** ${r.probability_weighted_fv:,.2f} "
      f"({_pct(r.probability_weighted_fv, r.current_price):+.1f}% vs price)")
    w(f"- **Breakeven TCE (scenario-invariant):** ${r.breakeven_tce:,.0f}/day — the value-"
      f"weighted blended rate (fleet-mix-adjusted) that justifies the current price. The "
      f"scenario sets the *probability* of clearing it, not the level.")
    w(f"- **Position (tool view):** {r.position_recommendation}\n")

    w("## Per-scenario fair value\n")
    w("| Scenario | Weight | Vessel× | NAV/sh | FV (base) | FV [low–high] | Cycle | w_nav | Strip NPV | Assumed TCE (12M) | Assumed / Breakeven |")
    w("|---|--:|--:|--:|--:|--:|--:|--:|--:|--:|--:|")
    for s in r.scenarios:
        ratio = s.assumed_tce / r.breakeven_tce if r.breakeven_tce else float("nan")
        w(f"| {_PRETTY.get(s.name, s.name)} | {s.weight:.0%} | {s.vessel_scale:.2f}× | "
          f"${s.nav_per_share:,.2f} | ${s.fair_value:,.2f} | "
          f"${s.fair_value_low:,.2f}–${s.fair_value_high:,.2f} | {s.cycle_position:.2f}× | "
          f"{s.w_nav:.2f} | ${s.divstrip_npv:,.2f} | ${s.assumed_tce:,.0f} | {ratio:.2f}× |")
    w(f"| **Probability-weighted** | | | | **${r.probability_weighted_fv:,.2f}** | | | | | | |\n")
    w("_Assumed TCE = the scenario's value-weighted 12-month forward (the model's rate "
      "assumption, NOT a breakeven). Assumed/Breakeven < 1 ⇒ that scenario's rates fall "
      "short of justifying the price; > 1 ⇒ they clear it._\n")

    w("## Decision signals\n")
    w(f"- **Upside (best scenario − price):** ${r.upside_best:+,.2f}")
    w(f"- **Downside (worst scenario − price):** ${r.downside_worst:+,.2f}")
    w(f"- **Expected value vs current** (weighted FV − price): ${r.expected_value_vs_current:+,.2f} "
      f"({_pct(r.probability_weighted_fv, r.current_price):+.1f}%)")
    w(f"- **Position:** {r.position_recommendation}\n")
    w("_Convention: FV above price = undervalued = BUY; FV below = overvalued = TRIM/SHORT. "
      "(This is the inverse of the literal buy/trim labels in scenario_inputs.yaml "
      "output_requirements.highlight — flagged for confirmation.)_\n")
    return "\n".join(out)


def write_scenario_report(r: ScenarioReport, outputs_dir: Path = OUTPUTS_DIR) -> Path:
    outputs_dir.mkdir(parents=True, exist_ok=True)
    path = outputs_dir / f"{r.ticker.lower()}_scenarios.md"
    path.write_text(render_scenario_markdown(r))
    return path


def write_scenario_summary(
    reports: list[ScenarioReport], outputs_dir: Path = OUTPUTS_DIR
) -> Path:
    """Scenario roll-up — one sheet per sector + a Pair-trades sheet.

    Per scenario_inputs.yaml output_requirements.watchlist_rollup; sector layer
    per METHODOLOGY §11. Each sector's scenarios drive its own FV columns. The
    crude sheet stays named ``Scenario summary`` and is the active sheet, so
    crude-tanker tooling and tests don't shift positions.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    outputs_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # we'll create explicit sheets per sector below

    # Group reports by sector, preserving each sector's scenario order from the
    # first report we see (a sector's scenario list is identical across tickers).
    by_sector: dict[str, list[ScenarioReport]] = {}
    sector_scenarios: dict[str, list[str]] = {}
    for r in reports:
        by_sector.setdefault(r.sector, []).append(r)
        sector_scenarios.setdefault(r.sector, [s.name for s in r.scenarios])

    # Crude comes first so it's the active sheet (preserves existing pipeline
    # consumers + the test_carveout INSW assertion on the active sheet).
    sector_order = (["crude"] if "crude" in by_sector else []) + [
        s for s in by_sector if s != "crude"
    ]

    for sector in sector_order:
        # Excel caps sheet names at 31 chars — "CONTAINERSHIPS" needs a tag.
        tag = "CTR" if sector == "containerships" else sector.upper()
        sheet_title = "Scenario summary" if sector == "crude" else f"Scenario summary ({tag})"
        ws = wb.create_sheet(title=sheet_title)
        scen_names = sector_scenarios[sector]
        scen_cols = [f"{name}_FV" for name in scen_names]
        # LNG sheet gets two extra columns (METHODOLOGY §13, §6 CCEC entry):
        #   FV_range  = max(scenario FV) - min(scenario FV) — total dispersion
        #   torque_5pp = 5pp weight shift from glut_intensifies → tight_resurgence
        #                applied to PW FV (a small constructive-tilt sensitivity)
        # The pair classifies LNG names as low-torque (FLNG-style TC-heavy book)
        # vs high-torque (CCEC-style NB-heavy book) so future LNG onboardings
        # surface the weight-driven-vs-weight-robust classification immediately.
        include_torque = (sector == "lng")
        torque_cols = ["FV_range", "torque_5pp"] if include_torque else []
        headers = [
            "ticker", "basis", "sector", "current_price",
            *scen_cols,
            "probability_weighted_FV", "upside_pct", "downside_pct",
            "expected_value_pct", "position_recommendation",
            *torque_cols,
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)

        ev_col = 4 + len(scen_cols) + 3  # 0-based index of expected_value_pct
        rows = []
        for r in by_sector[sector]:
            fv = {s.name: s.fair_value for s in r.scenarios}
            if r.basis.startswith("WHOLE-COMPANY"):
                basis = "WHOLE-COMPANY (hybrid aggregation)"
            elif r.basis.startswith("CRUDE SLEEVE"):
                basis = "CRUDE SLEEVE (allocated price)"
            else:
                basis = "whole-company"
            row = [
                r.ticker, basis, r.sector, r.current_price,
                *[round(fv.get(name, float("nan")), 2) for name in scen_names],
                round(r.probability_weighted_fv, 2),
                round(_pct(r.current_price + r.upside_best, r.current_price), 1),
                round(_pct(r.current_price + r.downside_worst, r.current_price), 1),
                round(_pct(r.probability_weighted_fv, r.current_price), 1),
                r.position_recommendation,
            ]
            if include_torque:
                fv_vals = [s.fair_value for s in r.scenarios]
                fv_range = round(max(fv_vals) - min(fv_vals), 2)
                # 5pp shift from glut_intensifies → tight_resurgence; if either
                # scenario is missing fall through to NaN so the cell renders blank.
                fv_tight = fv.get("tight_resurgence")
                fv_bear = fv.get("glut_intensifies")
                torque = round(0.05 * (fv_tight - fv_bear), 2) if (
                    fv_tight is not None and fv_bear is not None
                ) else None
                row.extend([fv_range, torque])
            rows.append(row)
        rows.sort(key=lambda row: row[ev_col], reverse=True)
        for row in rows:
            ws.append(row)

        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    wb.active = 0

    # Pair-trade sheet: relative implied return of long A / short B if both
    # converge to their probability-weighted FV (= EV%_long - EV%_short).
    # Cross-sector pairs are flagged (different rate cycles — judgement call).
    wsp = wb.create_sheet("Pair trades")
    wsp.append(["Long", "Short", "Long EV %", "Short EV %", "Relative return %", "Comparable?"])
    for c in wsp[1]:
        c.font = Font(bold=True)
    def _label(r):
        if r.basis.startswith("WHOLE-COMPANY"):
            return f"{r.ticker} (WHOLE-CO)"
        if r.basis.startswith("CRUDE SLEEVE"):
            return f"{r.ticker} (CRUDE SLEEVE)"
        return r.ticker
    by_ticker = {r.ticker: r for r in reports}
    ev = {r.ticker: _pct(r.probability_weighted_fv, r.current_price) for r in reports}
    pairs = []
    for long_t in ev:
        for short_t in ev:
            if long_t == short_t or ev[long_t] - ev[short_t] <= 0:
                continue
            long_r, short_r = by_ticker[long_t], by_ticker[short_t]
            tags = []
            if long_r.sector != short_r.sector:
                tags.append(f"MIXED SECTOR ({long_r.sector} vs {short_r.sector})")
            if bool(long_r.basis) ^ bool(short_r.basis):
                tags.append("MIXED BASIS (crude sleeve vs whole-co)")
            comparable = "; ".join(tags) if tags else "yes"
            pairs.append([_label(long_r), _label(short_r),
                          round(ev[long_t], 1), round(ev[short_t], 1),
                          round(ev[long_t] - ev[short_t], 1), comparable])
    pairs.sort(key=lambda p: p[4], reverse=True)
    for p in pairs:
        wsp.append(p)
    for col in wsp.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        wsp.column_dimensions[col[0].column_letter].width = min(width + 2, 28)

    path = outputs_dir / "scenario_summary.xlsx"
    wb.save(path)
    return path
