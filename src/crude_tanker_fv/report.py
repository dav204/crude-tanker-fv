"""Output formatting (METHODOLOGY.md section 7).

Two outputs:

- Per-company report: ``outputs/{ticker}_fv_report.md`` plus
  ``outputs/{ticker}_fv_report.xlsx``. Sections: header (ticker, date, price,
  model FV, analyst target), NAV breakdown, dividend-strip table, cycle
  weighting, blended fair value, implied breakeven TCE, 5x5 sensitivity grid,
  and a divergence diagnosis.
- Watchlist roll-up: ``outputs/fair_value_summary.xlsx`` — one sortable row per
  ticker (Current, Tool FV, Watchlist Target, deltas, implied breakeven, cycle).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .blend import BlendedFairValue
from .breakeven import BreakevenResult
from .cycle import CycleResult
from .dividend_strip import DividendStripResult
from .nav import NavResult
from .sensitivity import SensitivityGrid

OUTPUTS_DIR = Path(__file__).resolve().parents[2] / "outputs"


@dataclass
class CompanyReport:
    """Everything needed to render one company's report (section 7)."""

    ticker: str
    report_date: str
    current_price: float
    analyst_target: float
    nav: NavResult
    strip: DividendStripResult
    cycle: CycleResult
    blended: BlendedFairValue
    breakeven: BreakevenResult
    sensitivity: SensitivityGrid
    warnings: list[str] = field(default_factory=list)
    payout_sensitivity: dict[float, float] = field(default_factory=dict)
    notes: list[str] = field(default_factory=list)
    # Explicit valuation basis (which FV vs which price) — set for hybrid carve-outs
    # so every FV-vs-price figure in the report is unambiguous (crude sleeve vs
    # crude-allocated price). Empty for pure-play names (whole-company vs price).
    basis: str = ""


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def _primary_class(report: CompanyReport) -> str:
    """The class carrying the most fleet value (drives the breakeven readout)."""
    by_class = report.nav.fleet_value_by_class
    return max(by_class, key=by_class.get) if by_class else next(iter(report.breakeven.implied_breakeven_tce), "VLCC")


def _pct(a: float, b: float) -> float:
    return (a / b - 1.0) * 100.0 if b else float("nan")


_NAV_COVERS_THRESHOLD = 0.05   # multiplier below this ⇒ NAV alone covers price


def _diagnosis(report: CompanyReport) -> str:
    fv = report.blended.fair_value_per_share
    cur, tgt = report.current_price, report.analyst_target
    be = report.breakeven.blended_breakeven_tce
    mean = report.breakeven.blended_mean_tce
    ffa = report.breakeven.blended_ffa_12m
    mult = report.breakeven.multiplier

    lines = [
        f"Tool fair value **${fv:,.2f}** is {_pct(fv, cur):+.1f}% vs the current "
        f"price (${cur:,.2f}) and {_pct(fv, tgt):+.1f}% vs the analyst target (${tgt:,.2f}).",
    ]
    if abs(_pct(fv, cur)) <= 5 and abs(_pct(fv, tgt)) <= 5:
        lines.append("Tool, market, and analyst are in broad agreement (all within ~5%).")

    if mult < _NAV_COVERS_THRESHOLD:
        lines.append(
            f"NAV alone covers the price (NAV/sh ${report.nav.nav_per_share:,.2f} ≥ "
            f"${cur:,.2f}); the dividend strip provides no extra hurdle, so the implied "
            f"breakeven floor is effectively zero — the market is pricing the fleet at a "
            f"discount to vessel value."
        )
    elif be and mean and ffa:
        peak_or_distress = "extended peak rates" if be > mean else "distress"
        bull = "below" if mult < 1.0 else "above"
        lines.append(
            f"The current price implies the fleet earning a value-weighted blended "
            f"**${be:,.0f}/day** ({mult:.2f}× the current forward) — "
            f"{be / mean:.1f}× the value-weighted 10-yr mean (${mean:,.0f}, i.e. the market is "
            f"pricing {peak_or_distress}), and the market is {bull} the forward curve."
        )
    return " ".join(lines)


def _strip_quarter_labels(n: int) -> list[str]:
    return [f"Q{i + 1}" for i in range(n)]


# --------------------------------------------------------------------------- #
# markdown
# --------------------------------------------------------------------------- #
def _render_markdown(report: CompanyReport) -> str:
    nav, strip, cyc, fv, be, sens = (
        report.nav, report.strip, report.cycle, report.blended,
        report.breakeven, report.sensitivity,
    )
    M = 1_000_000
    cls = _primary_class(report)
    out: list[str] = []
    w = out.append

    w(f"# {report.ticker} — Fair Value Report\n")
    if report.basis:
        w(f"> **Valuation basis:** {report.basis}\n")
    w(f"- **Report date:** {report.report_date}")
    price_label = "Current price (crude-allocated)" if report.basis else "Current price"
    w(f"- **{price_label}:** ${report.current_price:,.2f}")
    w(f"- **Model fair value:** ${fv.fair_value_per_share:,.2f}")
    target_label = "Analyst target (crude-allocated)" if report.basis else "Analyst target"
    w(f"- **{target_label}:** ${report.analyst_target:,.2f}\n")

    if report.warnings:
        w("## Data validation warnings\n")
        for warning in report.warnings:
            w(f"- {warning}")
        w("")

    w("## NAV breakdown\n")
    w("| Item | $M |")
    w("|---|---:|")
    for cls_name, val in nav.fleet_value_by_class.items():
        w(f"| Fleet value — {cls_name} | {val / M:,.1f} |")
    w(f"| + Cash & equivalents | {nav.cash_and_equivalents / M:,.1f} |")
    w(f"| + Working capital (net) | {nav.working_capital_net / M:,.1f} |")
    w(f"| − Total debt | {nav.total_debt / M:,.1f} |")
    w(f"| − Lease liabilities | {nav.lease_liabilities / M:,.1f} |")
    w(f"| − Newbuild commitments | {nav.newbuild_capex_commitments / M:,.1f} |")
    w(f"| + Newbuild advances | {nav.newbuild_advances_paid / M:,.1f} |")
    w(f"| **= NAV total** | **{nav.nav_total / M:,.1f}** |")
    w(f"| Diluted shares | {nav.diluted_shares_outstanding:,.0f} |")
    w(f"| **NAV / share** | **${nav.nav_per_share:,.2f}** |")
    if abs(nav.nav_per_share_ex_yard_discount - nav.nav_per_share) > 0.005:
        impact = nav.nav_per_share - nav.nav_per_share_ex_yard_discount
        w(f"| NAV / share (ex yard discount) | ${nav.nav_per_share_ex_yard_discount:,.2f} |")
        w(f"| Yard-discount impact / share | ${impact:,.2f} |")
    w("")

    ffa_12m = strip.ffa_12m_by_class.get(cls)
    tc_12m = cyc.twelve_month_tc_by_class.get(cls)
    mean = be.historical_mean_tce.get(cls)
    w(f"## Dividend strip (r = {strip.discount_rate:.0%})\n")
    w(f"| Quarter | FFA spot ({cls}, $/day) | Blended TCE ($/day) | EPS | DPS | Disc. DPS |")
    w("|---|---:|---:|---:|---:|---:|")
    labels = _strip_quarter_labels(len(strip.dps_by_quarter))
    tce_series = strip.blended_tce_by_quarter.get(cls, [None] * len(labels))
    ffa_series = strip.ffa_spot_by_quarter.get(cls, [None] * len(labels))
    for i, lab in enumerate(labels):
        tce_s = f"{tce_series[i]:,.0f}" if tce_series[i] is not None else "—"
        ffa_s = f"{ffa_series[i]:,.0f}" if ffa_series[i] is not None else "—"
        w(f"| {lab} | {ffa_s} | {tce_s} | {strip.eps_by_quarter[i]:.3f} | "
          f"{strip.dps_by_quarter[i]:.3f} | {strip.discounted_dps[i]:.3f} |")
    w(f"| Σ discounted DPS | | | | | {sum(strip.discounted_dps):.2f} |")
    w(f"| Terminal value (NAV, q9) | | | | {strip.terminal_value:.2f} | "
      f"{strip.discounted_terminal_value:.2f} |")
    w(f"| **DivStrip implied price** | | | | | **${strip.implied_price:,.2f}** |")
    if ffa_12m is not None:
        note = (f"\n_FFA spot is the {cls} forward curve that drives the strip cash flows; "
                f"its 12-month average is **${ffa_12m:,.0f}/day**. Blended TCE is that spot "
                f"dampened by charter coverage.")
        if tc_12m is not None:
            note += (f" Cycle weighting (below) uses a different, more conservative input — "
                     f"the 12-month TC of **${tc_12m:,.0f}/day** — not this FFA average.")
        w(note + "_\n")

    w("## Cycle weighting\n")
    if tc_12m is not None and mean:
        w(f"- Cycle position = 12M TC (Compass) ${tc_12m:,.0f} / 10-yr mean ${mean:,.0f} = "
          f"**{cyc.cycle_position:.2f}×** → **{cyc.band_label}**")
    else:
        w(f"- Cycle position (12M TC / 10yr mean): **{cyc.cycle_position:.2f}×** → "
          f"**{cyc.band_label}**")
    w(f"- Weights: w_nav = {cyc.w_nav:.2f}, w_earn = {cyc.w_earn:.2f}\n")

    w("## Blended fair value\n")
    # BUG-4 (2026-06-22): show the EFFECTIVE (post-§15-haircut) NAV that actually enters
    # the blend, so the line foots for §15 names (raw nav.nav_per_share did not).
    w(f"{cyc.w_nav:.2f} × ${fv.nav_per_share_effective:,.2f} (NAV) + "
      f"{cyc.w_earn:.2f} × ${strip.implied_price:,.2f} (strip) = "
      f"**${fv.fair_value_per_share:,.2f}**\n")

    if report.payout_sensitivity:
        w("## Payout sensitivity\n")
        w("| Dividend payout | Fair value |")
        w("|---|---:|")
        for payout in sorted(report.payout_sensitivity):
            w(f"| {payout:.0%} | ${report.payout_sensitivity[payout]:,.2f} |")
        w("\n_80% = stated-floor / discipline-reasserts; ~95% = base (recent peak behaviour "
          "with some conservatism); 100% = peak persists._\n")

    w("## Implied breakeven TCE\n")
    be_val = be.blended_breakeven_tce
    if be.multiplier < _NAV_COVERS_THRESHOLD:
        w(f"**NAV alone covers the price.** NAV/share **${nav.nav_per_share:,.2f}** ≥ price "
          f"**${report.current_price:,.2f}** at base cycle weighting, so the strip provides no "
          f"extra hurdle — the implied breakeven floor is effectively zero (rates could fall "
          f"to ~0 and the price would still be justified by vessel value alone). The market is "
          f"pricing the fleet at a discount to NAV.\n")
    else:
        w(f"The current price requires the fleet to run at **{be.multiplier:.2f}× the current "
          f"forward curve** (inter-class rate ratios preserved). Headline is the value-weighted "
          f"blended TCE across the fleet; per-class detail below.\n")
    w("| Benchmark (value-weighted blended) | $/day | vs breakeven |")
    w("|---|---:|---:|")
    w(f"| **Implied breakeven (blended)** | **{be_val:,.0f}** | — |")
    if be.blended_mean_tce:
        w(f"| 10-year mean | {be.blended_mean_tce:,.0f} | {be_val / be.blended_mean_tce:.2f}× |")
    if be.blended_ffa_12m:
        w(f"| 12-month FFA | {be.blended_ffa_12m:,.0f} | {be_val / be.blended_ffa_12m:.2f}× |")
    if be.blended_spot_tce:
        w(f"| Current spot | {be.blended_spot_tce:,.0f} | {be_val / be.blended_spot_tce:.2f}× |")
    w("")
    if len(be.implied_breakeven_tce) > 1:
        w("| Per-class implied breakeven | $/day | × its 10-yr mean |")
        w("|---|---:|---:|")
        for c in sorted(be.implied_breakeven_tce, key=lambda x: -be.value_weights.get(x, 0.0)):
            rate = be.implied_breakeven_tce[c]
            m = be.historical_mean_tce.get(c)
            mult_txt = f"{rate / m:.2f}×" if m else "—"
            w(f"| {c} ({be.value_weights.get(c, 0.0):.0%} of fleet value) | {rate:,.0f} | {mult_txt} |")
        w("")

    w("## Sensitivity — fair value (rows: TCE shock, cols: vessel-value shock)\n")
    header = "| TCE \\ Vessel | " + " | ".join(f"{v:+.0%}" for v in sens.vessel_value_shocks) + " |"
    w(header)
    w("|---|" + "---:|" * len(sens.vessel_value_shocks))
    for i, ts in enumerate(sens.tce_shocks):
        row = " | ".join(f"${sens.grid[i][j]:,.2f}" for j in range(len(sens.vessel_value_shocks)))
        w(f"| **{ts:+.0%}** | {row} |")
    w(f"\n_Current price ${sens.current_price:,.2f}. Cycle weights held at base across the grid._\n")

    w("## Divergence diagnosis\n")
    w(_diagnosis(report))
    w("")

    if report.notes:
        w("## Modeling notes\n")
        for note in report.notes:
            w(f"- {note}")
        w("")

    # Per-ticker sidecar-diagnostic discovery (METHODOLOGY §9.10, §13).
    # Any file at outputs/{ticker.lower()}_*_diagnostic.md is auto-linked here,
    # so per-name deep-dive analyses (CCEC BUY-actionability, future weight-
    # sensitivity reviews, transaction-anchor case studies) surface in the
    # FV report without coupling the report renderer to specific tickers.
    sidecar_links = _discover_sidecar_diagnostics(report.ticker)
    if sidecar_links:
        w("## Additional diagnostics\n")
        for link in sidecar_links:
            w(f"- {link}")
        w("")
    return "\n".join(out)


def _discover_sidecar_diagnostics(ticker: str, outputs_dir: Path = OUTPUTS_DIR) -> list[str]:
    """Return relative-link bullets for any `{ticker.lower()}_*_diagnostic.md`
    sidecar file in outputs/. Sorted for deterministic output.
    """
    if not outputs_dir.exists():
        return []
    prefix = ticker.lower() + "_"
    suffix = "_diagnostic.md"
    out: list[str] = []
    for p in sorted(outputs_dir.glob(f"{prefix}*{suffix}")):
        # Pretty-name the diagnostic from the filename:
        #   ccec_buy_diagnostic.md → "CCEC — BUY-actionability diagnostic"
        # Falls back to the bare filename if the naming pattern doesn't match.
        stem = p.stem.removeprefix(prefix).removesuffix("_diagnostic")
        if stem:
            label = f"{ticker} — {stem.replace('_', ' ')}-actionability diagnostic"
            out.append(f"[`{p.name}`]({p.name}) — {label}")
        else:
            out.append(f"[`{p.name}`]({p.name})")
    return out


# --------------------------------------------------------------------------- #
# xlsx
# --------------------------------------------------------------------------- #
def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 40)


def _render_xlsx(report: CompanyReport, path: Path) -> None:
    nav, strip, cyc, fv, be, sens = (
        report.nav, report.strip, report.cycle, report.blended,
        report.breakeven, report.sensitivity,
    )
    M = 1_000_000
    cls = _primary_class(report)
    bold = Font(bold=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    basis_label = "CRUDE SLEEVE (allocated price)" if report.basis else "whole-company"
    summary = [
        ("Ticker", report.ticker),
        ("Valuation basis", basis_label),
        ("Report date", report.report_date),
        ("Current price" + (" (crude-allocated)" if report.basis else ""), report.current_price),
        ("Model fair value", fv.fair_value_per_share),
        ("Analyst target" + (" (crude-allocated)" if report.basis else ""), report.analyst_target),
        ("NAV / share", nav.nav_per_share),
        ("DivStrip implied price", strip.implied_price),
        (f"12M TC cycle input {cls} ($/day)", cyc.twelve_month_tc_by_class.get(cls)),
        (f"12M FFA strip {cls} ($/day)", strip.ffa_12m_by_class.get(cls)),
        ("Cycle position", cyc.cycle_position),
        ("Cycle band", cyc.band_label),
        ("w_nav", cyc.w_nav),
        ("w_earn", cyc.w_earn),
        ("Implied breakeven blended ($/day)", be.blended_breakeven_tce),
        ("Breakeven multiplier (× forward)", be.multiplier),
    ]
    for r, (k, v) in enumerate(summary, 1):
        ws.cell(r, 1, k).font = bold
        ws.cell(r, 2, v)
    _autosize(ws)

    ws = wb.create_sheet("NAV")
    ws.append(["Item", "$M"])
    for c in ws[1]:
        c.font = bold
    rows = [(f"Fleet value — {k}", v / M) for k, v in nav.fleet_value_by_class.items()]
    rows += [
        ("+ Cash & equivalents", nav.cash_and_equivalents / M),
        ("+ Working capital (net)", nav.working_capital_net / M),
        ("- Total debt", -nav.total_debt / M),
        ("- Lease liabilities", -nav.lease_liabilities / M),
        ("- Newbuild commitments", -nav.newbuild_capex_commitments / M),
        ("+ Newbuild advances", nav.newbuild_advances_paid / M),
        ("= NAV total", nav.nav_total / M),
        ("Diluted shares", nav.diluted_shares_outstanding),
        ("NAV / share", nav.nav_per_share),
        ("NAV / share (ex yard discount)", nav.nav_per_share_ex_yard_discount),
        ("Yard-discount impact / share", nav.nav_per_share - nav.nav_per_share_ex_yard_discount),
    ]
    for k, v in rows:
        ws.append([k, v])
    _autosize(ws)

    ws = wb.create_sheet("DivStrip")
    ws.append(["Quarter", f"FFA spot {cls} ($/day)", "Blended TCE ($/day)",
               "EPS", "DPS", "Disc. DPS"])
    for c in ws[1]:
        c.font = bold
    labels = _strip_quarter_labels(len(strip.dps_by_quarter))
    tce_series = strip.blended_tce_by_quarter.get(cls, [None] * len(labels))
    ffa_series = strip.ffa_spot_by_quarter.get(cls, [None] * len(labels))
    for i, lab in enumerate(labels):
        ws.append([lab, ffa_series[i], tce_series[i], strip.eps_by_quarter[i],
                   strip.dps_by_quarter[i], strip.discounted_dps[i]])
    ws.append(["Sum disc. DPS", None, None, None, None, sum(strip.discounted_dps)])
    ws.append(["Terminal value (q9)", None, None, None, strip.terminal_value,
               strip.discounted_terminal_value])
    ws.append(["DivStrip implied price", None, None, None, None, strip.implied_price])
    ws.append(["12M FFA strip (front 4q)", strip.ffa_12m_by_class.get(cls)])
    _autosize(ws)

    ws = wb.create_sheet("Sensitivity")
    ws.append(["TCE \\ Vessel"] + [f"{v:+.0%}" for v in sens.vessel_value_shocks])
    for c in ws[1]:
        c.font = bold
    for i, ts in enumerate(sens.tce_shocks):
        ws.append([f"{ts:+.0%}"] + [sens.grid[i][j] for j in range(len(sens.vessel_value_shocks))])
    _autosize(ws)

    if report.payout_sensitivity:
        wsp = wb.create_sheet("Payout")
        wsp.append(["Dividend payout", "Fair value"])
        for c in wsp[1]:
            c.font = bold
        for payout in sorted(report.payout_sensitivity):
            wsp.append([payout, round(report.payout_sensitivity[payout], 2)])
        _autosize(wsp)

    if report.warnings:
        wsw = wb.create_sheet("Warnings")
        wsw.append(["Data validation warnings"])
        wsw["A1"].font = bold
        for warning in report.warnings:
            wsw.append([warning])
        _autosize(wsw)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# --------------------------------------------------------------------------- #
# public API
# --------------------------------------------------------------------------- #
def write_company_report(report: CompanyReport, outputs_dir: Path = OUTPUTS_DIR) -> Path:
    """Render ``outputs/{ticker}_fv_report.md`` and the matching xlsx; return the .md path."""
    outputs_dir.mkdir(parents=True, exist_ok=True)
    stem = report.ticker.lower() + "_fv_report"
    md_path = outputs_dir / f"{stem}.md"
    md_path.write_text(_render_markdown(report))
    _render_xlsx(report, outputs_dir / f"{stem}.xlsx")
    return md_path


def write_watchlist_summary(
    reports: list[CompanyReport], outputs_dir: Path = OUTPUTS_DIR
) -> Path:
    """Render the watchlist roll-up ``outputs/fair_value_summary.xlsx`` (section 7)."""
    outputs_dir.mkdir(parents=True, exist_ok=True)
    bold = Font(bold=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Fair value summary"
    headers = [
        "Ticker", "Basis", "Current", "Tool FV", "Watchlist Target",
        "Tool vs Current", "Tool vs Target", "Implied Breakeven TCE (blended)", "Cycle Position",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold

    for r in reports:
        fv = r.blended.fair_value_per_share
        basis = "CRUDE SLEEVE (allocated price)" if r.basis else "whole-company"
        ws.append([
            r.ticker,
            basis,
            r.current_price,
            round(fv, 2),
            r.analyst_target,
            f"{_pct(fv, r.current_price):+.1f}%",
            f"{_pct(fv, r.analyst_target):+.1f}%",
            round(r.breakeven.blended_breakeven_tce, 0),
            round(r.cycle.cycle_position, 2),
        ])
    _autosize(ws)
    path = outputs_dir / "fair_value_summary.xlsx"
    wb.save(path)
    return path
