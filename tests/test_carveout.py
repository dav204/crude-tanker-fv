"""Crude carve-out tests (hybrid operators, METHODOLOGY.md section 6)."""

import pytest

from crude_tanker_fv.carveout import crude_carve_out, product_read
from crude_tanker_fv.nav import compute_nav
from crude_tanker_fv.schemas import (
    BalanceSheet,
    CompanyInputs,
    CostStructure,
    DividendPolicy,
    FleetManifest,
    MarketData,
    Vessel,
    VesselValueCurve,
)

M = 1_000_000


def _vlcc_curve():
    return VesselValueCurve(cls="VLCC", dwt=300_000, newbuild=175 * M,
                            five_year_benchmark=138 * M, ten_year_benchmark=111 * M, scrap_25yr=22 * M)


def _mr_curve():
    return VesselValueCurve(cls="MR", dwt=50_000, newbuild=59 * M,
                            five_year_benchmark=48 * M, ten_year_benchmark=40 * M, scrap_25yr=6 * M)


def _hybrid(corporate_only=True):
    bs = BalanceSheet(
        ticker="HYB", quarter="2026-Q1", cash_and_equivalents=100 * M, working_capital_net=0,
        total_debt=200 * M, lease_liabilities=0, newbuild_capex_commitments=0,
        newbuild_advances_paid=0, diluted_shares_outstanding=10 * M,
        product_specific_debt=0 if corporate_only else 50 * M,
    )
    return CompanyInputs(
        fleet=FleetManifest(ticker="HYB", report_date="2026-Q1", vessels=[
            Vessel(id="v", cls="VLCC", dwt=300_000, age=5, count=2, sleeve="crude"),
            Vessel(id="m", cls="MR", dwt=50_000, age=5, count=2, sleeve="product"),
        ]),
        balance_sheet=bs,
        dividend_policy=DividendPolicy(ticker="HYB", policy_type="variable", payout_ratio=1.0),
        cost_structure=CostStructure(ticker="HYB", opex_per_day={"VLCC": 9000},
                                     annual_G_and_A=20 * M, annual_interest_expense=10 * M),
        market_data=MarketData(vessel_value_curves={"VLCC": _vlcc_curve(), "MR": _mr_curve()}),
    )


def test_crude_share_by_vessel_value():
    co = crude_carve_out(_hybrid())
    # 2x VLCC @ $138M = $276M crude; 2x MR @ $48M = $96M product.
    assert co.crude_value == pytest.approx(276 * M)
    assert co.product_value == pytest.approx(96 * M)
    assert co.crude_share == pytest.approx(276 / 372)


def test_balance_sheet_allocated_pro_rata():
    co = crude_carve_out(_hybrid())
    s = co.crude_share
    bs = co.crude_inputs.balance_sheet
    assert bs.cash_and_equivalents == pytest.approx(100 * M * s)
    assert bs.total_debt == pytest.approx(200 * M * s)            # all corporate -> pro-rated
    assert co.crude_inputs.cost_structure.annual_G_and_A == pytest.approx(20 * M * s)
    # Crude fleet excludes the product vessels.
    assert {v.cls for v in co.crude_inputs.fleet.vessels} == {"VLCC"}


def test_vessel_specific_debt_allocated_directly():
    co = crude_carve_out(_hybrid(corporate_only=False))  # $50M product-secured debt
    s = co.crude_share
    corporate = (200 - 50) * M
    # Product-secured debt is NOT borne by the crude sleeve; only corporate is pro-rated.
    assert co.crude_inputs.balance_sheet.total_debt == pytest.approx(corporate * s)


def test_preferred_equity_pro_rated_by_sleeve():
    """preferred_equity is a corporate-stack claim and should pro-rate the same
    way as corporate/unsecured debt — by vessel-value share (METHODOLOGY §4.2)."""
    from dataclasses import replace
    from crude_tanker_fv.carveout import product_carve_out
    base = _hybrid()
    # Add $40M of preferreds; everything else identical.
    hybrid_with_pref = replace(base,
        balance_sheet=replace(base.balance_sheet, preferred_equity=40 * M))
    co = crude_carve_out(hybrid_with_pref)
    po = product_carve_out(hybrid_with_pref)
    assert co.crude_inputs.balance_sheet.preferred_equity == pytest.approx(40 * M * co.crude_share)
    assert po.product_inputs.balance_sheet.preferred_equity == pytest.approx(40 * M * po.product_share)
    # The two sleeves' preferreds reaggregate to the whole-co figure.
    crude_pref = co.crude_inputs.balance_sheet.preferred_equity
    product_pref = po.product_inputs.balance_sheet.preferred_equity
    assert crude_pref + product_pref == pytest.approx(40 * M)


def _lngc_curve():
    return VesselValueCurve(cls="LNGC", dwt=174_000, newbuild=260 * M,
                            five_year_benchmark=235 * M, ten_year_benchmark=200 * M, scrap_25yr=50 * M)


def _three_sleeve_hybrid():
    """Synthetic 3-sleeve hybrid: 2 VLCC + 2 MR + 1 LNGC, all age 5."""
    bs = BalanceSheet(
        ticker="HYB3", quarter="2026-Q1", cash_and_equivalents=200 * M, working_capital_net=0,
        total_debt=400 * M, lease_liabilities=0, newbuild_capex_commitments=0,
        newbuild_advances_paid=0, diluted_shares_outstanding=20 * M,
        preferred_equity=60 * M,
        shuttle_contracted_book=80 * M,
    )
    return CompanyInputs(
        fleet=FleetManifest(ticker="HYB3", report_date="2026-Q1", vessels=[
            Vessel(id="v", cls="VLCC", dwt=300_000, age=5, count=2),
            Vessel(id="m", cls="MR", dwt=50_000, age=5, count=2),
            Vessel(id="g", cls="LNGC", dwt=174_000, age=5, count=1),
        ]),
        balance_sheet=bs,
        dividend_policy=DividendPolicy(ticker="HYB3", policy_type="variable", payout_ratio=1.0),
        cost_structure=CostStructure(ticker="HYB3", opex_per_day={"VLCC": 9000, "MR": 8000, "LNGC": 14000},
                                     annual_G_and_A=30 * M, annual_interest_expense=20 * M),
        market_data=MarketData(vessel_value_curves={"VLCC": _vlcc_curve(), "MR": _mr_curve(), "LNGC": _lngc_curve()}),
    )


def test_three_sleeve_carve_outs_sum_to_one():
    """For a 3-sleeve hybrid, crude_share + product_share + lng_share = 1.0."""
    from crude_tanker_fv.carveout import lng_carve_out, product_carve_out
    h = _three_sleeve_hybrid()
    co = crude_carve_out(h)
    po = product_carve_out(h)
    lo = lng_carve_out(h)
    # 2 VLCC @ $138M = $276M; 2 MR @ $48M = $96M; 1 LNGC @ $235M = $235M. Total $607M.
    assert co.crude_value == pytest.approx(276 * M)
    assert po.product_value == pytest.approx(96 * M)
    assert lo.lng_value == pytest.approx(235 * M)
    assert co.crude_share + po.product_share + lo.lng_share == pytest.approx(1.0)


def test_three_sleeve_corporate_stack_aggregates_to_whole_co():
    """Corporate-stack items (cash, debt, preferred, shuttle_contracted_book)
    pro-rated across all three sleeves should sum back to the whole-co figure."""
    from crude_tanker_fv.carveout import lng_carve_out, product_carve_out
    h = _three_sleeve_hybrid()
    co = crude_carve_out(h)
    po = product_carve_out(h)
    lo = lng_carve_out(h)
    # Cash: $200M whole-co split across three sleeves
    assert (co.crude_inputs.balance_sheet.cash_and_equivalents
            + po.product_inputs.balance_sheet.cash_and_equivalents
            + lo.lng_inputs.balance_sheet.cash_and_equivalents) == pytest.approx(200 * M)
    # Total debt: $400M whole-co (all corporate, no specific debt)
    assert (co.crude_inputs.balance_sheet.total_debt
            + po.product_inputs.balance_sheet.total_debt
            + lo.lng_inputs.balance_sheet.total_debt) == pytest.approx(400 * M)
    # Preferred: $60M whole-co
    assert (co.crude_inputs.balance_sheet.preferred_equity
            + po.product_inputs.balance_sheet.preferred_equity
            + lo.lng_inputs.balance_sheet.preferred_equity) == pytest.approx(60 * M)
    # Shuttle contracted-book: $80M whole-co
    assert (co.crude_inputs.balance_sheet.shuttle_contracted_book
            + po.product_inputs.balance_sheet.shuttle_contracted_book
            + lo.lng_inputs.balance_sheet.shuttle_contracted_book) == pytest.approx(80 * M)


def test_three_sleeve_fleet_split_is_clean():
    """Each sleeve's fleet should contain ONLY vessels of its sector."""
    from crude_tanker_fv.carveout import lng_carve_out, product_carve_out
    h = _three_sleeve_hybrid()
    co = crude_carve_out(h)
    po = product_carve_out(h)
    lo = lng_carve_out(h)
    assert {v.cls for v in co.crude_inputs.fleet.vessels} == {"VLCC"}
    assert {v.cls for v in po.product_inputs.fleet.vessels} == {"MR"}
    assert {v.cls for v in lo.lng_inputs.fleet.vessels} == {"LNGC"}


def test_three_sleeve_aggregator_sums_per_scenario():
    """``_aggregate_three_sleeve_report`` should sum per-scenario fair_value
    and nav_per_share across the three sleeves, weighted by the input
    scenarios' weights (METHODOLOGY §11.6). Tests the aggregation primitive
    directly without depending on real scenario / rate plumbing.
    """
    from crude_tanker_fv.pipeline import _aggregate_three_sleeve_report
    from crude_tanker_fv.scenarios import ScenarioFV, ScenarioReport

    def _r(name_prefix, fvs):
        scenarios = [ScenarioFV(
            name=f"{name_prefix}_s{i}", weight=w, fair_value=fv,
            fair_value_low=fv * 0.95, fair_value_high=fv * 1.05,
            nav_per_share=fv * 1.1, vessel_scale=1.0,
            divstrip_npv=fv * 0.4, cycle_position=1.5, w_nav=0.6,
            assumed_tce=40000,
        ) for i, (w, fv) in enumerate(fvs)]
        return ScenarioReport(
            ticker="X", current_price=0, analyst_target=0,
            base_nav_per_share=sum(fv * 1.1 for _, fv in fvs) / len(fvs),
            breakeven_tce=20000, scenarios=scenarios,
            probability_weighted_fv=sum(w * fv for w, fv in fvs),
            upside_best=0, downside_worst=0, expected_value_vs_current=0,
            position_recommendation="HOLD", basis="", sector="crude",
        )

    # Three weight-aligned scenario lists; weights {0.10, 0.50, 0.40}.
    crude = _r("crude", [(0.10, 50.0), (0.50, 40.0), (0.40, 30.0)])
    product = _r("prod",  [(0.10, 15.0), (0.50, 12.0), (0.40, 10.0)])
    lng = _r("lng",   [(0.10, 25.0), (0.50, 22.0), (0.40, 18.0)])

    headline = _aggregate_three_sleeve_report(
        crude, product, lng,
        ticker="HYB3", whole_price=80.0, whole_target=100.0,
        crude_share=0.6, product_share=0.2, lng_share=0.2,
    )
    # Per-scenario fair_value sums.
    for i, s in enumerate(headline.scenarios):
        expected = crude.scenarios[i].fair_value + product.scenarios[i].fair_value + lng.scenarios[i].fair_value
        assert s.fair_value == pytest.approx(expected)
    # PW FV = weighted sum of summed FVs.
    expected_pw_fv = 0.10 * 90.0 + 0.50 * 74.0 + 0.40 * 58.0
    assert headline.probability_weighted_fv == pytest.approx(expected_pw_fv)
    # Base NAV/sh sums across sleeves.
    assert headline.base_nav_per_share == pytest.approx(
        crude.base_nav_per_share + product.base_nav_per_share + lng.base_nav_per_share
    )
    # Basis text identifies as 3-sleeve.
    assert "3-SLEEVE" in headline.basis


def test_carved_inputs_run_through_nav():
    co = crude_carve_out(_hybrid())
    nav = compute_nav(co.crude_inputs)
    # crude NAV = 276 fleet + 0.742*(100 cash - 200 debt) over 10M shares.
    expected = (276 * M + co.crude_share * (100 - 200) * M) / (10 * M)
    assert nav.nav_per_share == pytest.approx(expected, rel=1e-6)


def _lr1_curve():
    return VesselValueCurve(cls="LR1", dwt=75_000, newbuild=59 * M,
                            five_year_benchmark=53.5 * M, ten_year_benchmark=42 * M, scrap_25yr=7.7 * M)


def test_dual_use_lr1_split():
    # 2 VLCC (crude) + 2 LR1 dual-use at crude_fraction 0.30.
    inputs = CompanyInputs(
        fleet=FleetManifest(ticker="HYB", report_date="2026-Q1", vessels=[
            Vessel(id="v", cls="VLCC", dwt=300_000, age=5, count=2, sleeve="crude"),
            Vessel(id="l", cls="LR1", dwt=75_000, age=5, count=2, sleeve="product", crude_fraction=0.30),
        ]),
        balance_sheet=BalanceSheet(
            ticker="HYB", quarter="2026-Q1", cash_and_equivalents=0, working_capital_net=0,
            total_debt=0, lease_liabilities=0, newbuild_capex_commitments=0,
            newbuild_advances_paid=0, diluted_shares_outstanding=10 * M),
        dividend_policy=DividendPolicy(ticker="HYB", policy_type="variable", payout_ratio=1.0),
        cost_structure=CostStructure(ticker="HYB", opex_per_day={"VLCC": 9000, "LR1": 7000}),
        market_data=MarketData(vessel_value_curves={"VLCC": _vlcc_curve(), "LR1": _lr1_curve()}),
    )
    co = crude_carve_out(inputs)
    lr1_total = 53.5 * M * 2          # 2x LR1 @ $53.5M
    assert co.product_value == pytest.approx(lr1_total * 0.70)     # 70% to product
    assert co.crude_value == pytest.approx(276 * M + lr1_total * 0.30)  # VLCC + 30% LR1
    # The LR1's crude portion is in the crude fleet at 0.30x count.
    lr1 = [v for v in co.crude_inputs.fleet.vessels if v.cls == "LR1"][0]
    assert lr1.count == pytest.approx(0.6)   # 2 * 0.30
    compute_nav(co.crude_inputs)             # runs without error (LR1 has a value curve)


def test_product_read_string():
    co = crude_carve_out(_hybrid())
    txt = product_read(co, crude_ev_pct=-5.0)
    assert "Product sleeve" in txt and "OVERSTATES" in txt and "v2" in txt


def test_product_carve_out_complements_crude(tmp_path):
    """v2 invariant: crude_value + product_value == total vessel value across both carves."""
    from crude_tanker_fv.carveout import product_carve_out
    ci = _hybrid()
    crude = crude_carve_out(ci)
    product = product_carve_out(ci)
    # Same total vessel value seen from both sleeves.
    assert crude.crude_value == pytest.approx(product.crude_value)
    assert crude.product_value == pytest.approx(product.product_value)
    # Shares add to 1.
    assert crude.crude_share + product.product_share == pytest.approx(1.0)
    # Pro-rated BS items sum exactly to whole-company values.
    crude_bs = crude.crude_inputs.balance_sheet
    product_bs = product.product_inputs.balance_sheet
    whole_bs = ci.balance_sheet
    for field in ("cash_and_equivalents", "working_capital_net", "total_debt",
                  "lease_liabilities", "newbuild_capex_commitments",
                  "newbuild_advances_paid"):
        assert getattr(crude_bs, field) + getattr(product_bs, field) == pytest.approx(
            getattr(whole_bs, field)
        )


def test_whole_company_nav_aggregation_invariant():
    """v2: sleeve-NAV-per-share sums equal the whole-company NAV/share."""
    from crude_tanker_fv.carveout import product_carve_out
    from crude_tanker_fv.nav import compute_nav
    ci = _hybrid()
    whole_nav = compute_nav(ci).nav_per_share
    crude_nav = compute_nav(crude_carve_out(ci).crude_inputs).nav_per_share
    product_nav = compute_nav(product_carve_out(ci).product_inputs).nav_per_share
    assert crude_nav + product_nav == pytest.approx(whole_nav, abs=0.01)


def test_product_rate_remap_uses_clean_variants():
    """Product carve-out remaps LR1 / LR2 / MR rate entries to the clean variants."""
    from crude_tanker_fv.carveout import product_carve_out
    from crude_tanker_fv.loaders import load_company_inputs
    insw = load_company_inputs("INSW", "2026-Q1")
    pco = product_carve_out(insw)
    md = pco.product_inputs.market_data
    base = insw.market_data
    # LR1 in product sleeve picks up the LR1_clean rates (not the dirty / Aframax-equiv).
    assert md.spot_tce["LR1"] == pytest.approx(base.spot_tce["LR1_clean"])
    assert md.twelve_month_tc["LR1"] == pytest.approx(base.twelve_month_tc["LR1_clean"])
    # LR2 similarly.
    assert md.spot_tce["LR2"] == pytest.approx(base.spot_tce["LR2_clean"])
    # MR carried through unchanged.
    assert md.spot_tce["MR"] == pytest.approx(base.spot_tce["MR"])


def test_insw_outputs_are_clearly_labeled_v2_whole_company(tmp_path):
    """Under v2 (METHODOLOGY 6 v2), scenario/sweep/comparison outputs aggregate
    BOTH sleeves into a WHOLE-COMPANY view; only the single-point FV report stays
    crude-sleeve (it's the detail report; whole-co lives in the scenario report)."""
    from openpyxl import load_workbook

    from crude_tanker_fv.pipeline import (
        run_broker_sweep,
        run_scenarios_watchlist,
        run_transaction_anchored_comparison,
        run_watchlist,
    )

    run_watchlist("2026-Q1", outputs_dir=tmp_path)
    run_scenarios_watchlist("2026-Q1", outputs_dir=tmp_path)
    run_broker_sweep("2026-Q1", outputs_dir=tmp_path)
    run_transaction_anchored_comparison("2026-Q1", outputs_dir=tmp_path)

    # FV report (single-point detail) STAYS crude-sleeve in v2 — the whole-co
    # number is in the scenario report.
    insw_fv = (tmp_path / "insw_fv_report.md").read_text()
    assert "CRUDE SLEEVE" in insw_fv
    assert "Current price (crude-allocated)" in insw_fv
    wb = load_workbook(tmp_path / "insw_fv_report.xlsx")
    summary = {row[0].value: row[1].value for row in wb["Summary"].iter_rows()}
    assert summary.get("Valuation basis") == "CRUDE SLEEVE (allocated price)"

    # Scenario report .md is the WHOLE-COMPANY headline now (v2).
    insw_sc = (tmp_path / "insw_scenarios.md").read_text()
    assert "[WHOLE-CO]" in insw_sc and "Valuation basis" in insw_sc
    assert "WHOLE-COMPANY" in insw_sc
    assert "Hybrid sleeve breakdown" in insw_sc      # per-sleeve table appended
    # Whole-company price should NOT carry the (crude-allocated) qualifier
    assert "Current price (crude-allocated)" not in insw_sc

    # Watchlist roll-up reflects the FV report (still crude-sleeve detail).
    wb = load_workbook(tmp_path / "fair_value_summary.xlsx")
    rows = list(wb.active.iter_rows(values_only=True))
    insw_row = next(r for r in rows[1:] if r[0] == "INSW")
    assert insw_row[1] == "CRUDE SLEEVE (allocated price)"

    # Scenario roll-up now flags WHOLE-COMPANY for INSW.
    wb = load_workbook(tmp_path / "scenario_summary.xlsx")
    rows = list(wb.active.iter_rows(values_only=True))
    insw_row = next(r for r in rows[1:] if r[0] == "INSW")
    assert insw_row[1] == "WHOLE-COMPANY (hybrid aggregation)"

    # Broker sweep + transaction-anchor comparison: WHOLE-CO for INSW.
    sweep_md = (tmp_path / "broker_nav_sweep.md").read_text()
    assert "INSW **(WHOLE-CO)**" in sweep_md
    wb = load_workbook(tmp_path / "broker_nav_sweep.xlsx")
    rows = list(wb.active.iter_rows(values_only=True))
    insw_row = next(r for r in rows[1:] if r[0] == "INSW")
    assert insw_row[1] == "WHOLE-COMPANY (hybrid aggregation)"

    txn_md = (tmp_path / "transaction_anchor_comparison.md").read_text()
    assert "INSW **(WHOLE-CO)**" in txn_md
    wb = load_workbook(tmp_path / "transaction_anchor_comparison.xlsx")
    rows = list(wb.active.iter_rows(values_only=True))
    insw_row = next(r for r in rows[1:] if r[0] == "INSW")
    assert insw_row[1] == "WHOLE-COMPANY (hybrid aggregation)"
