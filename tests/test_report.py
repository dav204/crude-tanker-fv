"""Report-rendering and pipeline-assembly tests (METHODOLOGY.md sections 7-8)."""

from conftest import BOOK_QUARTER  # follows the book across quarter rolls
import pytest
from openpyxl import load_workbook

from crude_tanker_fv.loaders import load_watchlist
from crude_tanker_fv.pipeline import run_watchlist, value_company
from crude_tanker_fv.report import write_company_report, write_watchlist_summary


@pytest.fixture
def dht_report():
    return value_company("DHT", "2026-Q1", current_price=16.35, analyst_target=16.00)


def test_value_company_assembles_consistent_report(dht_report):
    r = dht_report
    assert r.ticker == "DHT"
    # Blend identity: FV = w_nav*NAV + w_earn*strip.
    expected = r.cycle.w_nav * r.nav.nav_per_share + r.cycle.w_earn * r.strip.implied_price
    assert r.blended.fair_value_per_share == pytest.approx(expected)
    # Band re-based 2026-06-09 (txn-anchored marks default-on, ~16.5→~14.3) then
    # 2026-06-22 (cycle-conditional terminal §9.2: DHT late-cycle/peak → 0.9x
    # mean-reversion of the terminal fleet value, ~14.3→~14.0).
    assert 13.5 < r.blended.fair_value_per_share < 15.0


def test_write_company_report_creates_md_and_xlsx(dht_report, tmp_path):
    md = write_company_report(dht_report, outputs_dir=tmp_path)
    assert md.exists() and md.suffix == ".md"
    xlsx = tmp_path / "dht_fv_report.xlsx"
    assert xlsx.exists()

    text = md.read_text()
    for needle in ["# DHT", "NAV breakdown", "Dividend strip", "Implied breakeven TCE",
                   "Sensitivity", "Divergence diagnosis", "$14.68", "FFA spot",
                   "Data validation warnings"]:
        assert needle in text

    wb = load_workbook(xlsx)
    assert wb.sheetnames[:4] == ["Summary", "NAV", "DivStrip", "Sensitivity"]
    assert "Warnings" in wb.sheetnames  # DHT's $462k spot is flagged
    assert wb["Sensitivity"].max_row == 6 and wb["Sensitivity"].max_column == 6


def test_write_watchlist_summary(dht_report, tmp_path):
    path = write_watchlist_summary([dht_report], outputs_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb.active
    assert ws.max_row == 2  # header + 1 ticker
    header = [c.value for c in ws[1]]
    assert header[:5] == ["Ticker", "Basis", "Current", "Tool FV", "Watchlist Target"]
    row = [c.value for c in ws[2]]
    assert row[0] == "DHT"
    assert row[1] == "whole-company"   # DHT is a pure-play
    assert row[3] == pytest.approx(14.68, abs=0.01)   # re-pinned 2026-07-18: VLCC mid-age slope eased on the C. Innovator age-14 print (marks-trail promotion; was 14.95 since §9.6 2026-06-30)


def test_run_watchlist_end_to_end(tmp_path):
    reports = run_watchlist(BOOK_QUARTER, outputs_dir=tmp_path)
    tickers = {r.ticker for r in reports}
    assert {"DHT", "FRO"} <= tickers  # both modeled (ECO/INSW skipped, no target yet)
    assert (tmp_path / "dht_fv_report.md").exists()
    assert (tmp_path / "fro_fv_report.md").exists()
    assert (tmp_path / "fair_value_summary.xlsx").exists()


def test_watchlist_loader_reads_dht():
    wl = load_watchlist()
    # re-pinned 2026-07-06 at the consensus-pair recapture (Pareto 3 Jul daily)
    assert wl["DHT"]["current_price"] == pytest.approx(17.20)
    assert wl["DHT"]["analyst_target"] == pytest.approx(16.00)
    assert wl["FRO"]["analyst_target"] == pytest.approx(30.50)
    assert wl["ECO"]["analyst_target"] == pytest.approx(45.00)
    assert wl["INSW"]["analyst_target"] == pytest.approx(79.50)   # hybrid, modeled via carve-out


def test_fv_attribution_block_foots_to_blend_fv(dht_report, tmp_path):
    """M-1 (methodology review 2026-07-14): the FV attribution rows must sum to the
    blend FV exactly, and the effective asset-value share must equal
    w_nav + w_earn x (discounted terminal / strip) — the decomposition that makes
    the marks-vs-rates effort gradient legible. Rendering-only: no FV changes."""
    r = dht_report
    md = write_company_report(r, outputs_dir=tmp_path).read_text()
    assert "### FV attribution" in md
    assert "Effective asset-value share" in md

    shares = r.nav.diluted_shares_outstanding
    fleet_ps = r.nav.fleet_value / shares
    bs_net_ps = r.nav.nav_per_share - fleet_ps
    dps_pv = r.strip.implied_price - r.strip.discounted_terminal_value
    haircut_ps = r.blended.nav_per_share_effective - r.nav.nav_per_share
    total = (r.cycle.w_nav * (fleet_ps + bs_net_ps + haircut_ps)
             + r.cycle.w_earn * (dps_pv + r.strip.discounted_terminal_value))
    assert total == pytest.approx(r.blended.fair_value_per_share, abs=1e-9)

    asset_share = (r.cycle.w_nav
                   + r.cycle.w_earn * r.strip.discounted_terminal_value / r.strip.implied_price)
    assert 0.0 < asset_share < 1.0
    # DHT at peak weighting: the memo's ~0.84 effective asset share reproduces.
    assert asset_share == pytest.approx(0.84, abs=0.03)
