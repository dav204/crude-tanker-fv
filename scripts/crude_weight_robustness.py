"""Crude tanker weight-robustness diagnostic (METHODOLOGY §9.10).

Runs all six crude tanker watchlist names (DHT, ECO, FRO, INSW, TNK, NAT) under
four defensible probability-weight sets and classifies each name's position
recommendation as weight-robust (same call across all four) or weight-driven
(call changes across sets).

Driver (2026-06-01): Catlin / VIE analysis (2026-05-25) plus the June 1 macro
briefing suggest current crude weights {escalation 0.10, pre_mou_baseline 0.15,
mou_base 0.50, mou_bear 0.25} may put too much weight on "deep normalization"
relative to "slow normalization with extended Phase 1." This diagnostic does
NOT change the locked Set A weights — it surfaces which calls would survive a
defensible reweighting (Set B Catlin-leaning) vs which depend on the specific
Set A prior.

Output:
  - outputs/weight_robustness_diagnostic.md   (markdown table + per-name detail)
  - outputs/weight_robustness_diagnostic.xlsx (single sheet, machine-readable)

Naming-namespace note: this script uses "Set A/B/C/D" for CRUDE weight families.
The LNG sector already uses "Set B" / "Set B-revised" for its own weight family
(METHODOLOGY §11.3). To avoid ambiguity, every label here carries the explicit
"(crude)" qualifier. Cross-sector conflation would be a methodology error.
"""

from __future__ import annotations

import copy
import sys
from pathlib import Path

# Make src importable
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from crude_tanker_fv.loaders import INPUTS_DIR, load_company_inputs, load_watchlist
from crude_tanker_fv.pipeline import (
    _load_all_sectors,
    _maybe_apply_transactions,
    _run_scenarios_for_ticker,
)

# ----------------------------------------------------------------------------
# Weight definitions (METHODOLOGY §9.10)
# ----------------------------------------------------------------------------
# All four sets sum to 1.0 across the four active crude scenarios.
# Set A is the production lock; B / C / D are defensible alternatives chosen to
# bracket the "how fast does normalisation arrive" axis (the dominant
# uncertainty for 2026 crude calls).
CRUDE_WEIGHT_SETS = {
    # Set A refreshed 2026-07-02 to the ACTUAL locked weights — it had silently
    # kept the pre-Jun-9 v1 set after the Jun-9 escalation reweight, so the
    # "current locked" column was not the production prior for three weeks.
    # Set A' adopted 2026-07-31 — the B' reweight executed per the 7/22 frozen
    # conditional (mou_scenario_reweight_proposal_2026-07-22.md; owner "Execute
    # B'" 7/31): dead-MoU mass retired into pre_mou; escalation untouched.
    "Crude Set A' (B' reweight, production 2026-07-31)": {
        "escalation": 0.25, "pre_mou_baseline": 0.57,
        "mou_base":   0.05, "mou_bear":         0.13,
    },
    "Crude Set A (Jun-9 war tilt, history bracket)": {
        "escalation": 0.25, "pre_mou_baseline": 0.45,
        "mou_base":   0.18, "mou_bear":         0.12,
    },
    "Crude Set B (Catlin-leaning, slow normalization)": {
        "escalation": 0.10, "pre_mou_baseline": 0.25,
        "mou_base":   0.45, "mou_bear":         0.20,
    },
    "Crude Set C (bullish, extended Phase 1)": {
        "escalation": 0.15, "pre_mou_baseline": 0.30,
        "mou_base":   0.40, "mou_bear":         0.15,
    },
    "Crude Set D (bearish, deep normalization)": {
        "escalation": 0.05, "pre_mou_baseline": 0.10,
        "mou_base":   0.55, "mou_bear":         0.30,
    },
    # Proposed 2026-07-02 (audit F-2 / §13.3 trigger — US–Iran stand-down):
    # v1 prior restored, adjusted for one observed ceasefire failure (esc 0.10
    # not 0.05) and no signed MoU yet (pre_mou 0.20 not 0.15). Was labeled
    # "current locked" — true Jul-2→Jul-14 only; the Hormuz re-tilt (RESTORE
    # BOTH, owner ruling 2026-07-14 EVE) put production back on Set A. The
    # label is data in the sidecar/outputs, so it must not claim "current";
    # main() asserts which set the production prior actually is.
    "Crude Set E (Jul-2 stand-down vintage)": {
        "escalation": 0.10, "pre_mou_baseline": 0.20,
        "mou_base":   0.45, "mou_bear":         0.25,
    },
}

for label, weights in CRUDE_WEIGHT_SETS.items():
    assert abs(sum(weights.values()) - 1.0) < 1e-9, f"{label} doesn't sum to 1"

# Extended 2026-07-02 (review W-1) from the six pure names to EVERY
# crude-exposed name — the hybrids/multi-sleeves (INSW/TEN/CMBT) route through
# the production scenario path, and the newbuild-torque names (BRUT/CAPT) are
# exactly where weight fragility bites hardest (BRUT: BUY +98% -> HOLD -5%
# across the family in the Jul-2 proposal run).
CRUDE_TICKERS = ["DHT", "ECO", "FRO", "INSW", "TNK", "NAT",
                 "TEN", "CMBT", "BRUT", "CAPT"]


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def doc_with_weights(base_doc: dict, weights: dict) -> dict:
    """Deep-copy the crude doc and overwrite scenario weights in place."""
    d = copy.deepcopy(base_doc)
    for name, w in weights.items():
        if name in d["scenarios"]:
            d["scenarios"][name]["weight"] = w
    return d


def position_label(ev_pct: float) -> str:
    """Same threshold as scenarios.position_recommendation but shorter labels."""
    if ev_pct > 5.0:
        return "BUY"
    if ev_pct < -5.0:
        return "TRIM/SHORT"
    return "HOLD"


# ----------------------------------------------------------------------------
# Per-ticker × per-weight-set run
# ----------------------------------------------------------------------------
def run_ticker_under_all_sets(ticker: str, watchlist: dict, base_sector_docs: dict):
    """Returns dict {set_label: {pw_fv, ev_pct, position}}.

    Uses the production scenario-routing path (`_run_scenarios_for_ticker`) so
    INSW's hybrid carve-out aggregation is handled identically to a real
    pipeline run. The weight override only touches `sector_docs["crude"]`, so
    INSW's product sleeve (running through `sectors.product`) is unchanged —
    the diagnostic correctly isolates the crude-weight sensitivity.
    """
    entry = watchlist[ticker]
    price = entry["current_price"]
    target = entry["analyst_target"]
    ci = load_company_inputs(ticker, "2026-Q1")
    # Production path parity (fixed 2026-07-02): transaction-anchored marks are
    # DEFAULT-ON in the pipeline; without this the diagnostic ran ~12% high on
    # un-anchored marks for every name since 2026-06-09.
    ci, _ = _maybe_apply_transactions(ci, INPUTS_DIR, True)

    results: dict[str, dict] = {}
    for set_label, weights in CRUDE_WEIGHT_SETS.items():
        # Build sector_docs with the overridden crude weights; LNG and product
        # docs pass through unchanged (their tickers aren't in this analysis
        # but the dict still needs to be complete for the routing logic).
        modified_docs = dict(base_sector_docs)
        modified_docs["crude"] = doc_with_weights(base_sector_docs["crude"], weights)

        report, _, _ = _run_scenarios_for_ticker(
            ticker, ci, price, target, modified_docs, watchlist,
        )
        ev_pct = report.expected_value_vs_current / price * 100.0
        results[set_label] = {
            "pw_fv":    round(report.probability_weighted_fv, 2),
            "ev_pct":   round(ev_pct, 1),
            "position": position_label(ev_pct),
        }
    return {"ticker": ticker, "price": price, "target": target, "results": results}


def classify_robustness(per_ticker: dict) -> tuple[str, list[str], str]:
    """Returns (verdict, positions_across_sets, robust_or_driven_note).

    WEIGHT-ROBUST = same position across all four sets.
    WEIGHT-DRIVEN = position changes; we identify which set produces the flip.
    """
    positions = [per_ticker["results"][s]["position"] for s in CRUDE_WEIGHT_SETS]
    unique_positions = set(positions)
    if len(unique_positions) == 1:
        return "WEIGHT-ROBUST", positions, (
            f"position {positions[0]} across all {len(CRUDE_WEIGHT_SETS)} weight sets")

    # Weight-driven — describe the flip. Find the "majority" position then
    # name the dissenting set(s).
    set_labels = list(CRUDE_WEIGHT_SETS.keys())
    by_position: dict[str, list[str]] = {}
    for s, pos in zip(set_labels, positions):
        by_position.setdefault(pos, []).append(s)
    parts = []
    for pos, sets in by_position.items():
        # Shorten set label for inline display
        short = [s.split("(")[0].strip().replace("Crude ", "") for s in sets]
        parts.append(f"{pos} under {'/'.join(short)}")
    return "WEIGHT-DRIVEN", positions, "; ".join(parts)


# ----------------------------------------------------------------------------
# Output rendering
# ----------------------------------------------------------------------------
def set_short_defs() -> list[str]:
    return [s.split("(")[0].strip().replace("Crude ", "") for s in CRUDE_WEIGHT_SETS]


def render_markdown(analyses: list[dict]) -> str:
    lines: list[str] = []
    lines.append("# Crude Weight-Robustness Diagnostic\n")
    lines.append("Diagnostic (METHODOLOGY §9.10) — does NOT change the locked "
                 "Crude Set A weights. Surfaces which crude tanker calls "
                 "survive defensible reweighting (call is **weight-robust**) "
                 "vs which depend on a specific weight prior (**weight-driven**).")
    lines.append("")
    lines.append("**Driver:** Catlin / VIE analysis (2026-05-25) plus the "
                 "June 1 macro briefing suggest current Set A weights may put "
                 "too much weight on \"deep normalisation\" relative to \"slow "
                 "normalisation with extended Phase 1.\" Sets B/C/D bracket the "
                 "normalisation-speed axis.")
    lines.append("")
    lines.append("**Naming namespace:** the labels below are CRUDE-sector "
                 "weight families. The LNG sector uses its own \"Set B\" / "
                 "\"Set B-revised\" naming (METHODOLOGY §11.3). Cross-sector "
                 "conflation would be a methodology error.\n")

    # Key findings — DERIVED from this run's analyses (the block was hand-
    # curated until 2026-07-02, when a reweight run left it contradicting the
    # generated table below it — the audit-F-8 pattern). The mark-spread
    # dimension lives in outputs/broker_nav_sweep.md; read the two together.
    lines.append("## Key findings (weight robustness, this run)\n")
    lines.append("Mark-spread robustness is the OTHER dimension — cross-read "
                 "with `outputs/broker_nav_sweep.md` before acting on any call.")
    lines.append("")
    lines.append("| Ticker | Weight robustness | What drives the call |")
    lines.append("|---|---|---|")
    for a in analyses:
        verdict, _, note = classify_robustness(a)
        badge = "✓ robust" if verdict == "WEIGHT-ROBUST" else "⚑ driven"
        lines.append(f"| {a['ticker']} | {badge} | {note} |")
    lines.append("")
    # Weight-set definitions table
    lines.append("## Weight sets compared\n")
    lines.append("| Scenario | " + " | ".join(set_short_defs()) + " |")
    lines.append("|---|" + "--:|" * len(CRUDE_WEIGHT_SETS))
    scen_order = ["escalation", "pre_mou_baseline", "mou_base", "mou_bear"]
    for scen in scen_order:
        vals = [CRUDE_WEIGHT_SETS[s][scen] for s in CRUDE_WEIGHT_SETS]
        lines.append(f"| {scen} | " + " | ".join(f"{v:.2f}" for v in vals) + " |")
    lines.append("")
    lines.append("Set B (Catlin-leaning) shifts 10pp from `mou_base` and 5pp "
                 "from `mou_bear` into `pre_mou_baseline` — i.e. Phase 1 "
                 "extends, Phase 2 normalisation arrives later. Set C is more "
                 "bullish (15pp into Phase 1). Set D is more bearish (15pp "
                 "deeper into MoU phase).\n")

    # Headline robustness table — columns derived from CRUDE_WEIGHT_SETS so
    # adding a set can never desynchronize header and cells (caught 2026-07-02
    # when Set E made the rows 5 cells against a hardcoded 4-column header).
    lines.append("## Summary — per-name robustness\n")
    set_short = [s.split("(")[0].strip().replace("Crude ", "") for s in CRUDE_WEIGHT_SETS]
    lines.append("| Ticker | " + " | ".join(f"{s} EV" for s in set_short) +
                 " | Robustness | Notes |")
    lines.append("|---|" + "--:|" * len(set_short) + "---|---|")
    for a in analyses:
        r = a["results"]
        verdict, positions, note = classify_robustness(a)
        cells = [f"{r[s]['ev_pct']:+.1f}% ({r[s]['position']})" for s in CRUDE_WEIGHT_SETS]
        badge = "✓ robust" if verdict == "WEIGHT-ROBUST" else "⚑ driven"
        lines.append(f"| {a['ticker']} | " + " | ".join(cells) + f" | {badge} | {note} |")
    lines.append("")

    # Per-name detail
    lines.append("## Per-name detail\n")
    for a in analyses:
        lines.append(f"### {a['ticker']} — price ${a['price']:.2f}, target ${a['target']:.2f}\n")
        verdict, positions, note = classify_robustness(a)
        lines.append(f"**Classification:** {verdict}. {note}.\n")
        lines.append("| Weight set | PW FV | EV % | Position |")
        lines.append("|---|--:|--:|---|")
        for s, r in a["results"].items():
            lines.append(f"| {s} | ${r['pw_fv']:.2f} | {r['ev_pct']:+.1f}% | {r['position']} |")
        lines.append("")

    # Methodology + combined diagnostic note
    lines.append("## Combined mark + weight robustness framework\n")
    lines.append("Pairing this diagnostic with the broker-NAV sweep (METHODOLOGY §9.9) "
                 "gives every name two robustness dimensions:")
    lines.append("")
    lines.append("- **Mark-robust + weight-robust** = highest-conviction signals "
                 "(call survives both vessel-mark uncertainty and probability-"
                 "weight reshuffling)")
    lines.append("- **Mark-driven OR weight-driven** (one of the two) = "
                 "moderate conviction; the call depends on one specific "
                 "judgemental input")
    lines.append("- **Mark-driven AND weight-driven** = lowest conviction; "
                 "two compounding judgemental dependencies. Treat with "
                 "explicit sizing discipline.\n")

    lines.append("See METHODOLOGY §9.9 (mark robustness) and §9.10 (weight "
                 "robustness) for the methodology. This diagnostic is the "
                 "§9.10 output for the crude sector; the LNG analogue lives "
                 "in `outputs/lng_weight_robustness.md`.\n")

    return "\n".join(lines)


def write_xlsx(analyses: list[dict], path: Path) -> None:
    """Single-sheet xlsx — same content as the markdown table, machine-readable."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Weight robustness (crude)"

    headers = ["ticker", "price", "target"]
    for s in CRUDE_WEIGHT_SETS:
        # Strip the long parenthetical for column-header brevity
        short = s.split("(")[0].strip()
        headers.extend([f"{short} PW FV", f"{short} EV %", f"{short} position"])
    headers.extend(["robustness", "notes"])
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    driven_fill = PatternFill("solid", fgColor="FFE6E6")  # soft pink for flagged

    for a in analyses:
        verdict, positions, note = classify_robustness(a)
        row = [a["ticker"], a["price"], a["target"]]
        for s in CRUDE_WEIGHT_SETS:
            r = a["results"][s]
            row.extend([r["pw_fv"], r["ev_pct"], r["position"]])
        row.extend([verdict, note])
        ws.append(row)
        if verdict == "WEIGHT-DRIVEN":
            for c in ws[ws.max_row]:
                c.fill = driven_fill

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    wb.save(path)


def sidecar_entries(analyses: list[dict]) -> dict:
    """Per name: does the EV SIGN survive the whole weight family (review
    2026-07-02, W-1)? Sign-stability, not position-label stability — a HOLD/BUY
    label flip inside a positive-EV range is magnitude sensitivity, while a sign
    flip means the direction of the call is a weight artifact. No timestamps
    (byte-stable regeneration); the vintage is the weight-set list itself PLUS
    the marks/tape the EVs were computed at — the WO1-F4 sha stamp can't see
    the latter, so the scorecard withholds any name whose live point EV exits
    the recorded range (containment guard, 2026-07-15: the MR age-0 re-anchor
    moved TEN +44.9 → +45.0 with this sidecar held). A range whose family
    includes the adopted set must contain the point, or it doesn't print.
    Written via scorecard.update_weight_fragility_sidecar (MERGE — S-2: the old
    whole-file write clobbered other families' entries)."""
    out = {}
    for a in analyses:
        evs = [a["results"][s]["ev_pct"] for s in CRUDE_WEIGHT_SETS]
        out[a["ticker"]] = {
            "ev_min_pct": min(evs),
            "ev_max_pct": max(evs),
            "ev_sign_stable": (min(evs) > 0) == (max(evs) > 0) and 0 not in (min(evs), max(evs)),
            "positions": [a["results"][s]["position"] for s in CRUDE_WEIGHT_SETS],
        }
    return out


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    # live_prices=True (2026-07-02): the diagnostic must value EV at the same
    # tape the pipeline does — on statics its Set A column silently disagreed
    # with the pipeline headline whenever the daily price had moved.
    watchlist = load_watchlist(live_prices=True)
    base_sector_docs = _load_all_sectors()

    # The family must INCLUDE the production prior — the scorecard's containment
    # guard asserts every printed point EV sits inside the family range, which
    # only holds if the adopted scenario_inputs.yaml weights are one of the sets
    # above. An owner reweight to a set not in the family fails HERE with the
    # cause, not downstream as an unexplained out-of-range withholding.
    adopted = {name: sc.get("weight", 0.0)
               for name, sc in base_sector_docs["crude"]["scenarios"].items()}
    member = next((label for label, ws in CRUDE_WEIGHT_SETS.items()
                   if all(abs(adopted.get(k, 0.0) - w) < 1e-9 for k, w in ws.items())),
                  None)
    if member is None:
        sys.exit("adopted crude weights "
                 f"{ {k: v for k, v in adopted.items() if v} } match no set in "
                 "CRUDE_WEIGHT_SETS — add the production prior to the family "
                 "before running the §9.10 diagnostic")
    print(f"production prior = {member}")

    analyses = []
    for ticker in CRUDE_TICKERS:
        if ticker not in watchlist:
            print(f"skip {ticker}: not in watchlist")
            continue
        a = run_ticker_under_all_sets(ticker, watchlist, base_sector_docs)
        analyses.append(a)

    project_root = Path(__file__).resolve().parents[1]
    out_md = project_root / "outputs" / "weight_robustness_diagnostic.md"
    out_xlsx = project_root / "outputs" / "weight_robustness_diagnostic.xlsx"

    out_md.write_text(render_markdown(analyses))
    write_xlsx(analyses, out_xlsx)
    from crude_tanker_fv.scorecard import update_weight_fragility_sidecar
    out_yaml = update_weight_fragility_sidecar(
        "crude", dict(CRUDE_WEIGHT_SETS), sidecar_entries(analyses))

    print(f"→ {out_md}")
    print(f"→ {out_xlsx}")
    print(f"→ {out_yaml}")
    print()

    # Compact terminal summary — columns derive from CRUDE_WEIGHT_SETS
    print(f"{'Ticker':6s}  " + " ".join(f"{s:12s}" for s in set_short_defs()) + "  Robustness")
    print("-" * (10 + 13 * len(CRUDE_WEIGHT_SETS) + 14))
    for a in analyses:
        verdict, positions, _ = classify_robustness(a)
        cells = [f"{a['results'][s]['ev_pct']:+5.1f}% {a['results'][s]['position'][:4]}"
                 for s in CRUDE_WEIGHT_SETS]
        print(f"{a['ticker']:6s}  " + " ".join(f"{c:12s}" for c in cells) + f"  {verdict}")


if __name__ == "__main__":
    main()
