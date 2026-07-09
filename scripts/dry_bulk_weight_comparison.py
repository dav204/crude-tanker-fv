"""Dry-bulk weight-robustness diagnostic (METHODOLOGY §9.10) — WO4.

Runs the four valued dry-bulk watchlist names (SBLK, GNK, CMDB, SB) under the
locked Bulk Set A plus two defensible bracketing weight sets, and classifies
each name's position recommendation as weight-robust (same call across all
sets) or weight-driven (call changes). This unblocks the consumer's Gate E,
whose "sign-flip at desk depth = FAIL" reads `weight_sign_stable` from the
`outputs/weight_robustness.yaml` sidecar (schema-2.2 handoff seam already
emits the family fields for any sector present in the sidecar).

DIAGNOSTIC ONLY (WO4 non-goal guardrail): this does NOT change the locked
Bulk Set A weights in inputs/scenario_inputs.yaml. The bracketing sets live
HERE, in the script, never in production. A weight-driven finding is a Gate-E
input to REPORT, not a license to reweight (that is a separate §11.7.x
revision with its own lock test).

Sensitivity axis (WO4 Phase 0): CHINA DRY-BULK DEMAND TENSION — the axis the
four scenarios already parameterize (china_acceleration ↔ china_property_drag /
coordinated_slowdown) and the charter thesis's own load-bearing variable
(Simandou ton-mile + supply discipline read against China property/steel drag).

Naming-namespace note: labels are DRY-BULK weight families ("Bulk Set …").
Both CRUDE ("Crude Set B") and LNG ("Set B-revised") already use "Set B" for
their own families; a bare unprefixed "Set B" would be a methodology error, so
every label here carries the explicit "Bulk" qualifier.

Output:
  - outputs/dry_bulk_weight_robustness.md    (markdown table + per-name detail)
  - outputs/dry_bulk_weight_robustness.xlsx  (single sheet, machine-readable)
  - outputs/weight_robustness.yaml           (SHARED sidecar; merge-aware)
"""

from __future__ import annotations

import copy
import sys
from pathlib import Path

# Make src importable
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from crude_tanker_fv.loaders import INPUTS_DIR, load_company_inputs, load_watchlist
from crude_tanker_fv.pipeline import (
    _load_all_sectors,
    _maybe_apply_transactions,
    _run_scenarios_for_ticker,
)

# ----------------------------------------------------------------------------
# Weight definitions (METHODOLOGY §9.10 / WO4 Phase 0)
# ----------------------------------------------------------------------------
# All sets sum to 1.0 across the four active dry-bulk scenarios. Set A is the
# production lock (2026-06-09, §11.7.4); B / C are defensible ±~10pp brackets on
# the China-demand-tension axis — reweightings a reasonable analyst would hold,
# NOT strawmen. Production stays locked; these exist only to measure survival.
DRY_BULK_WEIGHT_SETS = {
    "Bulk Set A (locked 2026-06-09, FFA-calibrated prior)": {
        "china_acceleration": 0.20, "moderate_growth":     0.40,
        "china_property_drag": 0.25, "coordinated_slowdown": 0.15,
    },
    "Bulk Set B (China-bull / Simandou super-cycle bracket)": {
        # +10pp into acceleration, pulled from the two bear legs; base held.
        "china_acceleration": 0.30, "moderate_growth":     0.40,
        "china_property_drag": 0.18, "coordinated_slowdown": 0.12,
    },
    "Bulk Set C (China-property-drag bracket)": {
        # +10pp into property_drag / +5pp slowdown, pulled from accel + base.
        "china_acceleration": 0.12, "moderate_growth":     0.33,
        "china_property_drag": 0.35, "coordinated_slowdown": 0.20,
    },
}

SCEN_ORDER = ["china_acceleration", "moderate_growth",
              "china_property_drag", "coordinated_slowdown"]

for label, weights in DRY_BULK_WEIGHT_SETS.items():
    assert abs(sum(weights.values()) - 1.0) < 1e-9, f"{label} doesn't sum to 1"
    assert set(weights) == set(SCEN_ORDER), f"{label} scenario keys mismatch"

# The four VALUED dry-bulk names (inputs/watchlist.yaml). Census-only names
# (SHIP / DSX / EDRY / PANL / HSHP) are not onboarded and get no read here.
DRY_BULK_TICKERS = ["SBLK", "GNK", "CMDB", "SB"]


# ----------------------------------------------------------------------------
# Helpers (parallel to scripts/crude_weight_robustness.py)
# ----------------------------------------------------------------------------
def doc_with_weights(base_doc: dict, weights: dict) -> dict:
    d = copy.deepcopy(base_doc)
    for name, w in weights.items():
        if name in d["scenarios"]:
            d["scenarios"][name]["weight"] = w
    return d


def position_label(ev_pct: float) -> str:
    if ev_pct > 5.0:
        return "BUY"
    if ev_pct < -5.0:
        return "TRIM/SHORT"
    return "HOLD"


def run_ticker_under_all_sets(ticker: str, watchlist: dict, base_sector_docs: dict):
    """{set_label: {pw_fv, ev_pct, position}} via the production scenario path.

    The weight override touches only sector_docs["dry_bulk"]; every other
    sector doc passes through unchanged, isolating dry-bulk-weight sensitivity.
    Transaction-anchored marks are DEFAULT-ON in the pipeline (crude-script
    lesson) — apply them so the diagnostic values EV on the same marks the
    pipeline headline does."""
    entry = watchlist[ticker]
    price = entry["current_price"]
    target = entry["analyst_target"]
    ci = load_company_inputs(ticker, "2026-Q1")
    ci, _ = _maybe_apply_transactions(ci, INPUTS_DIR, True)

    results: dict[str, dict] = {}
    for set_label, weights in DRY_BULK_WEIGHT_SETS.items():
        modified_docs = dict(base_sector_docs)
        modified_docs["dry_bulk"] = doc_with_weights(base_sector_docs["dry_bulk"], weights)
        report, _, _ = _run_scenarios_for_ticker(
            ticker, ci, price, target, modified_docs, watchlist,
        )
        ev_pct = report.expected_value_vs_current / price * 100.0
        results[set_label] = {
            "pw_fv":    round(report.probability_weighted_fv, 2),
            "ev_pct":   round(ev_pct, 1),
            "position": position_label(ev_pct),
        }
    return {"ticker": ticker, "price": price, "target": target, "results": results}


def classify_robustness(per_ticker: dict) -> "tuple[str, list[str], str]":
    positions = [per_ticker["results"][s]["position"] for s in DRY_BULK_WEIGHT_SETS]
    if len(set(positions)) == 1:
        return "WEIGHT-ROBUST", positions, (
            f"position {positions[0]} across all {len(DRY_BULK_WEIGHT_SETS)} weight sets")
    by_position: dict[str, list[str]] = {}
    for s, pos in zip(DRY_BULK_WEIGHT_SETS, positions):
        by_position.setdefault(pos, []).append(s)
    parts = []
    for pos, sets in by_position.items():
        short = [s.split("(")[0].strip().replace("Bulk ", "") for s in sets]
        parts.append(f"{pos} under {'/'.join(short)}")
    return "WEIGHT-DRIVEN", positions, "; ".join(parts)


def set_short_defs() -> "list[str]":
    return [s.split("(")[0].strip().replace("Bulk ", "") for s in DRY_BULK_WEIGHT_SETS]


# ----------------------------------------------------------------------------
# Output rendering
# ----------------------------------------------------------------------------
def render_markdown(analyses: "list[dict]") -> str:
    lines: list[str] = []
    lines.append("# Dry-Bulk Weight-Robustness Diagnostic (§9.10 — WO4)\n")
    lines.append("Diagnostic (METHODOLOGY §9.10) — does NOT change the locked Bulk Set A "
                 "weights. Surfaces which dry-bulk calls survive a defensible reweighting "
                 "(**weight-robust**) vs which depend on a specific prior (**weight-driven**). "
                 "Unblocks the consumer's Gate E (`weight_sign_stable`).")
    lines.append("")
    lines.append("**Axis:** China dry-bulk demand tension — the four scenarios' own "
                 "parameter (china_acceleration ↔ china_property_drag / coordinated_slowdown) "
                 "and the charter thesis's load-bearing variable (Simandou ton-mile + supply "
                 "discipline vs China property/steel drag). Bulk Set B brackets the China-bull / "
                 "super-cycle case; Bulk Set C the property-drag case; both are ±~10pp shifts.")
    lines.append("")
    lines.append("**Naming namespace:** labels are DRY-BULK families (\"Bulk Set …\"); crude "
                 "and LNG both use \"Set B\" for their own — a bare unprefixed label would be a "
                 "methodology error.\n")

    lines.append("## Key findings (weight robustness, this run)\n")
    lines.append("Mark-spread robustness is the OTHER dimension — cross-read with "
                 "`outputs/broker_nav_sweep.md` before acting on any call.\n")
    lines.append("| Ticker | Weight robustness | What drives the call |")
    lines.append("|---|---|---|")
    for a in analyses:
        verdict, _, note = classify_robustness(a)
        badge = "✓ robust" if verdict == "WEIGHT-ROBUST" else "⚑ driven"
        lines.append(f"| {a['ticker']} | {badge} | {note} |")
    lines.append("")

    lines.append("## Weight sets compared\n")
    lines.append("| Scenario | " + " | ".join(set_short_defs()) + " |")
    lines.append("|---|" + "--:|" * len(DRY_BULK_WEIGHT_SETS))
    for scen in SCEN_ORDER:
        vals = [DRY_BULK_WEIGHT_SETS[s][scen] for s in DRY_BULK_WEIGHT_SETS]
        lines.append(f"| {scen} | " + " | ".join(f"{v:.2f}" for v in vals) + " |")
    lines.append("")

    lines.append("## Summary — per-name robustness\n")
    short = set_short_defs()
    lines.append("| Ticker | " + " | ".join(f"{s} EV" for s in short) +
                 " | Robustness | Notes |")
    lines.append("|---|" + "--:|" * len(short) + "---|---|")
    for a in analyses:
        r = a["results"]
        verdict, _, note = classify_robustness(a)
        cells = [f"{r[s]['ev_pct']:+.1f}% ({r[s]['position']})" for s in DRY_BULK_WEIGHT_SETS]
        badge = "✓ robust" if verdict == "WEIGHT-ROBUST" else "⚑ driven"
        lines.append(f"| {a['ticker']} | " + " | ".join(cells) + f" | {badge} | {note} |")
    lines.append("")

    lines.append("## Per-name detail\n")
    for a in analyses:
        lines.append(f"### {a['ticker']} — price ${a['price']:.2f}, target ${a['target']:.2f}\n")
        verdict, _, note = classify_robustness(a)
        lines.append(f"**Classification:** {verdict}. {note}.\n")
        lines.append("| Weight set | PW FV | EV % | Position |")
        lines.append("|---|--:|--:|---|")
        for s, r in a["results"].items():
            lines.append(f"| {s} | ${r['pw_fv']:.2f} | {r['ev_pct']:+.1f}% | {r['position']} |")
        lines.append("")

    lines.append("See METHODOLOGY §9.9 (mark robustness) and §9.10 (weight robustness). "
                 "This is the §9.10 output for the dry-bulk sector; crude/LNG analogues live "
                 "in `outputs/weight_robustness_diagnostic.md` / `outputs/lng_weight_robustness.md`.\n")
    return "\n".join(lines)


def write_xlsx(analyses: "list[dict]", path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Weight robustness (dry bulk)"
    headers = ["ticker", "price", "target"]
    for s in DRY_BULK_WEIGHT_SETS:
        sh = s.split("(")[0].strip()
        headers.extend([f"{sh} PW FV", f"{sh} EV %", f"{sh} position"])
    headers.extend(["robustness", "notes"])
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    driven_fill = PatternFill("solid", fgColor="FFE6E6")
    for a in analyses:
        verdict, _, note = classify_robustness(a)
        row = [a["ticker"], a["price"], a["target"]]
        for s in DRY_BULK_WEIGHT_SETS:
            r = a["results"][s]
            row.extend([r["pw_fv"], r["ev_pct"], r["position"]])
        row.extend([verdict, note])
        ws.append(row)
        if verdict == "WEIGHT-DRIVEN":
            for c in ws[ws.max_row]:
                c.fill = driven_fill
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 44)
    wb.save(path)


def sidecar_entries(analyses: "list[dict]") -> dict:
    """Per name: does the EV SIGN survive the whole weight family (the field
    the consumer's Gate E reads as weight_sign_stable)? Sign-stability, not
    label stability — a HOLD/BUY flip inside a positive range is magnitude
    sensitivity; a sign flip means the call's DIRECTION is a weight artifact."""
    out = {}
    for a in analyses:
        evs = [a["results"][s]["ev_pct"] for s in DRY_BULK_WEIGHT_SETS]
        out[a["ticker"]] = {
            "ev_min_pct": min(evs),
            "ev_max_pct": max(evs),
            "ev_sign_stable": (min(evs) > 0) == (max(evs) > 0) and 0 not in (min(evs), max(evs)),
            "positions": [a["results"][s]["position"] for s in DRY_BULK_WEIGHT_SETS],
        }
    return out


def main():
    watchlist = load_watchlist(live_prices=True)
    base_sector_docs = _load_all_sectors()

    analyses = []
    for ticker in DRY_BULK_TICKERS:
        if ticker not in watchlist:
            print(f"skip {ticker}: not in watchlist")
            continue
        analyses.append(run_ticker_under_all_sets(ticker, watchlist, base_sector_docs))

    root = Path(__file__).resolve().parents[1]
    out_md = root / "outputs" / "dry_bulk_weight_robustness.md"
    out_xlsx = root / "outputs" / "dry_bulk_weight_robustness.xlsx"
    out_md.write_text(render_markdown(analyses))
    write_xlsx(analyses, out_xlsx)

    from crude_tanker_fv.scorecard import update_weight_fragility_sidecar
    out_yaml = update_weight_fragility_sidecar(
        "dry_bulk", dict(DRY_BULK_WEIGHT_SETS), sidecar_entries(analyses))

    print(f"→ {out_md}")
    print(f"→ {out_xlsx}")
    print(f"→ {out_yaml}\n")
    print(f"{'Ticker':6s}  " + " ".join(f"{s:10s}" for s in set_short_defs()) + "  Robustness")
    print("-" * (10 + 11 * len(DRY_BULK_WEIGHT_SETS) + 14))
    for a in analyses:
        verdict, _, _ = classify_robustness(a)
        cells = [f"{a['results'][s]['ev_pct']:+5.1f}% {a['results'][s]['position'][:4]}"
                 for s in DRY_BULK_WEIGHT_SETS]
        print(f"{a['ticker']:6s}  " + " ".join(f"{c:10s}" for c in cells) + f"  {verdict}")


if __name__ == "__main__":
    main()
