"""Ad-hoc diagnostic: terminal-NAV-multiple sensitivity sweep (METHODOLOGY §9.2).

Open methodology decision #2 asks whether the dividend-strip terminal value at
q9 should stay at 1.0× NAV (current production), drop to 0.9× (mid-cycle
discount), or lift to 1.1× (structural undersupply). This script sweeps the
multiple across {0.85, 0.9, 1.0, 1.1, 1.15} for every watchlist ticker,
recording how single-point FV, scenario PW FV, EV%, and position move.

Mechanism: monkey-patch `crude_tanker_fv.dividend_strip.TERMINAL_NAV_MULTIPLE`
before each pipeline pass. Both the single-point path
(`pipeline.value_company`) and the scenario path
(`pipeline._run_scenarios_for_ticker`) re-import the module-level constant on
every `compute_dividend_strip` call, so the patch flows through both.

Does NOT mutate inputs/* or src/*. Outputs:
  - outputs/terminal_value_sensitivity.md  (per-name table + summary)
  - outputs/terminal_value_sensitivity.xlsx (single sheet, machine-readable)
"""

from __future__ import annotations

import sys
from pathlib import Path

# Make src importable
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from crude_tanker_fv import dividend_strip
from crude_tanker_fv.loaders import load_company_inputs, load_watchlist
from crude_tanker_fv.pipeline import (
    _load_all_sectors,
    _run_scenarios_for_ticker,
    value_company,
)

# ----------------------------------------------------------------------------
# Sweep configuration
# ----------------------------------------------------------------------------
MULTIPLES = [0.85, 0.90, 1.00, 1.10, 1.15]
BASELINE = 1.00
QUARTER = "2026-Q1"


# ----------------------------------------------------------------------------
# Per-ticker run
# ----------------------------------------------------------------------------
def run_under_multiple(ticker: str, watchlist: dict, sector_docs: dict, multiple: float) -> dict:
    """Patch TERMINAL_NAV_MULTIPLE and run the full per-ticker pipeline pass."""
    dividend_strip.TERMINAL_NAV_MULTIPLE = multiple

    entry = watchlist[ticker]
    price = entry["current_price"]
    target = entry["analyst_target"]

    # Single-point FV path
    report = value_company(ticker, QUARTER, price, target)

    # Scenario path (PW FV + position). For hybrids this aggregates sleeves.
    ci = load_company_inputs(ticker, QUARTER)
    headline, _crude_r, _product_r = _run_scenarios_for_ticker(
        ticker, ci, price, target, sector_docs, watchlist,
    )

    point_fv = report.blended.fair_value_per_share
    pw_fv = headline.probability_weighted_fv
    return {
        "multiple": multiple,
        "point_fv": point_fv,
        "point_ev_pct": (point_fv / price - 1.0) * 100.0,
        "pw_fv": pw_fv,
        "pw_ev_pct": (pw_fv / price - 1.0) * 100.0,
        "position": headline.position_recommendation,
        "nav_per_share": report.nav.nav_per_share,
        "strip_implied_price": report.strip.implied_price,
        "discounted_terminal": report.strip.discounted_terminal_value,
        "w_nav": report.blended.w_nav,
        "w_earn": report.blended.w_earn,
        "cycle_label": report.cycle.band_label,
        "cycle_position": report.cycle.cycle_position,
    }


def analyze_ticker(ticker: str, watchlist: dict, sector_docs: dict) -> dict:
    """Run all multiples for one ticker; return per-multiple metrics + summary."""
    entry = watchlist[ticker]
    runs = {m: run_under_multiple(ticker, watchlist, sector_docs, m) for m in MULTIPLES}
    baseline = runs[BASELINE]
    pos_set = {r["position"] for r in runs.values()}
    point_fv_range = max(r["point_fv"] for r in runs.values()) - min(
        r["point_fv"] for r in runs.values()
    )
    point_fv_pct_range = point_fv_range / baseline["point_fv"] * 100.0 if baseline["point_fv"] else 0.0
    pw_fv_range = max(r["pw_fv"] for r in runs.values()) - min(r["pw_fv"] for r in runs.values())
    pw_fv_pct_range = pw_fv_range / baseline["pw_fv"] * 100.0 if baseline["pw_fv"] else 0.0
    return {
        "ticker": ticker,
        "price": entry["current_price"],
        "target": entry["analyst_target"],
        "runs": runs,
        "baseline": baseline,
        "positions_seen": pos_set,
        "position_flips": len(pos_set) > 1,
        "point_fv_pct_range": point_fv_pct_range,
        "pw_fv_pct_range": pw_fv_pct_range,
    }


# ----------------------------------------------------------------------------
# Reporting
# ----------------------------------------------------------------------------
def _short_pos(label: str) -> str:
    """Collapse 'BUY (undervalued)' → 'BUY' for compact tables."""
    return label.split(" (")[0]


def render_markdown(analyses: list[dict]) -> str:
    out: list[str] = []
    out.append("# Terminal-NAV-Multiple Sensitivity Sweep — METHODOLOGY §9.2")
    out.append("")
    out.append(
        "**Open methodology decision #2.** The dividend strip terminates at "
        "q9 with a terminal value = `TERMINAL_NAV_MULTIPLE × NAV(aged 9q)`. "
        "Production multiple is **1.0×**. §9.2 flags two open priors: 0.9× "
        "(mid-cycle discount) and 1.1× (structural undersupply). This sweep "
        "tests {0.85, 0.9, 1.0, 1.1, 1.15} for every watchlist name."
    )
    out.append("")
    out.append(
        "**How to read it.** The strip terminal sits inside the *earnings* "
        "leg of the blend, so the FV impact scales with `w_earn`. Late-cycle "
        "names (w_earn = 0.30) absorb the smallest swing; below-mid-cycle "
        "names (w_earn = 0.60) the largest. **Position flips are what "
        "matter** — they identify calls that are sensitive to the §9.2 "
        "choice. A name with no flip across the full ±15% range is "
        "*multiple-robust* on its current position."
    )
    out.append("")

    # Per-name detail
    for a in analyses:
        t = a["ticker"]
        b = a["baseline"]
        out.append(f"## {t} — price ${a['price']:.2f}, baseline 1.0× FV ${b['point_fv']:.2f}")
        out.append("")
        out.append(
            f"_Cycle band: **{b['cycle_label']}** (w_nav = {b['w_nav']:.2f}, "
            f"w_earn = {b['w_earn']:.2f}). Discounted terminal at 1.0× = "
            f"${b['discounted_terminal']:.2f}/sh._"
        )
        out.append("")
        out.append(
            "| Multiple | Pt FV | Δ vs 1.0× | Pt EV% | PW FV | PW EV% | Position |"
        )
        out.append("|---:|--:|--:|--:|--:|--:|---|")
        for m in MULTIPLES:
            r = a["runs"][m]
            delta_pt = (r["point_fv"] / b["point_fv"] - 1.0) * 100.0
            flag = "  ←" if m == BASELINE else ""
            pos = _short_pos(r["position"])
            pos_flagged = f"**{pos}** ⚑" if pos != _short_pos(b["position"]) else pos
            out.append(
                f"| {m:.2f}×{flag} | ${r['point_fv']:.2f} | {delta_pt:+.1f}% | "
                f"{r['point_ev_pct']:+.1f}% | ${r['pw_fv']:.2f} | "
                f"{r['pw_ev_pct']:+.1f}% | {pos_flagged} |"
            )
        out.append("")
        if a["position_flips"]:
            seen = " / ".join(sorted({_short_pos(p) for p in a["positions_seen"]}))
            out.append(f"**⚑ Position flips across the sweep: {seen}.**")
            out.append("")

    # Summary
    out.append("---")
    out.append("")
    out.append("## Summary")
    out.append("")

    flippers = [a for a in analyses if a["position_flips"]]
    if flippers:
        out.append("**Position flips (multiple-driven calls):**")
        out.append("")
        out.append("| Ticker | Cycle | w_earn | Positions seen | Pt FV range | PW FV range |")
        out.append("|---|---|--:|---|--:|--:|")
        for a in flippers:
            b = a["baseline"]
            seen = " → ".join([_short_pos(a["runs"][m]["position"]) for m in MULTIPLES])
            out.append(
                f"| {a['ticker']} | {b['cycle_label']} | {b['w_earn']:.2f} | "
                f"{seen} | {a['point_fv_pct_range']:.1f}% | {a['pw_fv_pct_range']:.1f}% |"
            )
        out.append("")
    else:
        out.append("**No position flips** across the full {0.85×, …, 1.15×} sweep "
                   "for any watchlist name. Every call is multiple-robust at the "
                   "§9.2 priors.")
        out.append("")

    # Sensitivity rank
    out.append("**Sensitivity rank** (single-point FV % range across 0.85× → 1.15×, "
               "highest → lowest):")
    out.append("")
    out.append("| Ticker | Cycle | w_earn | Pt FV @ 0.85× | Pt FV @ 1.15× | % range |")
    out.append("|---|---|--:|--:|--:|--:|")
    ranked = sorted(analyses, key=lambda x: -x["point_fv_pct_range"])
    for a in ranked:
        b = a["baseline"]
        low = a["runs"][MULTIPLES[0]]["point_fv"]
        high = a["runs"][MULTIPLES[-1]]["point_fv"]
        out.append(
            f"| {a['ticker']} | {b['cycle_label']} | {b['w_earn']:.2f} | "
            f"${low:.2f} | ${high:.2f} | {a['point_fv_pct_range']:.1f}% |"
        )
    out.append("")

    # Interpretation
    out.append("## Interpretation")
    out.append("")
    out.append(
        "The single-point FV % range tracks `w_earn` as expected: names at "
        "peak (w_earn = 0.30) move ~3% across the full ±15% multiple range; "
        "names below mid-cycle (w_earn = 0.60) move ~6-8%. Any position flip "
        "is therefore a name that already sits close to the ±5% HOLD band "
        "and gets pushed across by the multiple alone. **For those names, "
        "the §9.2 choice is a material input to the call** and should be "
        "named explicitly in the decision log; for non-flippers, the §9.2 "
        "decision is informational only."
    )
    out.append("")
    out.append(
        "**This diagnostic does not pick a multiple.** It quantifies how "
        "much the open §9.2 decision matters. A resolution still needs a "
        "methodological prior — mid-cycle reversion (0.9×), undisturbed "
        "(1.0×), or structural undersupply (1.1×) — applied uniformly."
    )
    out.append("")
    out.append(
        "See METHODOLOGY §9 (open methodology decisions) and §3.2 (dividend "
        "strip / terminal value construction)."
    )
    out.append("")

    return "\n".join(out)


def write_xlsx(analyses: list[dict], path: Path) -> None:
    """Single-sheet xlsx — same content as the markdown tables, machine-readable."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Terminal-value sensitivity"

    headers = ["ticker", "price", "cycle_label", "w_nav", "w_earn"]
    for m in MULTIPLES:
        tag = f"{m:.2f}x"
        headers.extend([f"{tag} pt_fv", f"{tag} pt_ev_pct", f"{tag} pw_fv",
                        f"{tag} pw_ev_pct", f"{tag} position"])
    headers.extend(["pt_fv_pct_range", "pw_fv_pct_range", "position_flips"])
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    flip_fill = PatternFill("solid", fgColor="FFE6E6")

    for a in analyses:
        b = a["baseline"]
        row = [a["ticker"], a["price"], b["cycle_label"], b["w_nav"], b["w_earn"]]
        for m in MULTIPLES:
            r = a["runs"][m]
            row.extend([r["point_fv"], r["point_ev_pct"], r["pw_fv"],
                        r["pw_ev_pct"], _short_pos(r["position"])])
        row.extend([a["point_fv_pct_range"], a["pw_fv_pct_range"], a["position_flips"]])
        ws.append(row)
        if a["position_flips"]:
            for c in ws[ws.max_row]:
                c.fill = flip_fill

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 30)

    wb.save(path)


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main() -> None:
    watchlist = load_watchlist()
    sector_docs = _load_all_sectors()

    original_multiple = dividend_strip.TERMINAL_NAV_MULTIPLE
    try:
        analyses = [analyze_ticker(t, watchlist, sector_docs) for t in watchlist]
    finally:
        # Always restore — defensive even though the script exits afterward.
        dividend_strip.TERMINAL_NAV_MULTIPLE = original_multiple

    project_root = Path(__file__).resolve().parents[1]
    out_md = project_root / "outputs" / "terminal_value_sensitivity.md"
    out_xlsx = project_root / "outputs" / "terminal_value_sensitivity.xlsx"

    out_md.write_text(render_markdown(analyses))
    write_xlsx(analyses, out_xlsx)

    print(f"→ {out_md}")
    print(f"→ {out_xlsx}")
    print()
    flippers = [a["ticker"] for a in analyses if a["position_flips"]]
    if flippers:
        print(f"Position flips: {', '.join(flippers)}")
    else:
        print("No position flips across the sweep.")


if __name__ == "__main__":
    main()
